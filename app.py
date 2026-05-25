import os
import sys
from datetime import datetime, timedelta, date
from io import BytesIO
import calendar
from flask import Flask, render_template, redirect, url_for, request, flash, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, inspect, text

# Importar modelos
from models import db, Bombeiro, Viatura, Avaria, Escala, TrocaServico, Dispensa, Checklist, Fardamento, Disponibilidade, CreditoDispensa, Oficina, GestaoFrota, StockFardamento, Ecin, StockFarmacia, FarmaciaCentral, StockAmbulancia, ChecklistAmbulancia, CategoriaFarmacia, ChecklistAmbulanciaItem, Nota, MensagemCorreio, FardamentoAtribuido, Reuniao, NotaComando, Deslocacao, TipoFardaMaterial, Ferias, Mobilidade, Monitor, StockFardamentoArmazem, PontoAgua, QuadroOperacional
app = Flask(__name__)

# Chave secreta (obrigatória para sessões e formulários)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'chave-local-insegura')

# Desativar o rastreio de modificações (poupa memória)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Base de dados: usa PostgreSQL se a variável DATABASE_URL existir (Render),
# senão mantém o SQLite local para desenvolvimento.
basedir = os.path.abspath(os.path.dirname(__file__))
db_url = os.environ.get('DATABASE_URL', 'sqlite:///' + os.path.join(basedir, 'bombeiros.db'))
if db_url:
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql+psycopg://', 1)
    elif db_url.startswith('postgresql://'):
        db_url = db_url.replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url

# Inicializar a extensão com a aplicação
db.init_app(app)


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Adicionar após db.init_app(app) e antes das rotas
with app.app_context():
    from sqlalchemy import inspect, text

    inspector = inspect(db.engine)
    if 'quadro_operacional' in inspector.get_table_names():
        columns = [col['name'] for col in inspector.get_columns('quadro_operacional')]
        print("Colunas existentes:", columns)

        novas_colunas = [
            ('motorista_inem_id', 'INTEGER', 'bombeiros'),
            ('motorista_inem_numero', 'VARCHAR(20)', None),
            ('motorista_inem_mec', 'VARCHAR(20)', None),
            ('reserva_1_id', 'INTEGER', 'bombeiros'),
            ('reserva_1_numero', 'VARCHAR(20)', None),
            ('reserva_1_mec', 'VARCHAR(20)', None),
            ('reserva_2_id', 'INTEGER', 'bombeiros'),
            ('reserva_2_numero', 'VARCHAR(20)', None),
            ('reserva_2_mec', 'VARCHAR(20)', None)
        ]

        for col_name, col_type, ref_table in novas_colunas:
            if col_name not in columns:
                try:
                    sql = f'ALTER TABLE quadro_operacional ADD COLUMN {col_name} {col_type}'
                    if ref_table:
                        sql += f' REFERENCES {ref_table}(id)'
                    db.session.execute(text(sql))
                    print(f"Coluna {col_name} adicionada")
                except Exception as e:
                    print(f"Erro ao adicionar {col_name}: {e}")

        db.session.commit()
        print("Migração concluída!")
    else:
        db.create_all()
        print("Tabela quadro_operacional criada")

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Bombeiro, int(user_id))


# ---------- Criação da BD e admin inicial ----------
@app.before_request
def create_tables():
    db.create_all()
    if not Bombeiro.query.filter_by(email='admin@quartel.pt').first():
        admin = Bombeiro(
            numero_interno='B000',
            mecanografico='M000',
            nome='Administrador',
            email='admin@quartel.pt',
            password_hash=generate_password_hash('admin123'),
            posto='Comandante',
            ativo=True,
            tipo_user='Admin',
            tipo_bombeiro='Voluntário'
        )
        db.session.add(admin)
        if not Viatura.query.first():
            v1 = Viatura(matricula='12-AB-34', tipo='ABSC', nomenclatura='ABSC 01', marca='Mercedes', modelo='Atego', ano=2020)
            v2 = Viatura(matricula='56-CD-78', tipo='VLCI', nomenclatura='VLCI 02', marca='MAN', modelo='TGM', ano=2019)
            db.session.add_all([v1, v2])
        db.session.commit()

    # Criar utilizador "Sistema" para mensagens automáticas
    if not Bombeiro.query.filter((Bombeiro.mecanografico == 'SISTEMA') | (Bombeiro.numero_interno == 'B998')).first():
        sistema = Bombeiro(
            numero_interno='B998',  # número interno livre
            mecanografico='SISTEMA',
            nome='Sistema de Alertas',
            nomecompleto='Sistema de Alertas Automáticos',
            email='sistema@quartel.pt',
            password_hash=generate_password_hash('sistema'),
            posto='Sistema',
            tipo_bombeiro='Voluntário',
            ativo=True,
            tipo_user='User'
        )
        db.session.add(sistema)
        db.session.commit()


# ---------- Autenticação ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        mecanografico = request.form['mecanografico'].strip()
        password = request.form['password']
        user = Bombeiro.query.filter_by(mecanografico=mecanografico).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash('Sessão iniciada com sucesso.', 'success')
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Mecanográfico ou password inválidos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Sessão terminada.', 'info')
    return redirect(url_for('login'))


def turno_para_horas(turno_str):
    """Converte string do turno (ex: '1 - 00h/08h') em (hora_inicio, hora_fim)."""
    try:
        partes = turno_str.split('-')[1].strip().split('/')
        inicio = partes[0].replace('h', ':00')
        fim = partes[1].replace('h', ':00')
        return inicio, fim
    except Exception:
        return '08:00', '20:00'


from sqlalchemy import func   # Certifique-se de que está no topo do app.py

def gerar_html_dia(data, categoria_filtro=''):
    # Escalas do dia
    query = Escala.query.filter(
        func.date(Escala.data_inicio) <= data,
        func.date(Escala.data_fim) >= data
    )
    if categoria_filtro:
        query = query.filter(Escala.categoria == categoria_filtro)
    escalas = query.order_by(Escala.categoria, Escala.turno, Escala.data_inicio).all()

    # IDs dos bombeiros que aparecem nas escalas (filtrados ou não)
    ids_bombeiros_escalados = {e.bombeiro_id for e in escalas}

    # Agrupar por categoria
    escalas_por_categoria = {}
    for esc in escalas:
        cat = esc.categoria or 'Outros'
        if cat not in escalas_por_categoria:
            escalas_por_categoria[cat] = []
        escalas_por_categoria[cat].append(esc)

    cores_categorias = {
        'Motorista': '#fd7e14',
        'Socorrista': '#20c997',
        'Centralista': '#0d6efd',
        'EIP': '#6f42c1',
        'ECIN': '#dc3545',
        'ELAC': '#fd7e14',
        'Piquete': '#198754',
        'Bombeiro': '#0dcaf0'
    }

    # Dispensas do dia
    dispensas_do_dia = Dispensa.query.filter(
        Dispensa.data_inicio <= data,
        Dispensa.data_fim >= data,
        Dispensa.aprovada == True
    )
    if categoria_filtro:
        dispensas_do_dia = dispensas_do_dia.filter(Dispensa.bombeiro_id.in_(ids_bombeiros_escalados))
    dispensas_do_dia = dispensas_do_dia.order_by(Dispensa.data_inicio).all()
    bombeiros_com_dispensa = {d.bombeiro_id for d in dispensas_do_dia}

    # Trocas do dia
    trocas_do_dia = TrocaServico.query.filter(
        (TrocaServico.data_origem == data) | (TrocaServico.data_destino == data),
        TrocaServico.estado == 'aprovada'
    )
    if categoria_filtro:
        trocas_do_dia = trocas_do_dia.filter(
            (TrocaServico.bombeiro_origem_id.in_(ids_bombeiros_escalados)) |
            (TrocaServico.bombeiro_destino_id.in_(ids_bombeiros_escalados))
        )
    trocas_do_dia = trocas_do_dia.order_by(TrocaServico.data_origem).all()

    html = ''
    if escalas_por_categoria:
        html += '<h6 class="text-secondary mt-2">Escalas</h6>'
        for cat, lista in escalas_por_categoria.items():
            cor = cores_categorias.get(cat, '#6c757d')
            html += f'<h6 style="color: {cor};"><i class="bi bi-person-badge me-1"></i>{cat}</h6>'
            html += '<ul class="list-group list-group-flush mb-2">'
            for e in lista:
                estilo_linha = ''
                if e.bombeiro_id in bombeiros_com_dispensa:
                    estilo_linha = 'style="background-color: #ffe5cc;"'
                html += f'<li class="list-group-item py-1 d-flex justify-content-between" {estilo_linha}>'
                html += f'<span>{e.bombeiro.nome} ({e.bombeiro.mecanografico})</span>'
                html += f'<span>{e.turno}</span>'
                html += '</li>'
            html += '</ul>'
    else:
        html += '<p class="text-muted">Nenhuma escala para este dia nesta categoria.</p>'

    if trocas_do_dia:
        html += '<hr><h6 class="text-secondary">Trocas</h6><ul class="list-group list-group-flush mb-2">'
        for t in trocas_do_dia:
            if t.data_origem == data:
                texto = f'{t.bombeiro_origem.nome} estava escalado para este dia, mas trocou com {t.bombeiro_destino.nome}.'
            else:
                texto = f'{t.bombeiro_destino.nome} estava escalado para este dia, mas trocou com {t.bombeiro_origem.nome}.'
            html += f'<li class="list-group-item py-1">{texto}</li>'
        html += '</ul>'
    else:
        html += '<hr><p class="text-muted">Sem trocas neste dia.</p>'

    if dispensas_do_dia:
        html += '<hr><h6 class="text-secondary">Dispensas</h6><ul class="list-group list-group-flush mb-2">'
        for d in dispensas_do_dia:
            html += f'<li class="list-group-item py-1">{d.bombeiro.nome} ({d.data_inicio.strftime("%d/%m/%Y")} a {d.data_fim.strftime("%d/%m/%Y")})</li>'
        html += '</ul>'
    else:
        html += '<hr><p class="text-muted">Sem dispensas neste dia.</p>'

    return html


# ---------- Dashboard ----------
@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html')


# ---------- Gestão de Bombeiros ----------
@app.route('/bombeiros', methods=['GET', 'POST'])
@login_required
def gerir_bombeiros():
    if current_user.tipo_user != 'Admin':
        flash('Acesso restrito ao administrador.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        numero = request.form['numero_interno']
        mecanografico = request.form['mecanografico']
        nome = request.form['nome']
        nomecompleto = request.form.get('nomecompleto', '')
        email = request.form['email']
        password = request.form['password']
        posto = request.form['posto']
        tipo_bombeiro = request.form.get('tipo_bombeiro', 'Voluntário')
        tipo_user = request.form['tipo_user']
        telemovel = request.form.get('telemovel', '')
        resp_departamento = request.form.get('resp_departamento', '')

        if Bombeiro.query.filter((Bombeiro.email == email) |
                                 (Bombeiro.numero_interno == numero) |
                                 (Bombeiro.mecanografico == mecanografico)).first():
            flash('Email, número interno ou mecanográfico já existe.', 'warning')
            return redirect(url_for('gerir_bombeiros'))

        if telemovel and Bombeiro.query.filter_by(telemovel=telemovel).first():
            flash('Telemóvel já está associado a outro bombeiro.', 'warning')
            return redirect(url_for('gerir_bombeiros'))

        novo = Bombeiro(
            numero_interno=numero,
            mecanografico=mecanografico,
            nome=nome,
            nomecompleto=nomecompleto if nomecompleto else None,
            email=email,
            password_hash=generate_password_hash(password),
            posto=posto,
            tipo_bombeiro=tipo_bombeiro,
            tipo_user=tipo_user,
            telemovel=telemovel if telemovel else None,
            resp_departamento=resp_departamento if resp_departamento else None,
            ativo=True
        )
        db.session.add(novo)
        db.session.commit()
        flash('Bombeiro criado com sucesso!', 'success')
        return redirect(url_for('gerir_bombeiros'))

    # GET – listagem com filtro e ordenação
    pesquisa = request.args.get('pesquisa', '').strip()
    query = Bombeiro.query
    if pesquisa:
        query = query.filter(
            (Bombeiro.nome.ilike(f'%{pesquisa}%')) |
            (Bombeiro.mecanografico.ilike(f'%{pesquisa}%'))
        )
    bombeiros = query.order_by(Bombeiro.nome.asc()).all()
    return render_template('bombeiros.html', bombeiros=bombeiros, pesquisa=pesquisa)

@app.route('/bombeiros/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_bombeiro(id):
    if current_user.tipo_user != 'Admin':
        flash('Acesso restrito ao administrador.', 'danger')
        return redirect(url_for('dashboard'))

    bombeiro = Bombeiro.query.get_or_404(id)

    if request.method == 'POST':
        novo_email = request.form['email']
        novo_numero = request.form['numero_interno']
        novo_mec = request.form['mecanografico']
        novo_telemovel = request.form.get('telemovel', '')

        conflito = Bombeiro.query.filter(
            (Bombeiro.id != id) &
            ((Bombeiro.email == novo_email) |
             (Bombeiro.numero_interno == novo_numero) |
             (Bombeiro.mecanografico == novo_mec))
        ).first()
        if conflito:
            flash('Email, nº interno ou mecanográfico já em uso.', 'warning')
            return redirect(url_for('gerir_bombeiros'))

        if novo_telemovel:
            telemovel_conflito = Bombeiro.query.filter(
                Bombeiro.id != id, Bombeiro.telemovel == novo_telemovel
            ).first()
            if telemovel_conflito:
                flash('Telemóvel já associado a outro bombeiro.', 'warning')
                return redirect(url_for('gerir_bombeiros'))

        bombeiro.numero_interno = novo_numero
        bombeiro.mecanografico = novo_mec
        bombeiro.nome = request.form['nome']
        bombeiro.nomecompleto = request.form.get('nomecompleto', '')
        bombeiro.email = novo_email
        bombeiro.telemovel = novo_telemovel if novo_telemovel else None
        bombeiro.posto = request.form['posto']
        bombeiro.tipo_bombeiro = request.form.get('tipo_bombeiro', 'Voluntário')
        bombeiro.tipo_user = request.form['tipo_user']
        bombeiro.resp_departamento = request.form.get('resp_departamento', '')
        bombeiro.ativo = request.form.get('ativo') == 'on'

        nova_password = request.form.get('password')
        if nova_password:
            bombeiro.password_hash = generate_password_hash(nova_password)

        db.session.commit()
        flash('Bombeiro atualizado com sucesso!', 'success')
        return redirect(url_for('gerir_bombeiros'))

    return redirect(url_for('gerir_bombeiros'))


@app.route('/bombeiros/apagar/<int:id>')
@login_required
def apagar_bombeiro(id):
    if current_user.tipo_user != 'Admin':
        flash('Acesso restrito ao administrador.', 'danger')
        return redirect(url_for('dashboard'))

    bombeiro = Bombeiro.query.get_or_404(id)
    if bombeiro.id == current_user.id:
        flash('Não pode apagar o seu próprio utilizador.', 'warning')
        return redirect(url_for('gerir_bombeiros'))

    db.session.delete(bombeiro)
    db.session.commit()
    flash(f'Bombeiro {bombeiro.nome} removido.', 'info')
    return redirect(url_for('gerir_bombeiros'))

# ---------- Exportar Bombeiros ----------

@app.route('/bombeiros/exportar')
@login_required
def exportar_bombeiros():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    bombeiros = Bombeiro.query.order_by(Bombeiro.numero_interno).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bombeiros"

    cabecalhos = ['Nº Interno', 'Mecanográfico', 'Nome', 'Nome Completo', 'Email', 'Telemóvel',
                  'Posto', 'Tipo Bombeiro', 'Resp. Departamento', 'Tipo Utilizador', 'Ativo']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for b in bombeiros:
        ws.append([
            b.numero_interno,
            b.mecanografico,
            b.nome,
            b.nomecompleto or '',
            b.email,
            b.telemovel or '',
            b.posto,
            b.tipo_bombeiro,
            b.resp_departamento or '',
            b.tipo_user,
            'Sim' if b.ativo else 'Não'
        ])

    col_widths = [12, 16, 25, 35, 30, 15, 20, 16, 22, 16, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='bombeiros.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- Importar Bombeiros ----------
@app.route('/bombeiros/importar', methods=['POST'])
@login_required
def importar_bombeiros():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('gerir_bombeiros'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('gerir_bombeiros'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido. Use .xlsx.', 'danger')
        return redirect(url_for('gerir_bombeiros'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('gerir_bombeiros'))

    linhas_importadas = 0
    erros = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        try:
            numero = str(row[0]).strip() if row[0] else None
            mecanografico = str(row[1]).strip() if len(row) > 1 and row[1] else None
            nome = str(row[2]).strip() if len(row) > 2 and row[2] else None
            nomecompleto = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            email = str(row[4]).strip().lower() if len(row) > 4 and row[4] else None
            telemovel = str(row[5]).strip() if len(row) > 5 and row[5] else None
            posto = str(row[6]).strip() if len(row) > 6 and row[6] else 'Bombeiro'
            tipo_bombeiro = str(row[7]).strip() if len(row) > 7 and row[7] else 'Voluntário'
            departamento = str(row[8]).strip() if len(row) > 8 and row[8] else None
            tipo_user = str(row[9]).strip() if len(row) > 9 and row[9] else 'User'
            ativo_str = str(row[10]).strip().lower() if len(row) > 10 and row[10] else 'sim'
            ativo = ativo_str == 'sim'
        except Exception:
            erros.append(f'Linha {row_num}: dados inválidos.')
            continue

        if not numero or not mecanografico or not nome or not email:
            erros.append(f'Linha {row_num}: campos obrigatórios em falta.')
            continue

        existente = Bombeiro.query.filter(
            (Bombeiro.numero_interno == numero) |
            (Bombeiro.mecanografico == mecanografico) |
            (Bombeiro.email == email)
        ).first()
        if existente:
            erros.append(f'Linha {row_num}: já existe um bombeiro com o Nº Interno, Mecanográfico ou Email indicado.')
            continue

        novo = Bombeiro(
            numero_interno=numero,
            mecanografico=mecanografico,
            nome=nome,
            nomecompleto=nomecompleto if nomecompleto else None,
            email=email,
            password_hash=generate_password_hash('123456'),
            telemovel=telemovel if telemovel else None,
            posto=posto,
            tipo_bombeiro=tipo_bombeiro,
            resp_departamento=departamento,
            tipo_user=tipo_user,
            ativo=ativo
        )
        db.session.add(novo)
        linhas_importadas += 1

    db.session.commit()

    if erros:
        flash(f'{linhas_importadas} importados. {len(erros)} erro(s): ' + '; '.join(erros), 'warning')
    else:
        flash(f'{linhas_importadas} bombeiros importados com sucesso!', 'success')
    return redirect(url_for('gerir_bombeiros'))


# ---------- Viaturas ----------
@app.route('/viaturas')
@login_required
def listar_viaturas():
    # Voluntários NÃO podem aceder à listagem de viaturas
    if current_user.tipo_bombeiro == 'Voluntário' and current_user.tipo_user != 'Admin':
        flash(
            'Acesso restrito. Apenas bombeiros profissionais, administração, comando, oficina e central podem aceder à listagem de viaturas.',
            'danger')
        return redirect(url_for('dashboard'))

    # Profissionais, Admin, Comando, Oficina, Central têm acesso
    if (current_user.tipo_bombeiro != 'Profissional' and
            current_user.tipo_user != 'Admin' and
            current_user.resp_departamento not in ['Comando', 'Oficina', 'Central']):
        flash(
            'Acesso restrito. Apenas bombeiros profissionais, administração, comando, oficina e central podem aceder a esta área.',
            'danger')
        return redirect(url_for('dashboard'))

    viaturas = Viatura.query.order_by(Viatura.tipo, Viatura.matricula).all()

    total_operacionais = Viatura.query.filter(func.lower(Viatura.estado) == 'operacional').count()
    total_inoperacionais = Viatura.query.filter(func.lower(Viatura.estado) == 'inoperacional').count()
    total_manutencao = Viatura.query.filter(func.lower(Viatura.estado) == 'manutenção').count()
    total_geral = Viatura.query.count()

    return render_template('viaturas.html',
                           viaturas=viaturas,
                           total_operacionais=total_operacionais,
                           total_inoperacionais=total_inoperacionais,
                           total_manutencao=total_manutencao,
                           total_geral=total_geral)


@app.route('/viaturas/adicionar', methods=['POST'])
@login_required
def adicionar_viatura():
    if current_user.tipo_user != 'Admin':
        flash('Apenas administradores podem adicionar viaturas.', 'danger')
        return redirect(url_for('listar_viaturas'))

    matricula = request.form['matricula']
    tipo = request.form['tipo']
    nomenclatura = request.form['nomenclatura']
    marca = request.form['marca']
    modelo = request.form['modelo']
    ano = request.form['ano']

    nova = Viatura(matricula=matricula, tipo=tipo, nomenclatura=nomenclatura,
                   marca=marca, modelo=modelo, ano=ano)
    db.session.add(nova)
    db.session.commit()

    gestao = GestaoFrota(viatura_id=nova.id)
    db.session.add(gestao)
    db.session.commit()

    flash('Viatura adicionada.', 'success')
    return redirect(url_for('listar_viaturas'))

@app.route('/viaturas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_viatura(id):
    if current_user.tipo_user != 'Admin':
        flash('Apenas administradores podem editar viaturas.', 'danger')
        return redirect(url_for('listar_viaturas'))

    viatura = Viatura.query.get_or_404(id)

    if request.method == 'POST':
        matricula = request.form['matricula']
        existente = Viatura.query.filter(Viatura.matricula == matricula, Viatura.id != id).first()
        if existente:
            flash('Matrícula já existe.', 'warning')
            return redirect(url_for('listar_viaturas'))

        viatura.matricula = matricula
        viatura.tipo = request.form['tipo']
        viatura.nomenclatura = request.form['nomenclatura']
        viatura.marca = request.form['marca']
        viatura.modelo = request.form['modelo']
        viatura.ano = request.form['ano']
        viatura.estado = request.form['estado']
        db.session.commit()
        flash('Viatura atualizada.', 'success')
        return redirect(url_for('listar_viaturas'))

    return redirect(url_for('listar_viaturas'))


@app.route('/viaturas/apagar/<int:id>')
@login_required
def apagar_viatura(id):
    if current_user.tipo_user != 'Admin':
        flash('Apenas administradores podem apagar viaturas.', 'danger')
        return redirect(url_for('listar_viaturas'))

    viatura = Viatura.query.get_or_404(id)
    db.session.delete(viatura)
    db.session.commit()
    flash(f'Viatura {viatura.matricula} removida.', 'info')
    return redirect(url_for('listar_viaturas'))


# ---------- Exportar Viaturas ----------
@app.route('/viaturas/exportar')
@login_required
def exportar_viaturas():
    viaturas = Viatura.query.order_by(Viatura.matricula).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Viaturas"

    cabecalhos = ['Matrícula', 'Tipo', 'Nomenclatura', 'Marca', 'Modelo', 'Ano', 'Estado']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for v in viaturas:
        ws.append([v.matricula, v.tipo, v.nomenclatura, v.marca, v.modelo, v.ano, v.estado])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='viaturas.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------- Importar Viaturas ----------

@app.route('/viaturas/importar', methods=['POST'])
@login_required
def importar_viaturas():
    if current_user.tipo_user != 'Admin':
        flash('Acesso restrito ao administrador.', 'danger')
        return redirect(url_for('listar_viaturas'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('listar_viaturas'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('listar_viaturas'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('listar_viaturas'))

    linhas_importadas = 0
    erros = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue
        try:
            matricula = str(row[0]).strip() if row[0] else None
            tipo = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            nomenclatura = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            marca = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            modelo = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            ano = int(row[5]) if len(row) > 5 and row[5] else None
            estado = str(row[6]).strip().lower() if len(row) > 6 and row[6] else 'operacional'
        except Exception:
            erros.append(f'Linha {row_num}: dados inválidos.')
            continue

        if not matricula:
            erros.append(f'Linha {row_num}: matrícula obrigatória.')
            continue
        if Viatura.query.filter_by(matricula=matricula).first():
            erros.append(f'Linha {row_num}: matrícula {matricula} já existe.')
            continue

        nova = Viatura(matricula=matricula, tipo=tipo, nomenclatura=nomenclatura,
                       marca=marca, modelo=modelo, ano=ano, estado=estado)
        db.session.add(nova)
        # Criar também registo na Gestão de Frota
        gestao = GestaoFrota(viatura_id=nova.id)
        db.session.add(gestao)
        linhas_importadas += 1

    db.session.commit()
    if erros:
        flash(f'{linhas_importadas} importadas. {len(erros)} erro(s): ' + '; '.join(erros), 'warning')
    else:
        flash(f'{linhas_importadas} viaturas importadas com sucesso.', 'success')
    return redirect(url_for('listar_viaturas'))


# ---------- Gestão Frota ----------
@app.route('/gestao-frota')
@login_required
def gestao_frota():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    # Criar registos de gestão em falta para qualquer viatura que ainda não tenha um
    viaturas_sem_gestao = Viatura.query.filter(~Viatura.gestao_frota.has()).all()
    for v in viaturas_sem_gestao:
        nova_gestao = GestaoFrota(viatura_id=v.id)
        db.session.add(nova_gestao)
    if viaturas_sem_gestao:
        db.session.commit()

    # Consulta que junta todas as viaturas com os seus dados de gestão (left join)
    registos = db.session.query(Viatura, GestaoFrota)\
                         .outerjoin(GestaoFrota, Viatura.id == GestaoFrota.viatura_id)\
                         .order_by(Viatura.matricula).all()

    viaturas = Viatura.query.order_by(Viatura.matricula).all()
    return render_template('gestao_frota.html', registos=registos, viaturas=viaturas)


# ---------- Editar Gestão Frota ----------
@app.route('/gestao-frota/editar/<int:id>', methods=['POST'])
@login_required
def editar_gestao_frota(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('gestao_frota'))

    registo = GestaoFrota.query.get_or_404(id)

    # Campos de data
    inspecao_str = request.form.get('inspecao_periodica')
    registo.inspecao_periodica = datetime.strptime(inspecao_str, '%Y-%m-%d').date() if inspecao_str else None

    # Campos numéricos (convertidos para int, None se vazios)
    registo.kms_ultima_revisao = request.form.get('kms_ultima_revisao', type=int)
    registo.kms_proxima_revisao = request.form.get('kms_proxima_revisao', type=int)
    registo.kms_pneus_dianteiros = request.form.get('kms_pneus_dianteiros', type=int)
    registo.kms_pneus_trazeiros = request.form.get('kms_pneus_trazeiros', type=int)
    registo.kms_correia = request.form.get('kms_correia', type=int)
    registo.outros_apontamentos = request.form.get('outros_apontamentos', '')

    db.session.commit()
    v = registo.viatura  # ← obter a viatura associada
    if v:
        verificar_inspecao_proxima(v)

    flash('Registo atualizado.', 'success')
    return redirect(url_for('gestao_frota'))

# ---------- Exportar Gestão Frota ----------

@app.route('/gestao-frota/exportar')
@login_required
def exportar_gestao_frota():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('gestao_frota'))

    registos = GestaoFrota.query.join(Viatura).order_by(Viatura.matricula).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gestao Frota"

    cabecalhos = ['ID', 'Matrícula', 'Nomenclatura', 'Marca/Modelo', 'Ano',
                  'Inspeção Periódica', 'Kms Última Revisão', 'Kms Próxima Revisão',
                  'Kms Pneus Diant.', 'Kms Pneus Tras.', 'Kms Correia', 'Outros Apontamentos']
    ws.append(cabecalhos)

    # Estilo cabeçalho
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for r in registos:
        v = r.viatura
        ws.append([
            r.id,
            v.matricula,
            v.nomenclatura,
            f"{v.marca} {v.modelo}",
            v.ano,
            r.inspecao_periodica.strftime('%d/%m/%Y') if r.inspecao_periodica else '',
            r.kms_ultima_revisao or '',
            r.kms_proxima_revisao or '',
            r.kms_pneus_dianteiros or '',
            r.kms_pneus_trazeiros or '',
            r.kms_correia or '',
            r.outros_apontamentos or ''
        ])

    # Ajustar larguras
    col_widths = [5, 12, 18, 20, 6, 14, 14, 14, 12, 12, 12, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='gestao_frota.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def verificar_inspecao_proxima(viatura):
    if not viatura.gestao_frota or not viatura.gestao_frota.inspecao_periodica:
        return
    hoje = date.today()
    dias = (viatura.gestao_frota.inspecao_periodica - hoje).days
    if 0 < dias <= 30:
        # Destinatários: Comando e Oficina
        destinatarios = Bombeiro.query.filter(
            (Bombeiro.resp_departamento == 'Comando') |
            (func.lower(Bombeiro.resp_departamento) == 'oficina'),
            Bombeiro.ativo == True
        ).all()

        remetente = Bombeiro.query.filter_by(mecanografico='SISTEMA').first()
        if not remetente:
            return

        corpo = (
            f"⚠️ Alerta de Inspeção Periódica\n\n"
            f"A viatura {viatura.matricula} ({viatura.nomenclatura}) "
            f"tem a inspeção agendada para {viatura.gestao_frota.inspecao_periodica.strftime('%d/%m/%Y')}.\n"
            f"Faltam {dias} dias.\n\n"
            f"Por favor, tome as devidas providências."
        )

        for dest in destinatarios:
            msg = MensagemCorreio(
                remetente_id=remetente.id,
                destinatario_id=dest.id,
                assunto=f'⚠️ Inspeção próxima – {viatura.matricula}',
                corpo=corpo,
                data_envio=datetime.utcnow(),
                lida=False,
                apagada_remetente=False,
                apagada_destinatario=False
            )
            db.session.add(msg)
        db.session.commit()

# ---------- Importar Gestão Frota ----------

@app.route('/gestao-frota/importar', methods=['POST'])
@login_required
def importar_gestao_frota():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('gestao_frota'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('gestao_frota'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('gestao_frota'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('gestao_frota'))

    linhas_importadas = 0
    erros = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue
        try:
            matricula = str(row[1]).strip() if len(row) > 1 and row[1] else None
            if not matricula:
                erros.append(f'Linha {row_num}: matrícula em falta.')
                continue
            viatura = Viatura.query.filter_by(matricula=matricula).first()
            if not viatura:
                erros.append(f'Linha {row_num}: viatura {matricula} não encontrada.')
                continue

            gestao = GestaoFrota.query.filter_by(viatura_id=viatura.id).first()
            if not gestao:
                # Cria se não existir
                gestao = GestaoFrota(viatura_id=viatura.id)
                db.session.add(gestao)

            # Atualizar campos
            # Inspeção periódica (coluna 6, índice 5)
            inspecao_str = str(row[5]).strip() if len(row) > 5 and row[5] else None
            if inspecao_str:
                gestao.inspecao_periodica = _parse_data(inspecao_str)
            else:
                gestao.inspecao_periodica = None

            gestao.kms_ultima_revisao = int(row[6]) if len(row) > 6 and row[6] else None
            gestao.kms_proxima_revisao = int(row[7]) if len(row) > 7 and row[7] else None
            gestao.kms_pneus_dianteiros = int(row[8]) if len(row) > 8 and row[8] else None
            gestao.kms_pneus_trazeiros = int(row[9]) if len(row) > 9 and row[9] else None
            gestao.kms_correia = int(row[10]) if len(row) > 10 and row[10] else None
            gestao.outros_apontamentos = str(row[11]).strip() if len(row) > 11 and row[11] else None

            linhas_importadas += 1
        except Exception as e:
            erros.append(f'Linha {row_num}: erro {str(e)}')
            continue

    db.session.commit()
    if erros:
        flash(f'{linhas_importadas} importadas com {len(erros)} erro(s).', 'warning')
    else:
        flash(f'{linhas_importadas} registos importados com sucesso.', 'success')
    return redirect(url_for('gestao_frota'))

# ---------- Imprimir Gestão Frota ----------
@app.route('/gestao-frota/imprimir/<int:viatura_id>')
@login_required
def imprimir_gestao_frota(viatura_id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('gestao_frota'))

    viatura = Viatura.query.get_or_404(viatura_id)
    gestao = GestaoFrota.query.filter_by(viatura_id=viatura_id).first()
    return render_template('imprimir_frota.html', viatura=viatura, gestao=gestao)


# ---------- Avarias ----------
from datetime import date

@app.route('/avarias', methods=['GET', 'POST'])
@login_required
def avarias():
    tab = request.args.get('tab', 'registo')

    if request.method == 'POST':
        viatura_id = request.form['viatura_id']
        descricao = request.form['descricao']
        kms = request.form.get('kms', '')

        ultima_avaria = Avaria.query.order_by(Avaria.id.desc()).first()
        proximo = 1
        if ultima_avaria:
            try:
                proximo = int(ultima_avaria.codigo[2:]) + 1
            except Exception:
                pass
        codigo = f"AV{proximo:04d}"

        nova = Avaria(
            codigo=codigo,
            viatura_id=viatura_id,
            descricao=descricao,
            reportado_por=current_user.id,
            kms=int(kms) if kms else None,
            responsavel_oficina=False,
            comando_verificado=False
        )
        db.session.add(nova)
        db.session.commit()
        flash(f'Avaria {codigo} registada.', 'success')
        return redirect(url_for('avarias', tab='registo'))

    # ---- GET ----
    filtro_viatura_id = request.args.get('viatura_id', type=int)
    filtro_mes = request.args.get('mes', type=int)
    filtro_ano = request.args.get('ano', type=int)
    filtro_estado = request.args.get('estado', '')

    todas_viaturas = Viatura.query.order_by(Viatura.matricula).all()

    # --- Aba Registo ---
    query_registo = Avaria.query
    # Voluntários veem apenas as suas avarias
    if current_user.tipo_bombeiro == 'Voluntário' and current_user.tipo_user != 'Admin':
        query_registo = query_registo.filter(Avaria.reportado_por == current_user.id)
    avarias_lista = query_registo.order_by(Avaria.data_reporte.desc()).limit(100).all()

    viaturas = Viatura.query.all()

    # --- Aba Histórico ---
    historico_avarias = []
    if tab == 'historico':
        query_hist = Avaria.query
        if filtro_viatura_id:
            query_hist = query_hist.filter_by(viatura_id=filtro_viatura_id)
        if filtro_mes:
            query_hist = query_hist.filter(db.extract('month', Avaria.data_reporte) == filtro_mes)
        if filtro_ano:
            query_hist = query_hist.filter(db.extract('year', Avaria.data_reporte) == filtro_ano)
        if filtro_estado:
            query_hist = query_hist.filter_by(estado=filtro_estado)
        # Voluntários veem apenas as suas avarias no histórico
        if current_user.tipo_bombeiro == 'Voluntário' and current_user.tipo_user != 'Admin':
            query_hist = query_hist.filter(Avaria.reportado_por == current_user.id)
        historico_avarias = query_hist.order_by(Avaria.data_reporte.desc()).all()

    return render_template('avarias.html',
                           avarias=avarias_lista,
                           viaturas=viaturas,
                           todas_viaturas=todas_viaturas,
                           historico_avarias=historico_avarias,
                           filtro_viatura_id=filtro_viatura_id,
                           filtro_mes=filtro_mes,
                           filtro_ano=filtro_ano,
                           filtro_estado=filtro_estado,
                           tab=tab,
                           hoje=date.today())


@app.route('/avarias/parecer_resp/<int:id>', methods=['POST'])
@login_required
def avaria_parecer_resp(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Oficina':
        return jsonify({'erro': 'Acesso restrito'}), 403
    avaria = Avaria.query.get_or_404(id)
    data = request.get_json()
    avaria.parecer_resp = data.get('parecer')
    avaria.urg_despacho = data.get('urg_despacho')
    avaria.data_resp = date.today()
    # Estado da avaria passa para "Analisar"
    avaria.estado = 'Analisar'
    # Atualizar estado da viatura conforme escolha do responsável
    if avaria.viatura:
        if data.get('estado_viatura') == 'inoperacional':
            avaria.viatura.estado = 'inoperacional'
        else:
            avaria.viatura.estado = 'operacional'
    db.session.commit()
    return jsonify({'sucesso': True, 'data_resp': avaria.data_resp.strftime('%d/%m/%Y')})


@app.route('/avarias/decisao_cmd/<int:id>', methods=['POST'])
@login_required
def avaria_decisao_cmd(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        return jsonify({'erro': 'Acesso restrito'}), 403
    avaria = Avaria.query.get_or_404(id)
    data = request.get_json()
    avaria.decisao_cmd = data.get('decisao')
    avaria.data_cmd = date.today()
    estado_viatura = data.get('estado_viatura')  # 'operacional', 'inoperacional' ou 'oficina'

    if avaria.viatura:
        if estado_viatura == 'oficina' or estado_viatura == 'inoperacional':
            avaria.viatura.estado = 'inoperacional'
            avaria.estado = 'Oficina'
        else:  # operacional
            avaria.viatura.estado = 'operacional'
            avaria.estado = 'Resolvido'

    db.session.commit()
    return jsonify({'sucesso': True, 'data_cmd': avaria.data_cmd.strftime('%d/%m/%Y')})


@app.route('/avarias/atualizar/<int:id>', methods=['POST'])
@login_required
def atualizar_avaria(id):
    avaria = Avaria.query.get_or_404(id)

    permitido = (current_user.tipo_user == 'Admin' or
                 current_user.resp_departamento in ['Oficina', 'Comando'])
    if not permitido:
        flash('Sem permissão para alterar.', 'danger')
        return redirect(url_for('avarias'))

    novo_estado = request.form.get('estado')
    if novo_estado in ['Pendente', 'Analisar', 'Resolvido']:
        avaria.estado = novo_estado

    kms = request.form.get('kms', '')
    if kms:
        avaria.kms = int(kms)

    if current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Oficina':
        avaria.responsavel_oficina = (request.form.get('resp_oficina') == 'on')

    if current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Comando':
        avaria.comando_verificado = (request.form.get('comando_verificado') == 'on')

    db.session.commit()
    flash('Avaria atualizada.', 'success')
    return redirect(url_for('avarias'))


@app.route('/avarias/historico/<int:viatura_id>')
@login_required
def historico_avaria(viatura_id):
    viatura = Viatura.query.get_or_404(viatura_id)
    avarias_lista = Avaria.query.filter_by(viatura_id=viatura_id)\
                                .order_by(Avaria.data_reporte.desc()).all()
    return render_template('_historico_avarias.html', viatura=viatura, avarias=avarias_lista)


# ---------- Exportar Avarias ----------
@app.route('/avarias/exportar')
@login_required
def exportar_avarias():
    avarias_lista = Avaria.query.order_by(Avaria.data_reporte.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avarias"

    cabecalhos = ['Código', 'Viatura', 'Descrição', 'Reportado por', 'Kms',
                  'Resp. Oficina', 'Comando', 'Estado', 'Data Reporte']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for a in avarias_lista:
        ws.append([
            a.codigo,
            a.viatura.matricula,
            a.descricao,
            a.reportador.nome,
            a.kms or '',
            'Sim' if a.responsavel_oficina else 'Não',
            'Sim' if a.comando_verificado else 'Não',
            a.estado,
            a.data_reporte.strftime('%d/%m/%Y %H:%M') if a.data_reporte else ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='avarias.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- Importar Avarias ----------

@app.route('/avarias/importar', methods=['POST'])
@login_required
def importar_avarias():
    if current_user.tipo_user != 'Admin':
        flash('Acesso restrito ao administrador.', 'danger')
        return redirect(url_for('avarias'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('avarias'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('avarias'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('avarias'))

    linhas_importadas = 0
    erros = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue
        try:
            codigo = str(row[0]).strip() if row[0] else None
            viatura_matricula = str(row[1]).strip() if len(row) > 1 and row[1] else None
            descricao = str(row[2]).strip() if len(row) > 2 and row[2] else None
            reportador_mec = str(row[3]).strip() if len(row) > 3 and row[3] else None
            kms = int(row[4]) if len(row) > 4 and row[4] else None
            resp_oficina = str(row[5]).strip().lower() == 'sim' if len(row) > 5 and row[5] else False
            comando = str(row[6]).strip().lower() == 'sim' if len(row) > 6 and row[6] else False
            estado = str(row[7]).strip() if len(row) > 7 and row[7] else 'Pendente'
            data_reporte = _parse_data(row[8]) if len(row) > 8 and row[8] else datetime.utcnow()
        except Exception:
            erros.append(f'Linha {row_num}: dados inválidos.')
            continue

        if not descricao or not viatura_matricula:
            erros.append(f'Linha {row_num}: descrição e matrícula obrigatórias.')
            continue
        viatura = Viatura.query.filter_by(matricula=viatura_matricula).first()
        if not viatura:
            erros.append(f'Linha {row_num}: viatura {viatura_matricula} não encontrada.')
            continue
        bombeiro = Bombeiro.query.filter_by(mecanografico=reportador_mec).first() if reportador_mec else None
        if not bombeiro and reportador_mec:
            erros.append(f'Linha {row_num}: mecanográfico {reportador_mec} não encontrado.')
            continue

        # Gerar código único se necessário
        if not codigo:
            ultima = Avaria.query.order_by(Avaria.id.desc()).first()
            proximo = 1 if not ultima else int(ultima.codigo[2:]) + 1
            codigo = f"AV{proximo:04d}"

        nova = Avaria(
            codigo=codigo,
            viatura_id=viatura.id,
            descricao=descricao,
            reportado_por=bombeiro.id if bombeiro else current_user.id,
            kms=kms,
            responsavel_oficina=resp_oficina,
            comando_verificado=comando,
            estado=estado,
            data_reporte=data_reporte if data_reporte else datetime.utcnow()
        )
        db.session.add(nova)
        linhas_importadas += 1

    db.session.commit()
    if erros:
        flash(f'{linhas_importadas} importadas. {len(erros)} erro(s): ' + '; '.join(erros), 'warning')
    else:
        flash(f'{linhas_importadas} avarias importadas com sucesso.', 'success')
    return redirect(url_for('avarias'))

# ---------- Oficina ----------
# ------------------------------
# Rota principal da Oficina (listagem, criação, filtros)
# ------------------------------
@app.route('/oficina', methods=['GET', 'POST'])
@login_required
def oficina():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    tab = request.args.get('tab', 'registo')

    # ----- Criação de novo registo (POST) -----
    if request.method == 'POST':
        # Gerar código único
        ultimo = Oficina.query.order_by(Oficina.id.desc()).first()
        proximo = 1 if not ultimo else int(ultimo.codigo[2:]) + 1
        codigo = f"OF{proximo:04d}"

        avaria_id = request.form.get('avaria_id', type=int)
        viatura_id = request.form.get('viatura_id', type=int)
        kms = request.form.get('kms', type=int)

        # Se veio de uma avaria, usar os dados dela
        if avaria_id:
            avaria = Avaria.query.get(avaria_id)
            if avaria:
                viatura_id = avaria.viatura_id
                kms = avaria.kms

        nome_oficina = request.form.get('nome_oficina')
        data_recepcao = datetime.strptime(request.form['data_recepcao'], '%Y-%m-%d').date()
        motivo = request.form.get('motivo', '')
        descricao_oficina = request.form.get('descricao_oficina', '')
        n_orc_fat = request.form.get('n_orc_fat', '')
        data_entrega_str = request.form.get('data_entrega')
        data_entrega = datetime.strptime(data_entrega_str, '%Y-%m-%d').date() if data_entrega_str else None

        inoperacional = request.form.get('inoperacional') == 'on'
        chefe_oficina = request.form.get('chefe_oficina') == 'on'
        comando = request.form.get('comando') == 'on'
        operacional = request.form.get('operacional') == 'on'

        # Determinar estado
        if operacional or (chefe_oficina and comando):
            estado = 'Resolvido'
        else:
            estado = 'Oficina'

        novo = Oficina(
            codigo=codigo,
            nome_oficina=nome_oficina,
            data_recepcao=data_recepcao,
            motivo=motivo,
            avaria_id=avaria_id,
            viatura_id=viatura_id,
            kms=kms,
            inoperacional=inoperacional,
            descricao_oficina=descricao_oficina,
            n_orc_fat=n_orc_fat,
            data_entrega=data_entrega,
            chefe_oficina=chefe_oficina,
            comando=comando,
            operacional=operacional,
            estado=estado
        )
        db.session.add(novo)

        # Atualizar estado da viatura
        viatura = Viatura.query.get(viatura_id)
        if viatura:
            if inoperacional:
                viatura.estado = 'Inoperacional'
            elif estado == 'Resolvido':
                viatura.estado = 'operacional'

        # Sincronizar avaria (se existir e estado for Resolvido)
        if estado == 'Resolvido' and avaria_id:
            avaria = Avaria.query.get(avaria_id)
            if avaria and avaria.estado != 'Resolvido':
                avaria.estado = 'Resolvido'
                db.session.add(avaria)

        db.session.commit()
        flash(f'Registo de oficina {codigo} criado.', 'success')
        return redirect(url_for('oficina', tab='registo'))

    # ----- GET (listagens) -----
    # Avarias com estado 'Oficina' (para dropdown)
    avarias_oficina = Avaria.query.filter_by(estado='Oficina').all()

    # Todos os registos de oficina (aba Registo)
    registos = Oficina.query.order_by(Oficina.id.desc()).all()

    # Listas para filtros (aba Histórico)
    filtro_viatura_id = request.args.get('viatura_id', type=int)
    filtro_nome_oficina = request.args.get('nome_oficina', '')
    filtro_mes = request.args.get('mes', type=int)
    filtro_ano = request.args.get('ano', type=int)
    filtro_estado = request.args.get('estado', '')

    todas_viaturas = Viatura.query.order_by(Viatura.matricula).all()
    nomes_oficina = [row[0] for row in db.session.query(Oficina.nome_oficina).distinct().all()]

    # Histórico filtrado
    historico_oficina = []
    if tab == 'historico':
        query = Oficina.query
        if filtro_viatura_id:
            query = query.filter_by(viatura_id=filtro_viatura_id)
        if filtro_nome_oficina:
            query = query.filter_by(nome_oficina=filtro_nome_oficina)
        if filtro_mes:
            query = query.filter(db.extract('month', Oficina.data_recepcao) == filtro_mes)
        if filtro_ano:
            query = query.filter(db.extract('year', Oficina.data_recepcao) == filtro_ano)
        if filtro_estado:
            query = query.filter_by(estado=filtro_estado)
        historico_oficina = query.order_by(Oficina.data_registo.desc()).all()

    # Viaturas para o formulário (fallback)
    viaturas = Viatura.query.all()

    return render_template('oficina.html',
                           registos=registos,
                           avarias_oficina=avarias_oficina,
                           viaturas=viaturas,
                           todas_viaturas=todas_viaturas,
                           nomes_oficina=nomes_oficina,
                           historico_oficina=historico_oficina,
                           filtro_viatura_id=filtro_viatura_id,
                           filtro_nome_oficina=filtro_nome_oficina,
                           filtro_mes=filtro_mes,
                           filtro_ano=filtro_ano,
                           filtro_estado=filtro_estado,
                           tab=tab)


# ------------------------------
# Editar registo de Oficina
# ------------------------------
@app.route('/oficina/editar/<int:id>', methods=['POST'])
@login_required
def editar_registo_oficina(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('oficina'))

    registo = Oficina.query.get_or_404(id)

    # Recolher dados do formulário
    registo.nome_oficina = request.form['nome_oficina']
    registo.data_recepcao = datetime.strptime(request.form['data_recepcao'], '%Y-%m-%d').date()
    registo.motivo = request.form.get('motivo', '')
    registo.descricao_oficina = request.form.get('descricao_oficina', '')
    registo.n_orc_fat = request.form.get('n_orc_fat', '')
    data_entrega_str = request.form.get('data_entrega')
    registo.data_entrega = datetime.strptime(data_entrega_str, '%Y-%m-%d').date() if data_entrega_str else None

    # Tratamento da avaria e viatura
    avaria_id_str = request.form.get('avaria_id')
    if avaria_id_str:
        avaria_id = int(avaria_id_str)
        avaria = Avaria.query.get(avaria_id)
        if avaria:
            registo.avaria_id = avaria_id
            registo.viatura_id = avaria.viatura_id
            registo.kms = avaria.kms
    else:
        registo.avaria_id = None
        viatura_id = request.form.get('viatura_id', type=int)
        if viatura_id:
            registo.viatura_id = viatura_id
        kms = request.form.get('kms', type=int)
        if kms is not None:
            registo.kms = kms

    # Checkboxes
    registo.inoperacional = request.form.get('inoperacional') == 'on'
    registo.chefe_oficina = request.form.get('chefe_oficina') == 'on'
    registo.comando = request.form.get('comando') == 'on'
    registo.operacional = request.form.get('operacional') == 'on'

    # Determinar estado
    if registo.operacional or (registo.chefe_oficina and registo.comando):
        registo.estado = 'Resolvido'
        registo.operacional = True
    else:
        registo.estado = 'Oficina'

    # Actualizar estado da viatura
    viatura = Viatura.query.get(registo.viatura_id)
    if viatura:
        if registo.inoperacional:
            viatura.estado = 'Inoperacional'
        elif registo.estado == 'Resolvido':
            viatura.estado = 'operacional'

    # ⭐ Sincronizar avaria (se o estado for Resolvido e existir avaria)
    if registo.estado == 'Resolvido' and registo.avaria_id:
        avaria_associada = Avaria.query.get(registo.avaria_id)
        if avaria_associada and avaria_associada.estado != 'Resolvido':
            avaria_associada.estado = 'Resolvido'
            db.session.add(avaria_associada)

    db.session.commit()
    flash(f'Registo {registo.codigo} atualizado e avaria sincronizada.', 'success')
    return redirect(url_for('oficina'))


# ------------------------------
# Apagar registo de Oficina
# ------------------------------
@app.route('/oficina/apagar/<int:id>')
@login_required
def apagar_registo_oficina(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('oficina'))

    registo = Oficina.query.get_or_404(id)
    db.session.delete(registo)
    db.session.commit()
    flash(f'Registo {registo.codigo} removido.', 'info')
    return redirect(url_for('oficina'))



# ---------- Histórico Oficina ----------
@app.route('/oficina/historico/<int:viatura_id>')
@login_required
def historico_oficina(viatura_id):
    viatura = Viatura.query.get_or_404(viatura_id)
    registos = Oficina.query.filter_by(viatura_id=viatura_id)\
                            .order_by(Oficina.data_registo.desc()).all()
    return render_template('_historico_oficina.html', viatura=viatura, registos=registos)

@app.route('/oficina/imprimir/<int:id>')
@login_required
def imprimir_registo_oficina(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('oficina'))
    registo = Oficina.query.get_or_404(id)
    return render_template('imprimir_oficina.html', registo=registo)

# ---------- Exportar Oficina ----------
@app.route('/oficina/exportar')
@login_required
def exportar_oficina():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('oficina'))

    registos = Oficina.query.order_by(Oficina.data_registo.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oficina"

    cabecalhos = ['Código', 'Data Registo', 'Nome Oficina', 'Data Recepção', 'Motivo',
                  'Nº Avaria', 'Viatura', 'Kms', 'Inoperacional', 'Descrição Oficina',
                  'Nº Orç/Fat', 'Data Entrega', 'Chefe Oficina', 'Comando', 'Operacional', 'Estado']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for r in registos:
        ws.append([
            r.codigo,
            r.data_registo.strftime('%d/%m/%Y %H:%M') if r.data_registo else '',
            r.nome_oficina,
            r.data_recepcao.strftime('%d/%m/%Y') if r.data_recepcao else '',
            r.motivo or '',
            r.avaria.codigo if r.avaria else '',
            r.viatura.matricula if r.viatura else '',
            r.kms or '',
            'Sim' if r.inoperacional else 'Não',
            r.descricao_oficina or '',
            r.n_orc_fat or '',
            r.data_entrega.strftime('%d/%m/%Y') if r.data_entrega else '',
            'Sim' if r.chefe_oficina else 'Não',
            'Sim' if r.comando else 'Não',
            'Sim' if r.operacional else 'Não',
            r.estado
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='oficina.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- Importar Oficina ----------
@app.route('/oficina/importar', methods=['POST'])
@login_required
def importar_oficina():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Oficina']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('oficina'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('oficina'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '':
        flash('Ficheiro vazio.', 'warning')
        return redirect(url_for('oficina'))
    if not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('oficina'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('oficina'))

    linhas_importadas = 0
    erros = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue
        try:
            nome_oficina = str(row[0]).strip() if row[0] else None
            data_recepcao = _parse_data(row[1]) if len(row) > 1 else None
            motivo = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            avaria_codigo = str(row[3]).strip() if len(row) > 3 and row[3] else None
            viatura_matricula = str(row[4]).strip() if len(row) > 4 and row[4] else None
            kms = int(row[5]) if len(row) > 5 and row[5] else 0
            inoperacional = str(row[6]).strip().lower() == 'sim' if len(row) > 6 and row[6] else False
            descricao = str(row[7]).strip() if len(row) > 7 and row[7] else ''
            n_orc_fat = str(row[8]).strip() if len(row) > 8 and row[8] else ''
            data_entrega = _parse_data(row[9]) if len(row) > 9 and row[9] else None
            chefe_oficina = str(row[10]).strip().lower() == 'sim' if len(row) > 10 and row[10] else False
            comando = str(row[11]).strip().lower() == 'sim' if len(row) > 11 and row[11] else False
            operacional = str(row[12]).strip().lower() == 'sim' if len(row) > 12 and row[12] else False
        except Exception as e:
            erros.append(f'Linha {row_num}: erro ao interpretar dados.')
            continue

        if not nome_oficina or not data_recepcao:
            erros.append(f'Linha {row_num}: nome oficina e data recepção obrigatórios.')
            continue

        viatura = Viatura.query.filter_by(matricula=viatura_matricula).first() if viatura_matricula else None
        avaria = Avaria.query.filter_by(codigo=avaria_codigo).first() if avaria_codigo else None

        if not viatura:
            erros.append(f'Linha {row_num}: viatura não encontrada.')
            continue

        # Gerar código único
        ultimo = Oficina.query.order_by(Oficina.id.desc()).first()
        proximo = 1 if not ultimo else int(ultimo.codigo[2:]) + 1
        codigo = f"OF{proximo:04d}"

        nova = Oficina(
            codigo=codigo,
            nome_oficina=nome_oficina,
            data_recepcao=data_recepcao,
            motivo=motivo,
            avaria_id=avaria.id if avaria else None,
            viatura_id=viatura.id,
            kms=kms,
            inoperacional=inoperacional,
            descricao_oficina=descricao,
            n_orc_fat=n_orc_fat,
            data_entrega=data_entrega,
            chefe_oficina=chefe_oficina,
            comando=comando,
            operacional=operacional,
            estado='Resolvido' if (chefe_oficina and comando) else 'Oficina'
        )
        db.session.add(nova)
        linhas_importadas += 1

    db.session.commit()
    if erros:
        flash(f'{linhas_importadas} importadas com {len(erros)} erro(s).', 'warning')
    else:
        flash(f'{linhas_importadas} importadas com sucesso.', 'success')
    return redirect(url_for('oficina'))

#---------------------Férias-----------------

@app.route('/ferias', methods=['GET', 'POST'])
@login_required
def ferias():
    if request.method == 'POST':
        datas_str = request.form.get('datas', '')  # datas separadas por vírgula
        if not datas_str:
            flash('Selecione pelo menos um dia.', 'warning')
            return redirect(url_for('ferias'))
        datas = [d.strip() for d in datas_str.split(',') if d.strip()]
        datas.sort()
        data_inicio = datetime.strptime(datas[0], '%Y-%m-%d').date()
        data_fim = datetime.strptime(datas[-1], '%Y-%m-%d').date()

        # Verificar sobreposição (opcional)
        existente = Ferias.query.filter(
            Ferias.bombeiro_id == current_user.id,
            Ferias.estado != 'Rejeitado',
            Ferias.data_inicio <= data_fim,
            Ferias.data_fim >= data_inicio
        ).first()
        if existente:
            flash('Já tem um pedido de férias que cobre este período.', 'warning')
            return redirect(url_for('ferias'))

        nova = Ferias(
            bombeiro_id=current_user.id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            estado='Pendente'
        )
        db.session.add(nova)
        db.session.commit()
        flash('Pedido de férias enviado.', 'success')
        return redirect(url_for('ferias'))

    # GET – filtros
    ano = request.args.get('ano', type=int, default=date.today().year)
    mes = request.args.get('mes', type=int)                     # ← NOVO
    bombeiro_id = request.args.get('bombeiro_id', type=int)
    estado_filtro = request.args.get('estado', '')

    query = Ferias.query
    if ano:
        query = query.filter(db.extract('year', Ferias.data_inicio) == ano)
    if mes:                                                     # ← NOVO
        query = query.filter(db.extract('month', Ferias.data_inicio) == mes)
    if bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)
    if estado_filtro:
        query = query.filter_by(estado=estado_filtro)

    # Se for user normal, só vê os seus pedidos
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria']:
        query = query.filter_by(bombeiro_id=current_user.id)

    pedidos = query.order_by(Ferias.data_inicio.desc()).all()
    bombeiros = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()

    return render_template('ferias.html', pedidos=pedidos, bombeiros=bombeiros,
                           ano=ano, mes=mes, bombeiro_id=bombeiro_id, estado_filtro=estado_filtro)



@app.route('/ferias/aprovar/<int:id>')
@login_required
def aprovar_ferias(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('ferias'))

    f = Ferias.query.get_or_404(id)
    f.estado = 'Aprovado'
    f.aprovado_por = current_user.id

    # Criar registos na escala para cada dia de férias
    dia_atual = f.data_inicio
    while dia_atual <= f.data_fim:
        nova_escala = Escala(
            bombeiro_id=f.bombeiro_id,
            data_inicio=datetime.combine(dia_atual, datetime.min.time()),
            data_fim=datetime.combine(dia_atual, datetime.max.time()),
            turno='Férias',
            categoria='Férias',
            funcao='Férias Aprovadas'
        )
        db.session.add(nova_escala)
        dia_atual += timedelta(days=1)

    db.session.commit()
    flash('Férias aprovadas e registadas na escala.', 'success')
    return redirect(url_for('ferias'))



@app.route('/ferias/rejeitar/<int:id>')
@login_required
def rejeitar_ferias(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('ferias'))
    ferias = Ferias.query.get_or_404(id)
    ferias.estado = 'Rejeitado'
    ferias.aprovado_por = current_user.id
    db.session.commit()
    flash('Férias rejeitadas.', 'info')
    return redirect(url_for('ferias'))

@app.route('/ferias/apagar/<int:id>')
@login_required
def apagar_ferias(id):
    ferias = Ferias.query.get_or_404(id)
    if current_user.id != ferias.bombeiro_id and current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('ferias'))
    db.session.delete(ferias)
    db.session.commit()
    flash('Pedido removido.', 'info')
    return redirect(url_for('ferias'))


def agrupar_ferias(lista_ferias):
    """Recebe uma lista de objetos Ferias e retorna uma lista de intervalos (data_inicio, data_fim) agrupados."""
    if not lista_ferias:
        return []
    dias = []
    for f in lista_ferias:
        d = f.data_inicio
        while d <= f.data_fim:
            dias.append(d)
            d += timedelta(days=1)
    dias = sorted(set(dias))
    intervalos = []
    inicio = dias[0]
    fim = dias[0]
    for d in dias[1:]:
        if d == fim + timedelta(days=1):
            fim = d
        else:
            intervalos.append((inicio, fim))
            inicio = d
            fim = d
    intervalos.append((inicio, fim))
    return intervalos



from datetime import date
from sqlalchemy import func

@app.route('/escala')
@login_required
def escala():
    # Filtros
    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)
    categoria = request.args.get('categoria', '')
    mecanografico = request.args.get('mecanografico', '')
    turno_filtro = request.args.get('turno', '')
    dia_filtro = request.args.get('dia', type=int)

    query = Escala.query.join(Bombeiro)

    if mes:
        query = query.filter(db.extract('month', Escala.data_inicio) == mes)
    if ano:
        query = query.filter(db.extract('year', Escala.data_inicio) == ano)
    if categoria:
        query = query.filter(Escala.categoria == categoria)
    if mecanografico:
        query = query.filter(Bombeiro.mecanografico == mecanografico)
    if turno_filtro:
        query = query.filter(Escala.turno == turno_filtro)
    if dia_filtro:
        data_ref = date(ano, mes, dia_filtro)
        query = query.filter(
            func.date(Escala.data_inicio) <= data_ref,
            func.date(Escala.data_fim) >= data_ref
        )

    # Permissões: utilizador normal vê apenas as suas escalas
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        query = query.filter(Escala.bombeiro_id == current_user.id)

    # Ordenação base
    escalas = query.order_by(Escala.data_inicio.asc()).all()

    # ---------- ORDENAÇÃO PERSONALIZADA (mantida) ----------
    categorias_ordem = ['Motorista', 'Socorrista', 'Centralista', 'EIP',
                        'ECIN', 'ELAC', 'Piquete']
    prioridades = {
        'Motorista': ['Luís Matias', 'Jorge Pereira', 'José Soldado', 'José Seco', 'Pedro Fernandes',
                      'David Charrinho', 'Fábio Leirinha', 'Soeiro Mendes', 'Ana Marzia',
                      'Filipe Martins', 'Eric Nobre'],
        'Socorrista': ['José Rodrigues', 'Paulo Branquinho', 'Sabrina Fernandes'],
        'Centralista': ['Mariana Charrinho', 'Ruben Ramos', 'António Pequito'],
        'EIP': ['José Fernandes', 'João Mateus', 'Tiago Bizarro', 'João Carita', 'João Silva']
    }
    for cat in prioridades:
        prioridades[cat] = {nome: i for i, nome in enumerate(prioridades[cat])}

    def chave_ordenacao(esc):
        data_ini = esc.data_inicio.date() if hasattr(esc.data_inicio, 'date') else esc.data_inicio
        cat = esc.categoria if esc.categoria else 'Outros'
        ordem_cat = categorias_ordem.index(cat) if cat in categorias_ordem else len(categorias_ordem)
        nome = esc.bombeiro.nome.strip()
        if cat in prioridades:
            pos_nome = prioridades[cat].get(nome, len(prioridades[cat]))
            return (data_ini, ordem_cat, pos_nome, nome)
        else:
            return (data_ini, ordem_cat, esc.turno, nome)

    escalas = sorted(escalas, key=chave_ordenacao)

    # Cartões de resumo
    total_escalas = len(escalas)
    total_bombeiros = len(set(e.bombeiro_id for e in escalas))
    total_categorias = len(set(e.categoria for e in escalas if e.categoria))
    total_turnos = len(set(e.turno for e in escalas))

    # Atividade do mês para o mini-calendário
    dias_com_escalas = []
    if mes:
        for dia in range(1, 32):
            try:
                data_ref = date(ano, mes, dia)
            except ValueError:
                break
            tem = Escala.query.filter(
                db.extract('month', Escala.data_inicio) == mes,
                db.extract('year', Escala.data_inicio) == ano,
                func.date(Escala.data_inicio) <= data_ref,
                func.date(Escala.data_fim) >= data_ref
            ).first() is not None
            if tem:
                dias_com_escalas.append(dia)

    # Trocas e dispensas de hoje
    hoje = date.today()
    trocas_hoje = set()
    for t in TrocaServico.query.filter(
        (TrocaServico.data_origem == hoje) | (TrocaServico.data_destino == hoje),
        TrocaServico.estado == 'aprovada'
    ).all():
        trocas_hoje.add(t.bombeiro_origem_id)
        trocas_hoje.add(t.bombeiro_destino_id)

    dispensas_hoje = set(d.bombeiro_id for d in Dispensa.query.filter(
        Dispensa.data_inicio <= hoje,
        Dispensa.data_fim >= hoje,
        Dispensa.aprovada == True
    ).all())

    # ---------- NOVO: Verificar se utilizador tem escalas ECIN/ELAC no mês/ano seleccionado ----------
    tem_ecin_no_mes = False
    tem_elac_no_mes = False
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        # Verifica na tabela Escala (não na Ecin, pois a escala é que define a categoria)
        tem_ecin_no_mes = Escala.query.filter(
            Escala.bombeiro_id == current_user.id,
            db.extract('year', Escala.data_inicio) == ano,
            db.extract('month', Escala.data_inicio) == mes,
            Escala.categoria == 'ECIN'
        ).first() is not None
        tem_elac_no_mes = Escala.query.filter(
            Escala.bombeiro_id == current_user.id,
            db.extract('year', Escala.data_inicio) == ano,
            db.extract('month', Escala.data_inicio) == mes,
            Escala.categoria == 'ELAC'
        ).first() is not None
    else:
        # Utilizadores com permissão podem sempre ver
        tem_ecin_no_mes = True
        tem_elac_no_mes = True

    # Listas auxiliares
    meses = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    categorias = ['Motorista','Socorrista','Centralista','EIP','ECIN','ELAC','Piquete','Férias']
    turnos = ['1 - 00h/08h','2 - 08h/16h','3 - 16h/24h','4 - 11h/19h','5 - 10h/18h','6 - 07h/19h','7 - 19h/07h','8 - 08h/20h','9 - 20h/08h']
    bombeiros_ativos = Bombeiro.query.filter_by(ativo=True).all()
    mecanograficos_ativos = [b.mecanografico for b in bombeiros_ativos]

    # Férias do profissional (mantido)
    ferias_intervalos = []
    if current_user.tipo_bombeiro == 'Profissional':
        mes_ref = mes or date.today().month
        ano_ref = date.today().year
        ferias_mes = Ferias.query.filter(
            Ferias.bombeiro_id == current_user.id,
            Ferias.estado.in_(['Pendente', 'Aprovado']),
            db.extract('month', Ferias.data_fim) >= mes_ref,
            db.extract('month', Ferias.data_inicio) <= mes_ref,
            db.extract('year', Ferias.data_inicio) == ano_ref,
            db.extract('year', Ferias.data_fim) == ano_ref
        ).order_by(Ferias.data_inicio).all()
        ferias_intervalos = agrupar_ferias(ferias_mes)

    if categoria == 'Férias':
        ferias_do_mes = Ferias.query.filter(
            Ferias.estado == 'Aprovado',
            db.extract('month', Ferias.data_fim) >= mes,
            db.extract('month', Ferias.data_inicio) <= mes,
            db.extract('year', Ferias.data_inicio) == ano,
            db.extract('year', Ferias.data_fim) == ano
        ).all()
        ids_ferias = [f.bombeiro_id for f in ferias_do_mes]
        query = query.filter(Escala.bombeiro_id.in_(ids_ferias))

    return render_template('escala.html',
                           escalas=escalas,
                           meses=meses,
                           categorias=categorias,
                           turnos=turnos,
                           mes_atual=mes,
                           categoria_atual=categoria,
                           turno_atual_filtro=turno_filtro,
                           dia_filtro=dia_filtro,
                           mecanografico_atual=mecanografico,
                           bombeiros_ativos=bombeiros_ativos,
                           mecanograficos_ativos=mecanograficos_ativos,
                           total_escalas=total_escalas,
                           total_bombeiros=total_bombeiros,
                           total_categorias=total_categorias,
                           total_turnos=total_turnos,
                           dias_com_escalas=dias_com_escalas,
                           trocas_hoje=trocas_hoje,
                           dispensas_hoje=dispensas_hoje,
                           hoje=hoje,
                           tem_ecin=tem_ecin_no_mes,     # ← variável renomeada para clareza
                           tem_elac=tem_elac_no_mes,     # ← variável renomeada
                           ferias_intervalos=ferias_intervalos,
                           now=date.today())



@app.route('/escala/imprimir-mes')
@login_required
def imprimir_escala_mes():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    if not mes or not ano:
        hoje = date.today()
        mes = hoje.month
        ano = hoje.year

    # Feriados fixos
    feriados = []
    feriados_fixos = {
        (1, 1): "Ano Novo", (4, 25): "Dia da Liberdade", (5, 1): "Dia do Trabalhador",
        (6, 10): "Dia de Portugal", (8, 15): "Assunção de Nossa Senhora",
        (10, 5): "Implantação da República", (11, 1): "Todos os Santos",
        (12, 1): "Restauração da Independência", (12, 8): "Imaculada Conceição",
        (12, 25): "Natal"
    }
    for (m, d), nome in feriados_fixos.items():
        try:
            feriados.append(date(ano, m, d))
        except ValueError:
            pass

    # Escalas do mês (excluindo ECIN, ELAC, Piquete)
    escalas = Escala.query.join(Bombeiro).filter(
        db.extract('month', Escala.data_inicio) == mes,
        db.extract('year', Escala.data_inicio) == ano,
        ~Escala.categoria.in_(['ECIN', 'ELAC', 'Piquete'])
    ).order_by(Escala.data_inicio.asc()).all()

    # Ordenação personalizada
    categorias_ordem = ['Motorista', 'Socorrista', 'Centralista', 'EIP']
    prioridades = {
        'Motorista': ['Luís Matias','Jorge Pereira','José Soldado','José Seco','Pedro Fernandes',
                      'David Charrinho','Fábio Leirinha','Soeiro Mendes','Ana Marzia',
                      'Filipe Martins','Eric Nobre'],
        'Socorrista': ['José Rodrigues','Paulo Branquinho','Sabrina Fernandes'],
        'Centralista': ['Mariana Charrinho','Ruben Ramos','António Pequito'],
        'EIP': ['José Fernandes','João Mateus','Tiago Bizarro','João Carita','João Silva']
    }
    for cat in prioridades:
        prioridades[cat] = {nome: i for i, nome in enumerate(prioridades[cat])}

    def chave_ordenacao(esc):
        cat = esc.categoria if esc.categoria else 'Outros'
        ordem_cat = categorias_ordem.index(cat) if cat in categorias_ordem else len(categorias_ordem)
        nome = esc.bombeiro.nome.strip()
        if cat in prioridades:
            pos_nome = prioridades[cat].get(nome, len(prioridades[cat]))
            return (ordem_cat, pos_nome, nome)
        else:
            return (ordem_cat, esc.turno, nome)

    escalas = sorted(escalas, key=chave_ordenacao)

    # Estrutura para o template
    from collections import OrderedDict
    estrutura = OrderedDict()
    for esc in escalas:
        bombeiro = esc.bombeiro
        cat = esc.categoria if esc.categoria else 'Outros'
        chave_bombeiro = (bombeiro.id, bombeiro.nome, bombeiro.mecanografico, cat)
        if chave_bombeiro not in estrutura:
            estrutura[chave_bombeiro] = {}
        try:
            num_turno = int(esc.turno.split('-')[0].strip())
        except (ValueError, IndexError):
            num_turno = 1
        dia = esc.data_inicio.day
        if dia not in estrutura[chave_bombeiro]:
            estrutura[chave_bombeiro][dia] = num_turno

    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dias = list(range(1, ultimo_dia + 1))

    # ---------- Férias aprovadas ----------
    ferias_query = Ferias.query.filter(
        Ferias.estado == 'Aprovado',
        db.extract('month', Ferias.data_fim) >= mes,
        db.extract('month', Ferias.data_inicio) <= mes,
        db.extract('year', Ferias.data_inicio) == ano,
        db.extract('year', Ferias.data_fim) == ano
    ).all()

    ferias_por_bombeiro = {}
    for f in ferias_query:
        d = f.data_inicio
        while d <= f.data_fim:
            if d.year == ano and d.month == mes:
                if f.bombeiro_id not in ferias_por_bombeiro:
                    ferias_por_bombeiro[f.bombeiro_id] = set()
                ferias_por_bombeiro[f.bombeiro_id].add(d.day)
            d += timedelta(days=1)

    meses = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    return render_template('imprimir_escala_mes.html',
                           estrutura=estrutura,
                           dias=dias,
                           mes=mes,
                           ano=ano,
                           meses=meses,
                           categorias_ordem=categorias_ordem,
                           feriados=feriados,
                           ferias_por_bombeiro=ferias_por_bombeiro,
                           date=date)



@app.route('/escala/adicionar', methods=['POST'])
@login_required
def adicionar_escala():
    if current_user.tipo_user != 'Admin':
        flash('Apenas administração pode inserir escalas.', 'danger')
        return redirect(url_for('escala'))

    bombeiro_id = request.form['bombeiro_id']
    data_str = request.form['data_inicio']  # formato YYYY-MM-DD
    turno = request.form['turno']
    categoria = request.form.get('categoria', 'Bombeiro')
    funcao = request.form.get('funcao', '')

    # Converter a data para objeto date
    try:
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Data de início inválida.', 'danger')
        return redirect(url_for('escala'))

    # Determinar horas de início/fim com base no turno
    inicio_str, fim_str = turno_para_horas(turno)
    data_inicio = datetime.combine(data, datetime.strptime(inicio_str, '%H:%M').time())
    data_fim = datetime.combine(data, datetime.strptime(fim_str, '%H:%M').time())

    # Se a hora de fim for menor ou igual à de início, acrescenta 1 dia
    if data_fim <= data_inicio:
        data_fim += timedelta(days=1)

    nova = Escala(
        bombeiro_id=bombeiro_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        turno=turno,
        categoria=categoria,
        funcao=funcao
    )
    db.session.add(nova)
    db.session.commit()
    flash('Escala inserida.', 'success')
    return redirect(url_for('escala'))


def _parse_datetime(valor):
    """Tenta converter uma string de data/hora em datetime. Retorna None se falhar."""
    if not valor:
        return None
    formatos = [
        '%d/%m/%Y %H:%M',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d/%m/%Y %H:%M:%S'
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(str(valor).strip(), fmt)
        except (ValueError, TypeError):
            continue
    return None

def _parse_data(valor):
    """Converte uma string de data (com ou sem hora) para um objeto date.
       Aceita vários formatos. Retorna None se falhar."""
    if not valor:
        return None
    s = str(valor).strip()
    # Se já veio como datetime (pode acontecer com openpyxl), extrai só a data
    if isinstance(valor, datetime):
        return valor.date()
    # Tenta formatos com hora (usa apenas a parte da data)
    formatos_com_hora = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d/%m/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M',
    ]
    for fmt in formatos_com_hora:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date()
        except (ValueError, TypeError):
            continue
    # Tenta formatos apenas com data
    formatos_data = ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y']
    for fmt in formatos_data:
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return None

@app.route('/escala/importar', methods=['POST'])
@login_required
def importar_escala():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito ao Comando/Admin.', 'danger')
        return redirect(url_for('escala'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('escala'))

    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '':
        flash('Ficheiro vazio.', 'warning')
        return redirect(url_for('escala'))

    if not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido. Use um ficheiro Excel (.xlsx).', 'danger')
        return redirect(url_for('escala'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler o ficheiro: {str(e)}', 'danger')
        return redirect(url_for('escala'))

    linhas_importadas = 0
    erros = []

    # Função auxiliar para converter data/hora
    def parse_datetime(valor):
        if not valor:
            return None
        # Se já for datetime (caso do openpyxl)
        if isinstance(valor, datetime):
            return valor
        # Se for date, converte para datetime (início do dia)
        if isinstance(valor, date):
            return datetime.combine(valor, datetime.min.time())
        # Caso contrário, tenta converter strings
        s = str(valor).strip()
        formatos = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%d/%m/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y'
        ]
        for fmt in formatos:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    # Cabeçalhos esperados:
    # Col0: Mecanográfico, Col1: Nome (opcional), Col2: Início, Col3: Fim, Col4: Turno, Col5: Categoria, Col6: Função (opcional)
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        try:
            mecanografico = str(row[0]).strip() if row[0] else None
            # A coluna Nome (índice 1) é ignorada na validação, mas pode existir
            inicio_val = row[2] if len(row) > 2 else None
            fim_val = row[3] if len(row) > 3 else None
            turno = str(row[4]).strip() if len(row) > 4 and row[4] else None
            categoria = str(row[5]).strip() if len(row) > 5 and row[5] else 'Bombeiro'
            funcao = str(row[6]).strip() if len(row) > 6 and row[6] else None
        except IndexError:
            erros.append(f'Linha {row_num}: número insuficiente de colunas.')
            continue

        if not mecanografico or not inicio_val or not fim_val or not turno:
            erros.append(f'Linha {row_num}: campos obrigatórios em falta (mecanográfico, início, fim, turno).')
            continue

        # Validar bombeiro
        bombeiro = Bombeiro.query.filter_by(mecanografico=mecanografico).first()
        if not bombeiro:
            erros.append(f'Linha {row_num}: mecanográfico "{mecanografico}" não encontrado.')
            continue

        # Converter datas (agora suporta datetime nativo)
        data_inicio = parse_datetime(inicio_val)
        data_fim = parse_datetime(fim_val)

        if not data_inicio:
            erros.append(f'Linha {row_num}: data de início inválida "{inicio_val}".')
            continue
        if not data_fim:
            erros.append(f'Linha {row_num}: data de fim inválida "{fim_val}".')
            continue

        # Validar turno
        turnos_validos = [
            '1 - 00h/08h', '2 - 08h/16h', '3 - 16h/24h', '4 - 11h/19h',
            '5 - 10h/18h', '6 - 07h/19h', '7 - 19h/07h', '8 - 08h/20h', '9 - 20h/08h'
        ]
        if turno not in turnos_validos:
            erros.append(f'Linha {row_num}: turno "{turno}" inválido.')
            continue

        # Validar categoria
        categorias_validas = ['Motorista', 'Socorrista', 'Centralista', 'EIP', 'ECIN', 'ELAC', 'Piquete', 'Bombeiro']
        if categoria not in categorias_validas:
            erros.append(f'Linha {row_num}: categoria "{categoria}" inválida.')
            continue

        # Criar a escala
        nova = Escala(
            bombeiro_id=bombeiro.id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            turno=turno,
            categoria=categoria,
            funcao=funcao if funcao else None
        )
        db.session.add(nova)
        linhas_importadas += 1

    db.session.commit()

    if erros:
        # Limita a 5 erros para não sobrecarregar a mensagem
        msg_erro = '; '.join(erros[:5])
        if len(erros) > 5:
            msg_erro += f' e mais {len(erros)-5} erro(s).'
        flash(f'{linhas_importadas} escala(s) importada(s) com sucesso. {len(erros)} erro(s): {msg_erro}', 'warning')
    else:
        flash(f'{linhas_importadas} escala(s) importada(s) com sucesso!', 'success')

    return redirect(url_for('escala'))


@app.route('/escala/exportar')
@login_required
def exportar_escala():
    # Se quiser exportar apenas as escalas visíveis de acordo com os filtros atuais,
    # pode copiar a lógica da função 'escala' para construir a query.
    # Para simplificar, vamos exportar todas as escalas do mês atual, por exemplo.
    # Se preferir aplicar os mesmos filtros do utilizador, teria de passar os parâmetros.
    # Vou usar uma abordagem simples (exporta tudo ordenado).

    escalas = Escala.query.join(Bombeiro).order_by(Escala.data_inicio.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Escalas"

    cabecalhos = ['Mecanográfico', 'Nome', 'Início', 'Fim', 'Turno', 'Categoria', 'Função']
    ws.append(cabecalhos)
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos) + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font

    for e in escalas:
        ws.append([
            e.bombeiro.mecanografico if e.bombeiro else '',
            e.bombeiro.nome if e.bombeiro else '',
            e.data_inicio.strftime('%d/%m/%Y %H:%M') if e.data_inicio else '',
            e.data_fim.strftime('%d/%m/%Y %H:%M') if e.data_fim else '',
            e.turno,
            e.categoria,
            e.funcao or ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='escalas.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------- Trocas de Serviço ----------
@app.route('/trocas', methods=['GET', 'POST'])
@login_required
def trocas():
    # Determinar se é voluntário
    is_voluntario = (current_user.tipo_bombeiro == 'Voluntário' and current_user.tipo_user != 'Admin')

    # Voluntários só podem ver ECIN, profissionais podem ver todos
    separador = request.args.get('tipo', 'ecin' if is_voluntario else 'todas')

    is_central = (current_user.resp_departamento == 'Central' and current_user.tipo_user != 'Admin')

    if request.method == 'POST' and is_central:
        flash('A Central não pode criar pedidos de troca.', 'danger')
        return redirect(url_for('trocas'))

    if request.method == 'POST':
        destino_id = request.form.get('destino_id', type=int)
        data_origem = datetime.strptime(request.form['data_origem'], '%Y-%m-%d').date()
        data_destino = datetime.strptime(request.form['data_destino'], '%Y-%m-%d').date()
        turno_origem = request.form.get('turno_origem', '')
        turno_destino = request.form.get('turno_destino', '')
        motivo = request.form.get('motivo', '')
        tipo_pedido = request.form.get('tipo_pedido', 'assalariado')

        # Voluntários só podem criar trocas ECIN
        if is_voluntario and tipo_pedido == 'assalariado':
            flash('Bombeiros voluntários só podem criar trocas ECIN/ELAC.', 'danger')
            return redirect(url_for('trocas', tipo='ecin'))

        # Validação extra para ECINs
        if tipo_pedido == 'ecin':
            escalado = Escala.query.filter(
                Escala.bombeiro_id == current_user.id,
                func.date(Escala.data_inicio) <= data_origem,
                func.date(Escala.data_fim) >= data_origem,
                Escala.categoria.in_(['ECIN', 'ELAC'])
            ).first()
            if not escalado:
                flash('Não está escalado em ECIN/ELAC para esse dia.', 'danger')
                return redirect(url_for('trocas', tipo='ecin'))

        nova = TrocaServico(
            bombeiro_origem_id=current_user.id,
            bombeiro_destino_id=destino_id,
            data_origem=data_origem,
            data_destino=data_destino,
            turno_origem=turno_origem,
            turno_destino=turno_destino,
            motivo=motivo,
            estado='pendente_colega'
        )
        db.session.add(nova)
        db.session.commit()
        flash('Pedido de troca enviado.', 'success')
        return redirect(url_for('trocas', tipo=tipo_pedido))

    # --- GET: construir a query base ---
    query = TrocaServico.query

    # Aplicar filtro por tipo de troca
    if separador == 'assalariado':
        sub_assalariado = db.session.query(TrocaServico.id) \
            .join(Escala, db.and_(
            Escala.bombeiro_id == TrocaServico.bombeiro_origem_id,
            func.date(Escala.data_inicio) <= TrocaServico.data_origem,
            func.date(Escala.data_fim) >= TrocaServico.data_origem,
            Escala.categoria.in_(['Motorista', 'Socorrista', 'Centralista'])
        )) \
            .join(Bombeiro, Bombeiro.id == TrocaServico.bombeiro_origem_id) \
            .filter(Bombeiro.tipo_bombeiro == 'Profissional') \
            .subquery()
        query = query.filter(TrocaServico.id.in_(sub_assalariado))
    elif separador == 'ecin':
        sub_ecin = db.session.query(TrocaServico.id) \
            .join(Escala, db.and_(
            Escala.bombeiro_id == TrocaServico.bombeiro_origem_id,
            func.date(Escala.data_inicio) <= TrocaServico.data_origem,
            func.date(Escala.data_fim) >= TrocaServico.data_origem,
            Escala.categoria.in_(['ECIN', 'ELAC'])
        )).subquery()
        query = query.filter(TrocaServico.id.in_(sub_ecin))

    # Permissões de visualização
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Comando':
        pedidos = query.order_by(TrocaServico.data_pedido.desc()).all()
    elif current_user.resp_departamento == 'Central':
        pedidos = query.order_by(TrocaServico.data_pedido.desc()).all()
    else:
        pedidos = query.filter(
            (TrocaServico.bombeiro_origem_id == current_user.id) |
            (TrocaServico.bombeiro_destino_id == current_user.id)
        ).order_by(TrocaServico.data_pedido.desc()).all()

    # Listas de bombeiros - ordenadas por nome
    bombeiros = Bombeiro.query.filter(
        Bombeiro.id != current_user.id,
        Bombeiro.ativo == True
    ).order_by(Bombeiro.nome.asc()).all()

    # Bombeiros elegíveis para troca assalariada (apenas profissionais)
    bombeiros_assalariados = Bombeiro.query.join(Escala, Escala.bombeiro_id == Bombeiro.id) \
        .filter(
        Bombeiro.id != current_user.id,
        Bombeiro.ativo == True,
        Bombeiro.tipo_bombeiro == 'Profissional',
        Escala.categoria.in_(['Motorista', 'Socorrista', 'Centralista'])
    ).distinct().order_by(Bombeiro.nome.asc()).all()

    return render_template('trocas.html',
                           pedidos=pedidos,
                           bombeiros=bombeiros,
                           bombeiros_assalariados=bombeiros_assalariados,
                           separador_atual=separador,
                           is_central=is_central,
                           is_voluntario=is_voluntario)


@app.route('/api/colegas_para_troca')
@login_required
def api_colegas_para_troca():
    tipo = request.args.get('tipo', 'assalariado')
    is_voluntario = (current_user.tipo_bombeiro == 'Voluntário' and current_user.tipo_user != 'Admin')

    # Voluntários só podem ver colegas para ECIN
    if is_voluntario and tipo == 'assalariado':
        tipo = 'ecin'

    if tipo == 'assalariado':
        # Apenas profissionais com escalas nas categorias certas
        colegas = Bombeiro.query.join(Escala, Escala.bombeiro_id == Bombeiro.id) \
            .filter(
            Bombeiro.id != current_user.id,
            Bombeiro.ativo == True,
            Bombeiro.tipo_bombeiro == 'Profissional',
            Escala.categoria.in_(['Motorista', 'Socorrista', 'Centralista'])
        ).distinct().order_by(Bombeiro.nome.asc()).all()
    else:
        # ECIN/ELAC: todos os bombeiros ativos (voluntários e profissionais)
        colegas = Bombeiro.query.filter(
            Bombeiro.id != current_user.id,
            Bombeiro.ativo == True
        ).order_by(Bombeiro.nome.asc()).all()

    return jsonify([{
        'id': b.id,
        'nome': b.nome,
        'mecanografico': b.mecanografico,
        'tipo_bombeiro': b.tipo_bombeiro
    } for b in colegas])


@app.route('/trocas/imprimir-ecin')
@login_required
def imprimir_troca_ecin():
    # Se quiser passar dados de uma troca específica, pode usar parâmetros
    return render_template('imprimir_troca_ecin.html')

@app.route('/trocas/imprimir-elac')
@login_required
def imprimir_troca_elac():
    # Se quiser passar dados de uma troca específica, pode usar parâmetros
    return render_template('imprimir_troca_elac.html')


@app.route('/trocas/imprimir/<int:id>')
@login_required
def imprimir_troca(id):
    troca = TrocaServico.query.get_or_404(id)
    tipo = determinar_tipo_troca(troca)
    if tipo == 'ecin':
        return render_template('imprimir_troca_ecin.html', troca=troca)
    elif tipo == 'elac':
        return render_template('imprimir_troca_elac.html', troca=troca)
    else:
        return render_template('imprimir_troca.html', troca=troca)


@app.route('/api/escala_usuario/<int:user_id>')
@login_required
def api_escala_usuario(user_id):
    ano = request.args.get('ano', type=int)
    mes = request.args.get('mes', type=int)
    tipo_filtro = request.args.get('tipo', 'todas')  # 'assalariado', 'ecin', 'todas'

    if not ano or not mes:
        return jsonify({'erro': 'Parâmetros ano e mes obrigatórios'}), 400

    # Query base
    query = Escala.query.filter(
        Escala.bombeiro_id == user_id,
        db.extract('year', Escala.data_inicio) == ano,
        db.extract('month', Escala.data_inicio) == mes
    )

    # Aplicar filtro por tipo
    if tipo_filtro == 'assalariado':
        query = query.filter(Escala.categoria.in_(['Motorista', 'Socorrista', 'Centralista']))
    elif tipo_filtro == 'ecin':
        query = query.filter(Escala.categoria.in_(['ECIN', 'ELAC']))
    # se for 'todas', não aplica filtro

    escalas = query.order_by(Escala.data_inicio.asc()).all()

    result = {}
    for e in escalas:
        dia = e.data_inicio.day
        if dia not in result:
            result[dia] = []
        result[dia].append({
            'turno': e.turno,
            'categoria': e.categoria,
            'funcao': e.funcao or ''
        })

    # Garantir que todos os dias do mês estão representados
    import calendar
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    for dia in range(1, ultimo_dia + 1):
        if dia not in result:
            result[dia] = []

    return jsonify(result)

@app.route('/escala/imprimir')
@login_required
def imprimir_escala():
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int, default=datetime.now().year)
    categoria = request.args.get('categoria', '')
    mecanografico = request.args.get('mecanografico', '')

    query = Escala.query.join(Bombeiro)

    if mes:
        query = query.filter(db.extract('month', Escala.data_inicio) == mes)
        query = query.filter(db.extract('year', Escala.data_inicio) == ano)
    if categoria:
        query = query.filter(Escala.categoria == categoria)
    if mecanografico:
        query = query.filter(Bombeiro.mecanografico == mecanografico)

    if current_user.tipo_user != 'Admin':
        query = query.filter(Escala.bombeiro_id == current_user.id)

    escalas = query.order_by(Escala.data_inicio.asc()).all()

    # Determinar nome e mecanográfico
    nome_bombeiro = None
    mecanografico_bombeiro = None
    if mecanografico:
        bombeiro = Bombeiro.query.filter_by(mecanografico=mecanografico).first()
        if bombeiro:
            nome_bombeiro = bombeiro.nome
            mecanografico_bombeiro = bombeiro.mecanografico
    elif current_user.tipo_user != 'Admin':
        nome_bombeiro = current_user.nome
        mecanografico_bombeiro = current_user.mecanografico

    # Agrupar por categoria (ordem fixa)
    ordem_categorias = ['Motorista', 'Socorrista', 'Centralista', 'EIP', 'ECIN', 'ELAC', 'Piquete', 'Bombeiro']
    categorias_agrupadas = {}
    for cat in ordem_categorias:
        cats = [e for e in escalas if e.categoria == cat]
        if cats:
            categorias_agrupadas[cat] = cats

    return render_template('imprimir_escala.html',
                           categorias_agrupadas=categorias_agrupadas,
                           nome_bombeiro=nome_bombeiro,
                           mecanografico_bombeiro=mecanografico_bombeiro,
                           mes=mes,
                           ano=ano,
                           now=datetime.now())

# ---------- Aceitar/Recusar pelo colega (destino) ----------
def determinar_tipo_troca(troca):
    """Retorna 'assalariado', 'ecin' ou 'elac' conforme a origem da troca."""
    bombeiro = Bombeiro.query.get(troca.bombeiro_origem_id)
    if bombeiro and bombeiro.tipo_bombeiro == 'Profissional':
        # Verifica se tem escala nas categorias de assalariado nesse dia
        escala = Escala.query.filter(
            Escala.bombeiro_id == troca.bombeiro_origem_id,
            func.date(Escala.data_inicio) == troca.data_origem,
            Escala.categoria.in_(['Motorista', 'Socorrista', 'Centralista'])
        ).first()
        if escala:
            return 'assalariado'
    # Se não for assalariado, verifica se é ECIN ou ELAC
    ecin = Ecin.query.filter(
        Ecin.bombeiro_id == troca.bombeiro_origem_id,
        Ecin.data == troca.data_origem,
        Ecin.turno == troca.turno_origem
    ).first()
    if ecin:
        if ecin.categoria == 'ECIN':
            return 'ecin'
        elif ecin.categoria == 'ELAC':
            return 'elac'
    return 'assalariado'  # fallback


@app.route('/trocas/aceitar/<int:id>')
@login_required
def aceitar_troca(id):
    troca = TrocaServico.query.get_or_404(id)

    # O destinatário pode aceitar (incluindo Central)
    if current_user.id != troca.bombeiro_destino_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('trocas'))

    if troca.estado == 'pendente_colega':
        troca.estado = 'aceite_colega'
        db.session.commit()
        flash('Troca aceite. Aguarda aprovação do Comando.', 'success')
    else:
        flash('Estado inválido.', 'warning')

    tipo = determinar_tipo_troca(troca)
    return redirect(url_for('trocas', tipo=tipo))


@app.route('/trocas/recusar/<int:id>')
@login_required
def recusar_troca(id):
    troca = TrocaServico.query.get_or_404(id)

    # O destinatário ou o requerente podem recusar (incluindo Central)
    if current_user.id not in (troca.bombeiro_origem_id, troca.bombeiro_destino_id) \
            and not (current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Comando'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('trocas'))

    if troca.estado in ('pendente_colega', 'aceite_colega'):
        troca.estado = 'recusada'
        db.session.commit()
        flash('Troca recusada.', 'info')
    else:
        flash('Estado inválido.', 'warning')

    tipo = determinar_tipo_troca(troca)
    return redirect(url_for('trocas', tipo=tipo))


@app.route('/trocas/aprovar/<int:id>')
@login_required
def aprovar_troca(id):
    # Central NÃO pode aprovar trocas
    if current_user.resp_departamento == 'Central' and current_user.tipo_user != 'Admin':
        flash('A Central não pode aprovar trocas.', 'danger')
        return redirect(url_for('trocas'))

    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Apenas o Comando pode aprovar trocas.', 'danger')
        return redirect(url_for('trocas'))

    troca = TrocaServico.query.get_or_404(id)
    if troca.estado != 'aceite_colega':
        flash('A troca precisa de ser aceite pelo colega primeiro.', 'warning')
        return redirect(url_for('trocas'))

    tipo = determinar_tipo_troca(troca)

    escalas_origem = Escala.query.filter(
        Escala.bombeiro_id == troca.bombeiro_origem_id,
        func.date(Escala.data_inicio) == troca.data_origem,
        Escala.turno == troca.turno_origem
    ).all()

    escalas_destino = Escala.query.filter(
        Escala.bombeiro_id == troca.bombeiro_destino_id,
        func.date(Escala.data_inicio) == troca.data_destino,
        Escala.turno == troca.turno_destino
    ).all()

    for escala in escalas_origem:
        escala.bombeiro_id = troca.bombeiro_destino_id
    for escala in escalas_destino:
        escala.bombeiro_id = troca.bombeiro_origem_id

    if tipo == 'ecin':
        ecins_origem = Ecin.query.filter(
            Ecin.bombeiro_id == troca.bombeiro_origem_id,
            Ecin.data == troca.data_origem,
            Ecin.turno == troca.turno_origem
        ).all()
        ecins_destino = Ecin.query.filter(
            Ecin.bombeiro_id == troca.bombeiro_destino_id,
            Ecin.data == troca.data_destino,
            Ecin.turno == troca.turno_destino
        ).all()
        for ec in ecins_origem:
            ec.bombeiro_id = troca.bombeiro_destino_id
        for ec in ecins_destino:
            ec.bombeiro_id = troca.bombeiro_origem_id

    troca.estado = 'aprovada'
    db.session.commit()
    flash('Troca aprovada e escalas atualizadas.', 'success')
    return redirect(url_for('trocas', tipo=tipo))




# ---------- Dispensas ----------
@app.route('/dispensas', methods=['GET', 'POST'])
@login_required
def dispensas():
    is_central = (current_user.resp_departamento == 'Central' and current_user.tipo_user != 'Admin')

    # Central pode criar a sua própria dispensa, mas não aprovar
    if request.method == 'POST':
        data_inicio = datetime.strptime(request.form['data_inicio'], '%Y-%m-%d').date()
        data_fim = datetime.strptime(request.form['data_fim'], '%Y-%m-%d').date()
        motivo = request.form.get('motivo', '')
        categoria = request.form.get('categoria_disp', '')
        turno = request.form.get('turno_disp', '')
        creditos_ids_str = request.form.get('creditos_selecionados', '')
        creditos_ids = [int(x.strip()) for x in creditos_ids_str.split(',') if
                        x.strip().isdigit()] if creditos_ids_str else []

        nova = Dispensa(
            bombeiro_id=current_user.id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            motivo=motivo,
            categoria=categoria,
            turno=turno,
            aprovada=False
        )
        db.session.add(nova)
        db.session.commit()

        for cid in creditos_ids:
            credito = CreditoDispensa.query.get(cid)
            if credito and credito.bombeiro_id == current_user.id and credito.observacao == 'Não Gozado':
                credito.dispensa_id = nova.id
        db.session.commit()

        flash('Pedido de dispensa enviado.', 'success')
        return redirect(url_for('dispensas'))

    # GET - listagem com permissões
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Comando':
        dispensas_lista = Dispensa.query.order_by(Dispensa.id.desc()).all()
    elif current_user.resp_departamento == 'Central':
        # Central pode ver todas as dispensas (apenas leitura para aprovações)
        dispensas_lista = Dispensa.query.order_by(Dispensa.id.desc()).all()
    else:
        # Utilizador normal vê apenas as suas dispensas
        dispensas_lista = Dispensa.query.filter_by(bombeiro_id=current_user.id).order_by(Dispensa.id.desc()).all()

    return render_template('dispensas.html', dispensas=dispensas_lista, is_central=is_central,now=date.today())


@app.route('/dispensas/detalhes/<int:id>')
@login_required
def detalhes_dispensa(id):
    dispensa = Dispensa.query.get_or_404(id)

    # Central pode ver detalhes de qualquer dispensa
    is_central = (current_user.resp_departamento == 'Central' and current_user.tipo_user != 'Admin')

    # Permissão: próprio, Admin, Comando ou Central
    if current_user.id != dispensa.bombeiro_id and current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando' and not is_central:
        return jsonify({'erro': 'Acesso negado'}), 403

    creditos = []
    for cred in dispensa.creditos:
        creditos.append({
            'data': cred.data.strftime('%d/%m/%Y'),
            'descricao': cred.descricao,
            'horas': cred.horas
        })

    dados = {
        'bombeiro': dispensa.bombeiro.nome,
        'mecanografico': dispensa.bombeiro.mecanografico,
        'data_inicio': dispensa.data_inicio.strftime('%d/%m/%Y'),
        'data_fim': dispensa.data_fim.strftime('%d/%m/%Y'),
        'categoria': dispensa.categoria or '-',
        'turno': dispensa.turno or '-',
        'motivo': dispensa.motivo or '-',
        'aprovada': dispensa.aprovada,
        'creditos': creditos
    }
    return jsonify(dados)


@app.route('/api/creditos_nao_gozados/<int:user_id>')
@login_required
def api_creditos_nao_gozados(user_id):
    # Apenas o próprio, Admin ou Comando podem ver
    if current_user.id != user_id and current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        return jsonify({'erro': 'Acesso negado'}), 403

    creditos = CreditoDispensa.query.filter(
        CreditoDispensa.bombeiro_id == user_id,
        CreditoDispensa.observacao == 'Não Gozado'
    ).order_by(CreditoDispensa.data).all()

    resultado = [{
        'id': c.id,
        'data': c.data.strftime('%d/%m/%Y'),
        'descricao': c.descricao or '',
        'horas': c.horas
    } for c in creditos]

    return jsonify(resultado)


@app.route('/deslocacoes/imprimir-minhas')
@login_required
def imprimir_minhas_deslocacoes():
    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)

    from sqlalchemy import extract
    deslocacoes = Deslocacao.query.filter(
        Deslocacao.bombeiro_id == current_user.id,
        extract('month', Deslocacao.data) == mes,
        extract('year', Deslocacao.data) == ano
    ).order_by(Deslocacao.data, Deslocacao.hora_inicio).all()

    linhas = []
    for d in deslocacoes:
        linha = {
            'dia': d.data.day,
            'localidade': f"{d.local_origem or ''} → {d.local_destino or ''}",
            'inicio_dia': d.data.day,
            'inicio_hora': d.hora_inicio,
            'fim_dia': d.data_fim.day if d.data_fim else '',
            'fim_hora': d.hora_fim or '',
            'viatura': d.viatura.matricula if d.viatura else '',
            'n_servico': d.n_servico or ''
        }
        linhas.append(linha)

    return render_template('imprimir_deslocacoes_bombeiro.html',
                           mes=mes, ano=ano,
                           nome_bombeiro=current_user.nome,
                           linhas=linhas)


@app.route('/dispensas/anular/<int:id>')
@login_required
def anular_dispensa(id):
    dispensa = Dispensa.query.get_or_404(id)

    # Verificar permissões: apenas o próprio bombeiro pode anular
    if current_user.id != dispensa.bombeiro_id:
        flash('Apenas o bombeiro que solicitou a dispensa pode anulá-la.', 'danger')
        return redirect(url_for('dispensas'))

    # Verificar se a dispensa já foi aprovada
    if dispensa.aprovada:
        flash('Não é possível anular uma dispensa já aprovada.', 'warning')
        return redirect(url_for('dispensas'))

    # Verificar se ainda é possível anular (até ao dia anterior)
    hoje = date.today()
    if dispensa.data_inicio <= hoje:
        flash('Não é possível anular a dispensa após o dia de início.', 'warning')
        return redirect(url_for('dispensas'))

    # Verificar se a dispensa já foi aprovada (redundante, mas seguro)
    if dispensa.aprovada:
        flash('Não é possível anular uma dispensa já aprovada.', 'warning')
        return redirect(url_for('dispensas'))

    # Devolver créditos (se houver)
    creditos = CreditoDispensa.query.filter_by(dispensa_id=dispensa.id, observacao='Gozado').all()
    for cred in creditos:
        cred.observacao = 'Não Gozado'
        cred.dispensa_id = None

    # Remover marcações de dispensa na escala
    from datetime import timedelta
    dia_atual = dispensa.data_inicio
    while dia_atual <= dispensa.data_fim:
        escalas = Escala.query.filter(
            Escala.bombeiro_id == dispensa.bombeiro_id,
            func.date(Escala.data_inicio) == dia_atual,
            Escala.observacao == 'dispensado'
        ).all()
        for e in escalas:
            e.observacao = None
        dia_atual += timedelta(days=1)

    # Remover a dispensa
    db.session.delete(dispensa)
    db.session.commit()

    flash('Dispensa anulada com sucesso. Os créditos foram restituídos.', 'success')
    return redirect(url_for('dispensas'))

@app.route('/dispensas/imprimir/<int:id>')
@login_required
def imprimir_dispensa(id):
    dispensa = Dispensa.query.get_or_404(id)
    # Apenas o próprio, Admin ou Comando podem imprimir
    if current_user.id != dispensa.bombeiro_id and current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dispensas'))
    return render_template('imprimir_dispensa.html', dispensa=dispensa)


@app.route('/dispensas/aprovar/<int:id>')
@login_required
def aprovar_dispensa(id):
    # Central NÃO pode aprovar dispensas
    if current_user.resp_departamento == 'Central' and current_user.tipo_user != 'Admin':
        flash('A Central não pode aprovar dispensas.', 'danger')
        return redirect(url_for('dispensas'))

    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Apenas o Comando pode aprovar dispensas.', 'danger')
        return redirect(url_for('dispensas'))

    dispensa = Dispensa.query.get_or_404(id)
    if dispensa.aprovada:
        flash('Dispensa já aprovada.', 'info')
        return redirect(url_for('dispensas'))

    dispensa.aprovada = True

    creditos = CreditoDispensa.query.filter_by(dispensa_id=dispensa.id, observacao='Não Gozado').all()
    for cred in creditos:
        cred.observacao = 'Gozado'

    from datetime import timedelta
    dia_atual = dispensa.data_inicio
    while dia_atual <= dispensa.data_fim:
        escalas = Escala.query.filter(
            Escala.bombeiro_id == dispensa.bombeiro_id,
            func.date(Escala.data_inicio) == dia_atual
        ).all()
        for e in escalas:
            e.observacao = 'dispensado'
        dia_atual += timedelta(days=1)

    db.session.commit()
    flash('Dispensa aprovada. Créditos atualizados e escala(s) marcada(s) como dispensado.', 'success')
    return redirect(url_for('dispensas'))

# ---------- Gestão de Créditos de Dispensa ----------
@app.route('/creditos')
@login_required
def listar_creditos():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito ao Comando/Admin.', 'danger')
        return redirect(url_for('dashboard'))

    bombeiro_id = request.args.get('bombeiro_id', type=int)
    query = CreditoDispensa.query
    if bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)

    # Ordenar por nome do bombeiro (fazendo join com Bombeiro)
    query = query.join(Bombeiro).order_by(Bombeiro.nome.asc())
    creditos = query.all()

    # Lista de bombeiros para o dropdown de filtro
    bombeiros = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()
    return render_template('creditos.html', creditos=creditos, bombeiros=bombeiros, bombeiro_id=bombeiro_id)



@app.route('/creditos/adicionar', methods=['POST'])
@login_required
def adicionar_credito():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_creditos'))

    bombeiro_id = request.form['bombeiro_id']
    data_str = request.form['data']
    descricao = request.form.get('descricao', '')
    horas = request.form.get('horas', 8, type=int)   # default 8

    data = datetime.strptime(data_str, '%Y-%m-%d').date()

    novo = CreditoDispensa(
        bombeiro_id=bombeiro_id,
        data=data,
        descricao=descricao,
        horas=horas,
        observacao='Em Análise'  # ← mude aqui (antes estava 'Não Gozado')
    )
    db.session.add(novo)
    db.session.commit()
    flash('Crédito adicionado.', 'success')
    return redirect(url_for('listar_creditos'))


@app.route('/creditos/apagar/<int:id>')
@login_required
def apagar_credito(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_creditos'))

    credito = CreditoDispensa.query.get_or_404(id)
    db.session.delete(credito)
    db.session.commit()
    flash('Crédito removido.', 'info')
    return redirect(url_for('listar_creditos'))


@app.route('/creditos/exportar')
@login_required
def exportar_creditos():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_creditos'))

    creditos = CreditoDispensa.query.order_by(CreditoDispensa.data.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Créditos Dispensa"

    # Cabeçalhos
    cabecalhos = ['ID', 'Mecanográfico', 'Nome', 'Data', 'Descrição', 'Horas', 'Estado']
    ws.append(cabecalhos)

    # Estilo do cabeçalho
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Dados
    for c in creditos:
        ws.append([
            c.id,
            c.bombeiro.mecanografico,
            c.bombeiro.nome,
            c.data.strftime('%d/%m/%Y') if c.data else '',
            c.descricao or '',
            c.horas,
            c.observacao
        ])

    # Ajustar largura das colunas
    col_widths = [5, 15, 30, 12, 30, 8, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        as_attachment=True,
        download_name='creditos_dispensa.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/creditos/importar', methods=['POST'])
@login_required
def importar_creditos():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_creditos'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('listar_creditos'))

    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '':
        flash('Ficheiro vazio.', 'warning')
        return redirect(url_for('listar_creditos'))

    if not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido. Use .xlsx.', 'danger')
        return redirect(url_for('listar_creditos'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {e}', 'danger')
        return redirect(url_for('listar_creditos'))

    linhas_importadas = 0
    erros = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        try:
            mecanografico_val = row[0]
            data_val = row[1]
            descricao_val = row[2] if len(row) > 2 else None
            horas_val = row[3] if len(row) > 3 else 8
        except IndexError:
            erros.append(f'Linha {row_num}: colunas insuficientes.')
            continue

        mecanografico = str(mecanografico_val).strip() if mecanografico_val is not None else ''
        if not mecanografico:
            erros.append(f'Linha {row_num}: mecanográfico em falta.')
            continue

        bombeiro = Bombeiro.query.filter_by(mecanografico=mecanografico).first()
        if not bombeiro:
            erros.append(f'Linha {row_num}: mecanográfico "{mecanografico}" não encontrado.')
            continue

        # Parse da data usando a função _parse_data (já existe no app.py)
        data_credito = _parse_data(data_val)
        if not data_credito:
            erros.append(f'Linha {row_num}: data inválida "{data_val}".')
            continue

        descricao = str(descricao_val).strip() if descricao_val is not None else ''
        try:
            horas = int(horas_val) if horas_val is not None else 8
        except (ValueError, TypeError):
            horas = 8

        novo = CreditoDispensa(
            bombeiro_id=bombeiro.id,
            data=data_credito,
            descricao=descricao,
            horas=horas,
            observacao='Não Gozado'
        )
        db.session.add(novo)
        linhas_importadas += 1

    db.session.commit()

    if erros:
        flash(f'{linhas_importadas} crédito(s) importado(s). {len(erros)} erro(s): ' + '; '.join(erros), 'warning')
    else:
        flash(f'{linhas_importadas} crédito(s) importado(s) com sucesso!', 'success')

    return redirect(url_for('listar_creditos'))

# ---------- Créditos users ----------

@app.route('/meus-creditos', methods=['GET', 'POST'])
@login_required
def meus_creditos():
    # Restrição: apenas Profissionais, Admin ou Comando
    if current_user.tipo_bombeiro != 'Profissional' and current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito apenas a bombeiros profissionais, administradores ou comando.', 'danger')
        return redirect(url_for('dashboard'))

    # Filtro por estado (Gozado / Não Gozado)
    estado = request.args.get('estado', '')

    if request.method == 'POST':
        data_str = request.form['data']
        descricao = request.form.get('descricao', '')
        horas = request.form.get('horas', 8, type=int)

        try:
            data = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data inválida.', 'danger')
            return redirect(url_for('meus_creditos'))

        novo = CreditoDispensa(
            bombeiro_id=current_user.id,
            data=data,
            descricao=descricao,
            horas=horas,
            observacao='Em Análise'
        )
        db.session.add(novo)
        db.session.commit()
        flash('Crédito registado e em análise.', 'success')
        return redirect(url_for('meus_creditos'))

    # GET – listagem dos créditos do utilizador
    query = CreditoDispensa.query.filter_by(bombeiro_id=current_user.id)
    if estado:
        query = query.filter_by(observacao=estado)

    creditos = query.order_by(CreditoDispensa.data.desc()).all()
    estados = ['Não Gozado', 'Gozado']

    return render_template('meus_creditos.html',
                           creditos=creditos,
                           estados=estados,
                           estado_atual=estado)


# ----------  Aprovação Admin Créditos  ----------
@app.route('/creditos/aprovar/<int:id>')
@login_required
def aprovar_credito(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_creditos'))

    credito = CreditoDispensa.query.get_or_404(id)
    if credito.observacao == 'Em Análise':
        credito.observacao = 'Não Gozado'
        db.session.commit()
        flash('Crédito aprovado (estado: Não Gozado).', 'success')
    else:
        flash('Este crédito não está em análise.', 'warning')

    return redirect(url_for('listar_creditos'))


# ---------- Disponibilidades ----------
@app.route('/disponibilidades', methods=['GET', 'POST'])
@login_required
def     disponibilidades():
    if request.method == 'POST':
        # Recolher as datas enviadas (separadas por vírgula)
        datas_str = request.form.get('datas', '')
        categoria = request.form.get('categoria', '')
        turno_extra = request.form.get('turno_extra', '')

        if not datas_str:
            flash('Selecione pelo menos uma data.', 'warning')
            return redirect(url_for('disponibilidades'))

        datas = [d.strip() for d in datas_str.split(',') if d.strip()]
        for data_str in datas:
            try:
                data = datetime.strptime(data_str, '%Y-%m-%d').date()
            except ValueError:
                continue
            # Verificar se já existe uma disponibilidade idêntica
            existente = Disponibilidade.query.filter_by(
                bombeiro_id=current_user.id,
                data=data,
                turno_extra=turno_extra,
                categoria=categoria
            ).first()
            if existente:
                continue
            nova = Disponibilidade(
                bombeiro_id=current_user.id,
                data=data,
                turno_extra=turno_extra,
                categoria=categoria,
                confirmada=False
            )
            db.session.add(nova)
        db.session.commit()
        flash('Disponibilidade(s) registada(s).', 'success')
        return redirect(url_for('disponibilidades'))

    # GET – listagem com filtros
    bombeiro_id = request.args.get('bombeiro_id', type=int)
    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    filtro = request.args.get('filtro', 'todas')   # NOVO

    query = Disponibilidade.query

    # Apenas Admin/Comando podem ver todos; os restantes ficam limitados ao seu ID
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'ECIN']:
        query = query.filter_by(bombeiro_id=current_user.id)
    elif bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)

    if mes and ano:
        query = query.filter(
            db.extract('month', Disponibilidade.data) == mes,
            db.extract('year', Disponibilidade.data) == ano
        )
    elif mes:
        query = query.filter(db.extract('month', Disponibilidade.data) == mes)
    elif ano:
        query = query.filter(db.extract('year', Disponibilidade.data) == ano)

    # Filtro por estado
    if filtro == 'confirmadas':
        query = query.filter_by(confirmada=True)
    elif filtro == 'nao_confirmadas':
        query = query.filter_by(confirmada=False)

    lista = query.order_by(Disponibilidade.data.desc()).all()

    # Lista de bombeiros ativos (apenas para Admin/Comando)
    bombeiros_ativos = []
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento in ('Comando', 'ECIN'):
        # Lista de bombeiros que têm pelo menos uma disponibilidade (nos filtros atuais ou não)
        subquery = db.session.query(Disponibilidade.bombeiro_id).distinct()
        bombeiros_ativos = Bombeiro.query.filter(Bombeiro.id.in_(subquery)).order_by(Bombeiro.nome).all()

    now = date.today()
    return render_template('disponibilidades.html',
                           disponibilidades=lista,
                           bombeiros=bombeiros_ativos,
                           now=now,
                           bombeiro_id=bombeiro_id,
                           mes=mes,
                           ano=ano,
                           filtro_atual=filtro)   # NOVO

@app.route('/disponibilidades/apagar', methods=['POST'])
@login_required
def apagar_disponibilidades_mes():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('disponibilidades'))

    mes = request.form.get('mes', type=int)
    ano = request.form.get('ano', type=int)
    bombeiro_id = request.form.get('bombeiro_id', type=int)  # opcional

    if not mes or not ano:
        flash('Mês e ano são obrigatórios para apagar.', 'warning')
        return redirect(url_for('disponibilidades'))

    query = Disponibilidade.query.filter(
        db.extract('month', Disponibilidade.data) == mes,
        db.extract('year', Disponibilidade.data) == ano
    )
    if bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)

    count = query.count()
    query.delete()
    db.session.commit()
    flash(f'{count} disponibilidade(s) apagada(s) (mês {mes}/{ano}).', 'success')
    return redirect(url_for('disponibilidades', mes=mes, ano=ano, bombeiro_id=bombeiro_id))


@app.route('/disponibilidades/aprovar', methods=['POST'])
@login_required
def aprovar_disponibilidades():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('disponibilidades'))

    ids_str = request.form.get('ids', '')
    if not ids_str:
        flash('Nenhuma disponibilidade selecionada.', 'warning')
        return redirect(url_for('disponibilidades'))

    ids = [int(i) for i in ids_str.split(',') if i.strip().isdigit()]
    for id in ids:
        disp = Disponibilidade.query.get(id)
        if disp and not disp.confirmada:
            disp.confirmada = True

            # Criar ECIN automaticamente
            existente = Ecin.query.filter_by(
                bombeiro_id=disp.bombeiro_id,
                data=disp.data,
                turno=disp.turno_extra
            ).first()
            if not existente:
                novo_ecin = Ecin(
                    bombeiro_id=disp.bombeiro_id,
                    data=disp.data,
                    turno=disp.turno_extra,
                    estado='Pendente'
                )
                db.session.add(novo_ecin)

    db.session.commit()
    flash(f'{len(ids)} disponibilidade(s) aprovada(s) e registada(s) em ECINS.', 'success')
    return redirect(url_for('disponibilidades'))

@app.route('/avarias/imprimir/<int:id>')
@login_required
def imprimir_avaria(id):
    avaria = Avaria.query.get_or_404(id)
    return render_template('imprimir_avaria.html', avaria=avaria)

@app.route('/disponibilidades/imprimir')
@login_required
def imprimir_disponibilidades():
    from datetime import date
    bombeiro_id = request.args.get('bombeiro_id', type=int)
    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)

    if current_user.tipo_user == 'Admin' or current_user.resp_departamento in ['Comando', 'ECIN']:
        if bombeiro_id:
            bombeiro = Bombeiro.query.get_or_404(bombeiro_id)
        else:
            bombeiro = current_user
    else:
        bombeiro = current_user

    # Buscar disponibilidades
    disponibilidades = Disponibilidade.query.filter(
        Disponibilidade.bombeiro_id == bombeiro.id,
        db.extract('year', Disponibilidade.data) == ano,
        db.extract('month', Disponibilidade.data) == mes
    ).all()

    # Dicionário com chave 'YYYY-MM-DD' para fácil acesso
    # Dicionário para guardar os dois turnos por dia
    disp_dict = {}
    for d in disponibilidades:
        key = d.data.isoformat()  # "YYYY-MM-DD"
        if key not in disp_dict:
            disp_dict[key] = {'07h/19h': None, '19h/07h': None}
        if d.turno_extra == '07h/19h':
            disp_dict[key]['07h/19h'] = d.categoria
        elif d.turno_extra == '19h/07h':
            disp_dict[key]['19h/07h'] = d.categoria

    import calendar
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dias = list(range(1, ultimo_dia + 1))

    meses = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    return render_template('imprimir_disponibilidades.html',
                           bombeiro=bombeiro,
                           dias=dias,
                           mes=mes,
                           ano=ano,
                           meses=meses,
                           disp_dict=disp_dict,
                           date=date)  # ← passa a função date para o template




# ---------- Confirmar Disponibiidade ----------
@app.route('/disponibilidades/confirmar', methods=['POST'])
@login_required
def confirmar_disponibilidade():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('disponibilidades'))

    id = request.form.get('id', type=int)
    if not id:
        flash('ID em falta.', 'warning')
        return redirect(url_for('disponibilidades'))

    disp = Disponibilidade.query.get_or_404(id)
    disp.confirmada = True

    # Criar automaticamente um registo ECIN, se ainda não existir
    existente = Ecin.query.filter_by(
        bombeiro_id=disp.bombeiro_id,
        data=disp.data,
        turno=disp.turno_extra
    ).first()
    if not existente:
        novo_ecin = Ecin(
            bombeiro_id=disp.bombeiro_id,
            data=disp.data,
            turno=disp.turno_extra,
            estado='Pendente'
        )
        db.session.add(novo_ecin)

    db.session.commit()
    flash('Disponibilidade confirmada e registada em ECINS.', 'success')
    return redirect(url_for('disponibilidades'))



#_________Exportar disponibilidades----
@app.route('/disponibilidades/exportar')
@login_required
def exportar_disponibilidades():
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento not in ['Comando', 'ECIN']:
        lista = Disponibilidade.query.order_by(Disponibilidade.data.desc()).all()
    else:
        lista = Disponibilidade.query.filter_by(bombeiro_id=current_user.id)\
                                     .order_by(Disponibilidade.data.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Disponibilidades"

    cabecalhos = ['Bombeiro', 'Data', 'Turno', 'Confirmada']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for d in lista:
        ws.append([
            d.bombeiro.nome,
            d.data.strftime('%d/%m/%Y') if d.data else '',
            d.turno_extra,
            'Sim' if d.confirmada else 'Não'
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='disponibilidades.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---------- Ecins ----------
@app.route('/ecins')
@login_required
def listar_ecins():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    ordem = request.args.get('ordem', 'data')
    mec = request.args.get('mec', '').strip()
    nome_filtro = request.args.get('nome', '')      # novo: bombeiro ID
    data_filtro = request.args.get('data', '')      # novo: data específica (YYYY-MM-DD)
    turno_filtro = request.args.get('turno', '')    # novo: '07h/19h' ou '19h/07h'

    query = Ecin.query

    # Filtro por mecanográfico (pesquisa)
    if mec:
        query = query.join(Bombeiro).filter(Bombeiro.mecanografico.ilike(f'%{mec}%'))

    # Filtro por nome (bombeiro específico)
    if nome_filtro and nome_filtro.isdigit():
        query = query.filter(Ecin.bombeiro_id == int(nome_filtro))

    # Filtro por data (dia exato)
    if data_filtro:
        try:
            data_obj = datetime.strptime(data_filtro, '%Y-%m-%d').date()
            query = query.filter(Ecin.data == data_obj)
        except ValueError:
            pass

    # Filtro por turno
    if turno_filtro:
        query = query.filter(Ecin.turno == turno_filtro)

    # Ordenação
    if ordem == 'nome':
        query = query.join(Bombeiro).order_by(Bombeiro.nome.asc(), Ecin.data.asc())
    else:
        query = query.order_by(Ecin.data.asc())

    registos = query.all()

    # Lista de bombeiros que têm registos ECIN/ELAC (para o combobox)
    bombeiros_com_registos = db.session.query(Bombeiro).join(Ecin).distinct().order_by(Bombeiro.nome).all()

    # Para o modal "Novo Registo" – todos os bombeiros ativos
    bombeiros_ativos = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()

    return render_template('ecins.html',
                           registos=registos,
                           bombeiros_ativos=bombeiros_ativos,
                           bombeiros_com_registos=bombeiros_com_registos,
                           agora=date.today(),
                           ordem_atual=ordem,
                           mec_pesquisa=mec,
                           nome_filtro=nome_filtro,
                           data_filtro=data_filtro,
                           turno_filtro=turno_filtro)

# ---------- Adicionar Ecins ----------
@app.route('/ecins/adicionar', methods=['POST'])
@login_required
def adicionar_ecin():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_ecins'))

    bombeiro_id = request.form['bombeiro_id']
    data_str = request.form['data']
    turno = request.form['turno']
    categoria = request.form['categoria']   # 'ECIN' ou 'ELAC'

    try:
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Data inválida.', 'danger')
        return redirect(url_for('listar_ecins'))

    novo = Ecin(
        bombeiro_id=bombeiro_id,
        data=data,
        turno=turno,
        categoria=categoria,
        estado='Pendente'
    )
    db.session.add(novo)
    db.session.commit()
    flash('Registo ECIN criado.', 'success')
    return redirect(url_for('listar_ecins'))

# ----------Escalar Ecins ----------

@app.route('/ecins/escalar/<int:id>')
@login_required
def escalar_ecin(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_ecins'))

    ecin = Ecin.query.get_or_404(id)
    funcao_cod = request.args.get('funcao', 'X')

    mapeamento = {
        'M':  {'funcao': 'Motorista', 'categoria': 'ECIN', 'estado': 'Motorista ECIN'},
        'C':  {'funcao': 'Chefe',     'categoria': 'ECIN', 'estado': 'Chefe ECIN'},
        'G':  {'funcao': 'Guarnição', 'categoria': 'ECIN', 'estado': 'Guarnição ECIN'},
        'Me': {'funcao': 'Motorista', 'categoria': 'ELAC', 'estado': 'Motorista ELAC'},
        'Ce': {'funcao': 'Chefe',     'categoria': 'ELAC', 'estado': 'Chefe ELAC'},
        'X':  {'estado': 'Não Escalado'}
    }

    if funcao_cod not in mapeamento:
        flash('Opção inválida.', 'danger')
        return redirect(url_for('listar_ecins'))

    dados = mapeamento[funcao_cod]

    # ---------- REMOVER QUALQUER ESCALA ANTERIOR DESTE ECIN ----------
    # Procura por data (ignorando horas) e turno, para o mesmo bombeiro
    escala_antiga = Escala.query.filter(
        Escala.bombeiro_id == ecin.bombeiro_id,
        func.date(Escala.data_inicio) == ecin.data,
        Escala.turno == ecin.turno,
        Escala.categoria.in_(['ECIN', 'ELAC'])
    ).first()
    if escala_antiga:
        db.session.delete(escala_antiga)

    # ---------- SE FOR "NÃO ESCALAR", REMOVE SÓ A ESCALA E ATUALIZA O ESTADO ----------
    if funcao_cod == 'X':
        ecin.estado = dados['estado']
        ecin.funcao = None
        ecin.categoria = None
        db.session.commit()
        flash('Bombeiro removido da escala e marcado como Não Escalado.', 'info')
        return redirect(url_for('listar_ecins'))

    # ---------- CRIAR A NOVA ESCALA ----------
    try:
        partes = ecin.turno.split('-')[1].strip().split('/')
        inicio_str = partes[0].replace('h', ':00')
        fim_str = partes[1].replace('h', ':00')
    except Exception:
        inicio_str, fim_str = '08:00', '20:00'

    data = ecin.data
    inicio_time = datetime.strptime(inicio_str, '%H:%M').time()
    fim_time = datetime.strptime(fim_str, '%H:%M').time()
    data_inicio = datetime.combine(data, inicio_time)
    data_fim = datetime.combine(data, fim_time)
    if data_fim <= data_inicio:
        data_fim += timedelta(days=1)

    nova_escala = Escala(
        bombeiro_id=ecin.bombeiro_id,
        data_inicio=data_inicio,
        data_fim=data_fim,
        turno=ecin.turno,
        categoria=dados['categoria'],
        funcao=dados['funcao']
    )
    db.session.add(nova_escala)

    ecin.funcao = dados['funcao']
    ecin.categoria = dados['categoria']
    ecin.estado = dados['estado']
    db.session.commit()
    flash(f'{ecin.bombeiro.nome} escalado como {dados["estado"]}.', 'success')
    return redirect(url_for('listar_ecins'))

@app.route('/ecins/escalar_ajax/<int:id>', methods=['GET'])
@login_required
def escalar_ecin_ajax(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        return jsonify({'error': 'Acesso restrito'}), 403

    ecin = Ecin.query.get_or_404(id)
    funcao_cod = request.args.get('funcao', 'X')

    mapeamento = {
        'M':  {'funcao': 'Motorista', 'categoria': 'ECIN', 'estado': 'Motorista ECIN'},
        'C':  {'funcao': 'Chefe',     'categoria': 'ECIN', 'estado': 'Chefe ECIN'},
        'G':  {'funcao': 'Guarnição', 'categoria': 'ECIN', 'estado': 'Guarnição ECIN'},
        'Me': {'funcao': 'Motorista', 'categoria': 'ELAC', 'estado': 'Motorista ELAC'},
        'Ce': {'funcao': 'Chefe',     'categoria': 'ELAC', 'estado': 'Chefe ELAC'},
        'X':  {'estado': 'Não Escalado'}
    }

    if funcao_cod not in mapeamento:
        return jsonify({'error': 'Opção inválida'}), 400

    dados = mapeamento[funcao_cod]

    # Remover escala antiga se existir
    from sqlalchemy import func
    if ecin.categoria and ecin.funcao:
        escalas = Escala.query.filter(
            Escala.bombeiro_id == ecin.bombeiro_id,
            func.date(Escala.data_inicio) == ecin.data,
            Escala.turno == ecin.turno,
            Escala.categoria == ecin.categoria,
            Escala.funcao == ecin.funcao
        ).all()
        for escala in escalas:
            db.session.delete(escala)

    if funcao_cod == 'X':
        ecin.estado = dados['estado']
        ecin.funcao = None
        ecin.categoria = None
    else:
        # Criar nova escala
        try:
            partes = ecin.turno.split('-')[1].strip().split('/')
            inicio_str = partes[0].replace('h', ':00')
            fim_str = partes[1].replace('h', ':00')
        except Exception:
            inicio_str, fim_str = '08:00', '20:00'

        data = ecin.data
        inicio_time = datetime.strptime(inicio_str, '%H:%M').time()
        fim_time = datetime.strptime(fim_str, '%H:%M').time()
        data_inicio = datetime.combine(data, inicio_time)
        data_fim = datetime.combine(data, fim_time)
        if data_fim <= data_inicio:
            data_fim += timedelta(days=1)

        nova_escala = Escala(
            bombeiro_id=ecin.bombeiro_id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            turno=ecin.turno,
            categoria=dados['categoria'],
            funcao=dados['funcao']
        )
        db.session.add(nova_escala)

        ecin.funcao = dados['funcao']
        ecin.categoria = dados['categoria']
        ecin.estado = dados['estado']

    db.session.commit()

    # Preparar resposta com o novo estado
    return jsonify({
        'success': True,
        'novo_estado': ecin.estado,
        'estado_class': get_estado_class(ecin.estado),
        'estado_texto': get_estado_texto(ecin.estado),
        'mostrar_botoes': ecin.estado == 'Pendente'
    })

def get_estado_class(estado):
    if estado == 'Pendente':
        return 'bg-warning text-dark'
    elif estado == 'Não Escalado':
        return 'bg-danger'
    else:
        return 'bg-success'

def get_estado_texto(estado):
    return estado



#----------Imprimir Disponibilidade ECINS----------------
@app.route('/ecins/imprimir')
@login_required
def imprimir_ecins():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    mes = request.args.get('mes', type=int)
    ano = request.args.get('ano', type=int)
    if not mes or not ano:
        hoje = date.today()
        mes = hoje.month
        ano = hoje.year

    # Filtrar ECINS do mês/ano, ordenados por data
    ecins = Ecin.query.filter(
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    ).order_by(Ecin.data).all()

    # Agrupar por dia e turno
    from collections import defaultdict
    escala = defaultdict(lambda: defaultdict(list))
    for ec in ecins:
        dia_str = ec.data.strftime('%d/%m/%Y')
        turno_str = ec.turno
        # Mapear turnos conhecidos: "07h/19h" e "19h/07h"
        if turno_str in ['07h/19h', '19h/07h']:
            escala[dia_str][turno_str].append(ec.bombeiro.nome)
        else:
            # Se não for um dos turnos padrão, colocar em "Outro"
            escala[dia_str]['Outro'].append(ec.bombeiro.nome)

    # Ordenar dias cronologicamente
    dias_ordenados = sorted(escala.keys(), key=lambda d: datetime.strptime(d, '%d/%m/%Y'))

    meses_nomes = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro',
                   'Novembro', 'Dezembro']

    return render_template('imprimir_ecins.html',
                           escala=escala,
                           dias=dias_ordenados,
                           mes=mes,
                           ano=ano,
                           meses=meses_nomes)

@app.route('/ecins/modificar/<int:id>')
@login_required
def modificar_ecin(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria', 'ECIN']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('listar_ecins'))

    ecin = Ecin.query.get_or_404(id)

    # Remove a escala associada (se existir)
    if ecin.categoria and ecin.funcao:
        escalas = Escala.query.filter_by(
            bombeiro_id=ecin.bombeiro_id,
            turno=ecin.turno,
            categoria=ecin.categoria,
            funcao=ecin.funcao
        ).filter(func.date(Escala.data_inicio) == ecin.data).all()
        for escala in escalas:
            db.session.delete(escala)

    # Força o estado para Pendente e limpa os campos de escalonamento
    ecin.estado = 'Pendente'
    ecin.funcao = None
    ecin.categoria = None
    db.session.commit()

    flash('Registo modificado. O bombeiro voltou a ficar Pendente.', 'info')
    return redirect(url_for('listar_ecins'))

# ---------- Imprimir Escala ECin----------

@app.route('/ecins/imprimir-escala-ecin')
@login_required
def imprimir_escala_ecin():
    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)

    # Verificar permissões: Admin, Comando, Secretaria, ECIN ou bombeiro escalado no mês/ano
    tem_permissao = (current_user.tipo_user == 'Admin' or
                     current_user.resp_departamento in ['Comando', 'Secretaria', 'ECIN'])

    if not tem_permissao:
        # Verificar se o bombeiro actual tem alguma escala ECIN no mês/ano na tabela Escala
        escala_user = Escala.query.filter(
            Escala.bombeiro_id == current_user.id,
            db.extract('year', Escala.data_inicio) == ano,
            db.extract('month', Escala.data_inicio) == mes,
            Escala.categoria == 'ECIN'
        ).first()
        if not escala_user:
            flash('Acesso restrito: não está escalado para ECIN neste mês.', 'danger')
            return redirect(url_for('escala'))

    # Buscar ECINs do mês/ano na tabela Ecin (para gerar o conteúdo)
    ecins = Ecin.query.filter(
        Ecin.categoria == 'ECIN',
        Ecin.estado.in_(['Motorista ECIN', 'Chefe ECIN', 'Guarnição ECIN']),
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    ).order_by(Ecin.data, Ecin.turno, Ecin.funcao).all()

    # Construir estrutura: escala[dia][turno][funcao] = lista de bombeiros
    from collections import defaultdict
    escala = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for ec in ecins:
        escala[ec.data.day][ec.turno][ec.funcao].append({
            'nome': ec.bombeiro.nome,
            'mecanografico': ec.bombeiro.mecanografico,
            'posto': ec.bombeiro.posto
        })

    import calendar
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dias = list(range(1, ultimo_dia + 1))

    weekend_days = []
    for dia in dias:
        if date(ano, mes, dia).weekday() >= 5:
            weekend_days.append(dia)

    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro',
             'Novembro', 'Dezembro']

    return render_template('imprimir_ecin_escala.html',
                           escala=escala,
                           dias=dias,
                           mes=mes,
                           ano=ano,
                           meses=meses,
                           weekend_days=weekend_days)


@app.route('/ecins/imprimir-escala-elac')
@login_required
def imprimir_escala_elac():
    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)

    # Verificar permissões: Admin, Comando, Secretaria, ECIN ou bombeiro escalado no mês/ano
    tem_permissao = (current_user.tipo_user == 'Admin' or
                     current_user.resp_departamento in ['Comando', 'Secretaria', 'ECIN'])

    if not tem_permissao:
        # Verificar se o bombeiro actual tem alguma escala ELAC no mês/ano na tabela Escala
        escala_user = Escala.query.filter(
            Escala.bombeiro_id == current_user.id,
            db.extract('year', Escala.data_inicio) == ano,
            db.extract('month', Escala.data_inicio) == mes,
            Escala.categoria == 'ELAC'
        ).first()
        if not escala_user:
            flash('Acesso restrito: não está escalado para ELAC neste mês.', 'danger')
            return redirect(url_for('escala'))

    # Buscar ELACs do mês/ano na tabela Ecin
    elacs = Ecin.query.filter(
        Ecin.categoria == 'ELAC',
        Ecin.estado.in_(['Motorista ELAC', 'Chefe ELAC']),
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    ).order_by(Ecin.data, Ecin.turno, Ecin.funcao).all()

    from collections import defaultdict
    escala = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for ec in elacs:
        escala[ec.data.day][ec.turno][ec.funcao].append({
            'nome': ec.bombeiro.nome,
            'mecanografico': ec.bombeiro.mecanografico,
            'posto': ec.bombeiro.posto
        })

    import calendar
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    dias = list(range(1, ultimo_dia + 1))

    weekend_days = []
    for dia in dias:
        if date(ano, mes, dia).weekday() >= 5:
            weekend_days.append(dia)

    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro',
             'Novembro', 'Dezembro']

    return render_template('imprimir_elac_escala.html',
                           escala=escala,
                           dias=dias,
                           mes=mes,
                           ano=ano,
                           meses=meses,
                           weekend_days=weekend_days)



# ---------- Bombeiro Perfil ----------

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    bombeiro = current_user  # já é o objeto do utilizador logado

    if request.method == 'POST':
        novo_email = request.form['email'].strip().lower()
        novo_telemovel = request.form.get('telemovel', '').strip()
        nova_password = request.form.get('password', '')

        # Verificar duplicação de email (exceto o próprio)
        conflito_email = Bombeiro.query.filter(
            Bombeiro.id != bombeiro.id,
            Bombeiro.email == novo_email
        ).first()
        if conflito_email:
            flash('Este email já está a ser usado por outro bombeiro.', 'warning')
            return redirect(url_for('perfil'))

        # Verificar duplicação de telemóvel (se preenchido)
        if novo_telemovel:
            conflito_tele = Bombeiro.query.filter(
                Bombeiro.id != bombeiro.id,
                Bombeiro.telemovel == novo_telemovel
            ).first()
            if conflito_tele:
                flash('Este número de telemóvel já está associado a outro bombeiro.', 'warning')
                return redirect(url_for('perfil'))

        # Atualizar campos
        bombeiro.email = novo_email
        bombeiro.telemovel = novo_telemovel if novo_telemovel else None

        # Só altera a password se foi preenchida
        if nova_password:
            bombeiro.password_hash = generate_password_hash(nova_password)

        db.session.commit()
        flash('Perfil atualizado com sucesso.', 'success')
        return redirect(url_for('perfil'))

    # GET → mostra o formulário pré‑preenchido
    return render_template('perfil.html')



# ---------- Checklist ----------
@app.route('/checklist', methods=['GET', 'POST'])
@login_required
def checklist():
    if request.method == 'POST':
        viatura_id = request.form['viatura_id']
        itens = request.form['itens']
        observacoes = request.form.get('observacoes', '')
        nova = Checklist(
            viatura_id=viatura_id,
            bombeiro_id=current_user.id,
            itens_verificados=itens,
            observacoes=observacoes
        )
        db.session.add(nova)
        db.session.commit()
        flash('Checklist registado.', 'success')
        return redirect(url_for('checklist'))

    checklists_lista = Checklist.query.order_by(Checklist.data.desc()).all()
    viaturas = Viatura.query.all()
    return render_template('checklist.html', checklists=checklists_lista, viaturas=viaturas)


# ---------- Fardamento ----------
@app.route('/fardamento', methods=['GET', 'POST'])
@login_required
def fardamento():
    aba = request.args.get('tab', 'pedidos')
    bombeiro_id_filtro = request.args.get('bombeiro_id', type=int)
    hoje = date.today()

    # ---- POST: processar formulários ----
    if request.method == 'POST':
        form_type = request.form.get('form_type')

        # ---------- NOVO PEDIDO ----------
        if form_type == 'pedido':
            tipo = request.form['tipo']
            nome = request.form['nome']
            tamanho = request.form['tamanho']
            motivo = request.form['motivo']
            descricao_motivo = request.form.get('descricao_motivo', '')
            stock_id = request.form.get('stock_id', type=int)

            descricao = ''
            if stock_id:
                item_stock = StockFardamentoArmazem.query.get(stock_id)
                if item_stock:
                    descricao = item_stock.descricao or ''

            novo = Fardamento(
                bombeiro_id=current_user.id,
                tipo=tipo,
                nome=nome,
                descricao=descricao,
                tamanho=tamanho,
                motivo=motivo,
                descricao_motivo=descricao_motivo,
                stock_id=stock_id,
                estado='Pedido'
            )
            db.session.add(novo)
            db.session.commit()
            flash('Pedido de fardamento registado.', 'success')
            return redirect(url_for('fardamento', tab='pedidos'))

        # ---------- ATRIBUIÇÃO MÚLTIPLA (novo) ----------
        elif form_type == 'atribuicao_multipla':
            bombeiro_id = request.form.get('bombeiro_id', type=int)
            data_entrega_str = request.form.get('data_entrega')
            if data_entrega_str:
                data_entrega = datetime.strptime(data_entrega_str, '%Y-%m-%d').date()
            else:
                data_entrega = hoje

            if not bombeiro_id:
                flash('Selecione um bombeiro.', 'danger')
                return redirect(url_for('fardamento', tab='atribuido'))

            atribuicoes_feitas = 0
            for key, value in request.form.items():
                if key.startswith('quantidade_') and value and int(value) > 0:
                    item_id = int(key.replace('quantidade_', ''))
                    quantidade = int(value)

                    item = StockFardamentoArmazem.query.get(item_id)
                    if not item or item.stock < quantidade:
                        flash(f'Stock insuficiente para {item.nome} ({item.tamanho or "-"}). Disponível: {item.stock}', 'danger')
                        continue

                    # Criar atribuição
                    atribuicao = FardamentoAtribuido(
                        bombeiro_id=bombeiro_id,
                        tipo=item.tipo,
                        nome=item.nome,
                        tamanho=item.tamanho,
                        data_entrega=data_entrega,
                        estado='Entregue'
                    )
                    db.session.add(atribuicao)

                    # Atualizar stock (diminuir)
                    item.stock -= quantidade
                    atribuicoes_feitas += 1

            db.session.commit()
            if atribuicoes_feitas > 0:
                flash(f'{atribuicoes_feitas} item(ns) atribuído(s) com sucesso.', 'success')
            else:
                flash('Nenhum item foi atribuído.', 'warning')
            return redirect(url_for('fardamento', tab='atribuido'))

        # ---------- ATRIBUIÇÃO ANTIGA (simples, mantida para compatibilidade) ----------
        elif form_type == 'atribuicao':
            idpedido = request.form.get('idpedido', type=int)
            if not idpedido:
                flash('É obrigatório selecionar um pedido de origem.', 'warning')
                return redirect(url_for('fardamento', tab='atribuido'))

            pedido = Fardamento.query.get_or_404(idpedido)
            # ... (código antigo, pode ser mantido ou removido)

    # ---- GET: listagens ----
    tipos = TipoFardaMaterial.query.order_by(TipoFardaMaterial.nome).all()
    bombeiros = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()

    # Pedidos (aba "Pedidos")
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento in ['Comando', 'Fardamento']:
        pedidos = Fardamento.query.order_by(Fardamento.data_registo.desc()).all()
    else:
        pedidos = Fardamento.query.filter_by(bombeiro_id=current_user.id).order_by(Fardamento.data_registo.desc()).all()

    # Atribuições (aba "Atribuído")
    query_atrib = FardamentoAtribuido.query
    if bombeiro_id_filtro:
        query_atrib = query_atrib.filter_by(bombeiro_id=bombeiro_id_filtro)
    atribuicoes = query_atrib.order_by(FardamentoAtribuido.data_entrega.desc()).all()

    return render_template('fardamento.html',
                           pedidos=pedidos,
                           atribuicoes=atribuicoes,
                           tipos=tipos,
                           bombeiros=bombeiros,
                           bombeiro_id_filtro=bombeiro_id_filtro,
                           aba=aba,
                           hoje=hoje)

@app.route('/fardamento/pedidos-analise/<int:bombeiro_id>')
@login_required
def pedidos_analise_bombeiro(bombeiro_id):
    pedidos = Fardamento.query.filter_by(
        bombeiro_id=bombeiro_id,
        estado='Análise',
        responsavel=True,
        comando=True
    ).order_by(Fardamento.data_registo.desc()).all()

    result = [{
        'id': p.id,
        'tipo': p.tipo,
        'nome': p.nome,
        'tamanho': p.tamanho,
        'data': p.data_registo.strftime('%d/%m/%Y') if p.data_registo else ''
    } for p in pedidos]
    return jsonify(result)




@app.route('/fardamento-atribuido/editar/<int:id>', methods=['POST'])
@login_required
def editar_fardamento_atribuido(id):
    item = FardamentoAtribuido.query.get_or_404(id)
    item.tipo = request.form['tipo']
    item.nome = request.form['nome']
    item.tamanho = request.form['tamanho']
    item.data_entrega = datetime.strptime(request.form['data_entrega'], '%Y-%m-%d').date()
    db.session.commit()
    flash('Registo atualizado.', 'success')
    return redirect(url_for('fardamento', tab='atribuido'))

@app.route('/fardamento-atribuido/apagar/<int:id>')
@login_required
def apagar_fardamento_atribuido(id):
    atribuicao = FardamentoAtribuido.query.get_or_404(id)
    devolver_stock = request.args.get('devolver_stock', '0') == '1'

    # Devolver ao stock apenas se o estado for 'Entregue' e o utilizador confirmou
    if devolver_stock and atribuicao.estado == 'Entregue':
        item = StockFardamento.query.filter_by(
            tipo=atribuicao.tipo,
            nome=atribuicao.nome,
            tamanho=atribuicao.tamanho
        ).first()
        if item:
            item.stock += 1
            db.session.add(item)

    # Se houver pedido associado, removê‑lo também
    if atribuicao.idpedido:
        pedido = Fardamento.query.get(atribuicao.idpedido)
        if pedido:
            db.session.delete(pedido)

    db.session.delete(atribuicao)
    db.session.commit()
    flash('Registo(s) removido(s).', 'info')
    return redirect(url_for('fardamento', tab='atribuido'))

@app.route('/fardamento-atribuido/devolver/<int:id>')
@login_required
def devolver_fardamento_atribuido(id):
    atribuicao = FardamentoAtribuido.query.get_or_404(id)
    if atribuicao.estado == 'Entregue':
        # Devolver ao stock
        item = StockFardamento.query.filter_by(
            tipo=atribuicao.tipo,
            nome=atribuicao.nome,
            tamanho=atribuicao.tamanho
        ).first()
        if item:
            item.stock += 1
            db.session.add(item)

        atribuicao.estado = 'Devolvido'
        atribuicao.data_devolucao = date.today()
        db.session.commit()
        flash('Devolução registada e stock atualizado.', 'success')
    else:
        flash('Esta atribuição já foi devolvida.', 'warning')
    return redirect(url_for('fardamento', tab='atribuido'))

# ---------- Fardamento p/tipo ----------
@app.route('/fardamento/nomes-por-tipo/<tipo>')
@login_required
def nomes_por_tipo(tipo):
    itens = StockFardamento.query.filter_by(tipo=tipo).order_by(StockFardamento.nome).all()
    result = [{
        'id': i.id,
        'nome': i.nome,
        'descricao': i.descricao or '',
        'tamanho': i.tamanho or '',
        'stock': i.stock                  # ← adicionado
    } for i in itens]
    from flask import jsonify
    return jsonify(result)

# ---------- Editar Fardamento ----------

@app.route('/fardamento/editar/<int:id>', methods=['POST'])
@login_required
def editar_fardamento(id):
    pedido = Fardamento.query.get_or_404(id)

    # Permissão: Admin/Comando/Fardamento ou o próprio bombeiro dono do pedido
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento'] and current_user.id != pedido.bombeiro_id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('fardamento'))

    # Atualizar campos básicos (só o dono ou admin pode mudar o pedido em estado 'Pedido')
    if pedido.estado == 'Pedido' or current_user.tipo_user == 'Admin' or current_user.resp_departamento in ['Comando', 'Fardamento']:
        pedido.tipo = request.form.get('tipo', pedido.tipo)
        pedido.nome = request.form.get('nome', pedido.nome)
        pedido.tamanho = request.form.get('tamanho', pedido.tamanho)
        pedido.motivo = request.form.get('motivo', pedido.motivo)
        pedido.descricao_motivo = request.form.get('descricao_motivo', '') if pedido.motivo in ['Troca', 'Abatido'] else ''
        stock_id = request.form.get('stock_id', type=int)
        if stock_id:
            pedido.stock_id = stock_id
            item = StockFardamento.query.get(stock_id)
            if item:
                pedido.descricao = item.descricao or ''

    # Campos de aprovação (só Admin/Comando/Fardamento)
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento in ['Comando', 'Fardamento']:
        # Responsável (apenas Admin ou Fardamento)
        if current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Fardamento':
            pedido.responsavel = request.form.get('responsavel') == 'on'
        # Comando
        if current_user.tipo_user == 'Admin' or current_user.resp_departamento == 'Comando':
            pedido.comando = request.form.get('comando') == 'on'

        # Entregue (só se responsável e comando estiverem marcados)
        if pedido.responsavel and pedido.comando:
            pedido.entregue = request.form.get('entregue') == 'on'
            if pedido.entregue:
                pedido.data_entrega = datetime.utcnow()
                pedido.estado = 'Concluido'
                # Atualizar stock
                if pedido.stock_id:
                    item_stock = StockFardamento.query.get(pedido.stock_id)
                    if item_stock and item_stock.stock > 0:
                        item_stock.stock -= 1
        else:
            pedido.entregue = False

        # Estado (pode ser alterado manualmente)
        novo_estado = request.form.get('estado')
        if novo_estado in ['Pedido', 'Análise', 'Concluido']:
            pedido.estado = novo_estado

    db.session.commit()
    flash('Pedido atualizado.', 'success')
    return redirect(url_for('fardamento'))

# ---------- Apagar Fardamento ----------

@app.route('/fardamento/apagar/<int:id>')
@login_required
def apagar_fardamento(id):
    pedido = Fardamento.query.get_or_404(id)
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento'] and current_user.id != pedido.bombeiro_id:
        flash('Sem permissão.', 'danger')
        return redirect(url_for('fardamento'))
    db.session.delete(pedido)
    db.session.commit()
    flash('Pedido removido.', 'info')
    return redirect(url_for('fardamento'))


# ---------- Exportar Fardamento ----------
@app.route('/fardamento/exportar')
@login_required
def exportar_fardamento():
    if current_user.tipo_user == 'Admin' or current_user.resp_departamento in ['Comando', 'Fardamento']:
        pedidos = Fardamento.query.order_by(Fardamento.data_registo.desc()).all()
    else:
        pedidos = Fardamento.query.filter_by(bombeiro_id=current_user.id)\
                                  .order_by(Fardamento.data_registo.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fardamento"

    cabecalhos = ['ID', 'Data', 'Mecanográfico', 'Bombeiro', 'Tipo', 'Nome', 'Tamanho',
                  'Motivo', 'Descrição Motivo', 'Estado', 'Resp.', 'Comando', 'Entregue', 'Data Entrega']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for p in pedidos:
        ws.append([
            p.id,
            p.data_registo.strftime('%d/%m/%Y %H:%M') if p.data_registo else '',
            p.bombeiro.mecanografico,
            p.bombeiro.nome,
            p.tipo,
            p.nome,
            p.tamanho,
            p.motivo,
            p.descricao_motivo or '',
            p.estado,
            'Sim' if p.responsavel else 'Não',
            'Sim' if p.comando else 'Não',
            'Sim' if p.entregue else 'Não',
            p.data_entrega.strftime('%d/%m/%Y %H:%M') if p.data_entrega else ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='fardamento.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ---------- Importar Fardamento ----------
@app.route('/fardamento/importar', methods=['POST'])
@login_required
def importar_fardamento():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('fardamento'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('fardamento'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('fardamento'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('fardamento'))

    linhas_importadas = 0
    erros = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue
        try:
            mecanografico = str(row[2]).strip() if len(row) > 2 and row[2] else None
            tipo = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            nome = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            tamanho = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            motivo = str(row[7]).strip() if len(row) > 7 and row[7] else 'Novo'
            estado = str(row[9]).strip() if len(row) > 9 and row[9] else 'Pedido'
        except Exception:
            erros.append(f'Linha {row_num}: dados inválidos.')
            continue

        bombeiro = Bombeiro.query.filter_by(mecanografico=mecanografico).first() if mecanografico else None
        if not bombeiro:
            erros.append(f'Linha {row_num}: mecanográfico não encontrado.')
            continue

        # Encontrar stock_id pelo nome e tipo (simplificação)
        stock_item = StockFardamento.query.filter_by(nome=nome, tipo=tipo).first()
        stock_id = stock_item.id if stock_item else None

        novo = Fardamento(
            bombeiro_id=bombeiro.id,
            tipo=tipo,
            nome=nome,
            tamanho=tamanho,
            motivo=motivo,
            estado=estado,
            stock_id=stock_id
        )
        db.session.add(novo)
        linhas_importadas += 1

    db.session.commit()
    if erros:
        flash(f'{linhas_importadas} importados. {len(erros)} erro(s): ' + '; '.join(erros), 'warning')
    else:
        flash(f'{linhas_importadas} pedidos importados com sucesso.', 'success')
    return redirect(url_for('fardamento'))


#-----------Stock Fardamento--------------
from models import StockFardamento, StockFardamentoArmazem, TipoFardaMaterial

@app.route('/stock-fardamento', methods=['GET', 'POST'])
@login_required
def stock_fardamento():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    # ---- POST: criar novo item (produto ou variação) ----
    if request.method == 'POST':
        tipo_adicao = request.form.get('tipo_adicao')

        # Caso 1: Novo produto (cria stock_fardamento + primeira variação)
        if tipo_adicao == 'novo_produto':
            tipo = request.form.get('tipo', 'Outro')
            nome = request.form['nome']
            descricao = request.form.get('descricao', '')
            tamanho = request.form.get('tamanho', '').strip()
            stock = request.form.get('stock', 0, type=int)

            # Gerar novo codigo_farda (FA001, FA002...)
            ultimo = StockFardamento.query.order_by(StockFardamento.codigo_farda.desc()).first()
            if ultimo and ultimo.codigo_farda.startswith('FA'):
                try:
                    num = int(ultimo.codigo_farda[2:]) + 1
                except:
                    num = 1
            else:
                num = 1
            codigo_farda = f"FA{num:03d}"

            # Criar produto principal
            produto = StockFardamento(
                codigo_farda=codigo_farda,
                tipo=tipo,
                nome=nome,
                descricao=descricao
            )
            db.session.add(produto)
            db.session.flush()

            # Criar primeira variação (sub_codigo = codigo_farda + '01')
            sub_codigo_farda = f"{codigo_farda}01"
            novo_item = StockFardamentoArmazem(
                codigo_farda=codigo_farda,
                sub_codigo_farda=sub_codigo_farda,
                tipo=tipo,
                nome=nome,
                descricao=descricao,
                tamanho=tamanho if tamanho else None,
                stock=stock
            )
            db.session.add(novo_item)
            db.session.commit()
            flash(f'Produto {codigo_farda} criado com sucesso.', 'success')

        # Caso 2: Nova variação (apenas stock_fardamento_armazem)
        elif tipo_adicao == 'nova_variacao':
            codigo_farda = request.form.get('codigo_farda')
            tamanho = request.form.get('tamanho', '').strip()
            stock = request.form.get('stock', 0, type=int)

            # LOGS PARA DEPURAÇÃO (ver no terminal do Render)
            print(f"DEBUG nova_variacao: codigo_farda={codigo_farda}, tamanho='{tamanho}', stock={stock}")

            if not codigo_farda:
                flash('Erro: Nenhum produto selecionado.', 'danger')
                return redirect(url_for('stock_fardamento'))

            # Tamanho opcional: se vazio, guarda como None
            if tamanho == '':
                tamanho = None

            produto = StockFardamento.query.filter_by(codigo_farda=codigo_farda).first()
            if not produto:
                flash(f'Erro: Produto com código {codigo_farda} não encontrado.', 'danger')
                return redirect(url_for('stock_fardamento'))

            # Gerar sub_codigo_farda sequencial
            ultimo_sub = StockFardamentoArmazem.query.filter(
                StockFardamentoArmazem.codigo_farda == codigo_farda
            ).order_by(StockFardamentoArmazem.sub_codigo_farda.desc()).first()
            if ultimo_sub and ultimo_sub.sub_codigo_farda:
                try:
                    num_sub = int(ultimo_sub.sub_codigo_farda[-2:]) + 1
                except:
                    num_sub = 1
            else:
                num_sub = 1
            sub_codigo_farda = f"{codigo_farda}{num_sub:02d}"

            novo_item = StockFardamentoArmazem(
                codigo_farda=codigo_farda,
                sub_codigo_farda=sub_codigo_farda,
                tipo=produto.tipo,
                nome=produto.nome,
                descricao=produto.descricao,
                tamanho=tamanho,
                stock=stock
            )
            db.session.add(novo_item)
            db.session.commit()
            print(f"DEBUG: Item criado - ID={novo_item.id}, tamanho={novo_item.tamanho}, stock={novo_item.stock}")
            flash(f'Variação {sub_codigo_farda} (tamanho {tamanho if tamanho else "não especificado"}) adicionada a {produto.nome}.',
                'success')

        else:
            flash('Opção inválida.', 'danger')

        return redirect(url_for('stock_fardamento'))

    # ---- GET: listagem (ordenada por codigo_farda) ----
    produtos = StockFardamento.query.order_by(StockFardamento.codigo_farda.asc()).all()
    tipos = TipoFardaMaterial.query.order_by(TipoFardaMaterial.nome).all()

    # Calcular estatísticas
    total_produtos = len(produtos)
    total_variacoes = sum(len(p.items_armazem) for p in produtos)
    total_esgotados = sum(1 for p in produtos for i in p.items_armazem if i.stock == 0)

    return render_template('stock_fardamento.html',
                           produtos=produtos,
                           tipos=tipos,
                           total_produtos=total_produtos,
                           total_variacoes=total_variacoes,
                           total_esgotados=total_esgotados)

#-----------Editar Stock Fardamento--------------

@app.route('/stock-fardamento/editar/<int:id>', methods=['POST'])
@login_required
def editar_stock_fardamento(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))

    item = StockFardamento.query.get_or_404(id)
    item.tipo = request.form.get('tipo', 'Outro')
    item.nome = request.form['nome']
    item.descricao = request.form.get('descricao', '')
    item.tamanho = request.form.get('tamanho', '')
    item.stock = request.form.get('stock', 0, type=int)
    db.session.commit()
    flash('Item atualizado.', 'success')
    return redirect(url_for('stock_fardamento'))

#-----------Editar Stock Fardamento--------------
@app.route('/stock-fardamento/apagar/<int:id>')
@login_required
def apagar_stock_fardamento(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))

    item = StockFardamento.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    flash('Item removido.', 'info')
    return redirect(url_for('stock_fardamento'))


@app.route('/stock-fardamento/editar-produto/<string:codigo_farda>', methods=['POST'])
@login_required
def editar_produto_fardamento(codigo_farda):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    produto = StockFardamento.query.filter_by(codigo_farda=codigo_farda).first_or_404()
    produto.tipo = request.form.get('tipo', 'Outro')
    produto.nome = request.form['nome']
    produto.descricao = request.form.get('descricao', '')
    db.session.commit()
    flash('Produto atualizado.', 'success')
    return redirect(url_for('stock_fardamento'))


@app.route('/stock-fardamento/editar-item/<int:id>', methods=['POST'])
@login_required
def editar_item_stock_fardamento(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    item = StockFardamentoArmazem.query.get_or_404(id)
    item.tipo = request.form.get('tipo', 'Outro')
    item.nome = request.form['nome']
    item.descricao = request.form.get('descricao', '')
    item.tamanho = request.form.get('tamanho', '')
    item.stock = request.form.get('stock', 0, type=int)
    db.session.commit()
    flash('Item atualizado.', 'success')
    return redirect(url_for('stock_fardamento'))


@app.route('/stock-fardamento/apagar-item/<int:id>')
@login_required
def apagar_item_stock_fardamento(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    item = StockFardamentoArmazem.query.get_or_404(id)
    codigo = item.codigo_farda
    db.session.delete(item)
    db.session.commit()
    # Se não restarem itens, apagar o produto principal
    restantes = StockFardamentoArmazem.query.filter_by(codigo_farda=codigo).count()
    if restantes == 0:
        produto = StockFardamento.query.filter_by(codigo_farda=codigo).first()
        if produto:
            db.session.delete(produto)
            db.session.commit()
    flash('Item removido.', 'info')
    return redirect(url_for('stock_fardamento'))


@app.route('/stock-fardamento/exportar')
@login_required
def exportar_stock_fardamento():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    produtos = StockFardamento.query.order_by(StockFardamento.codigo_farda.asc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Fardamento"
    cabecalhos = ['Código Produto', 'Tipo', 'Nome', 'Descrição', 'Sub-código', 'Tamanho', 'Stock']
    ws.append(cabecalhos)
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
    for p in produtos:
        for item in p.items_armazem:
            ws.append([p.codigo_farda, p.tipo, p.nome, p.descricao or '',
                       item.sub_codigo_farda, item.tamanho or '', item.stock])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='stock_fardamento.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/stock-fardamento/imprimir')
@login_required
def imprimir_stock_fardamento():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    produtos = StockFardamento.query.order_by(StockFardamento.codigo_farda.asc()).all()
    return render_template('imprimir_stock_fardamento.html', produtos=produtos, now=date.today())


@app.route('/stock-fardamento/importar', methods=['POST'])
@login_required
def importar_stock_fardamento():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('stock_fardamento'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('stock_fardamento'))
    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('stock_fardamento'))
    linhas_importadas = 0
    erros = []
    # Colunas: Código Produto, Tipo, Nome, Descrição, Sub-código, Tamanho, Stock
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        try:
            codigo_farda = str(row[0]).strip() if row[0] else ''
            tipo = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            nome = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            descricao = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            sub_codigo = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            tamanho = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            stock = int(row[6]) if len(row) > 6 and row[6] is not None else 0
        except Exception:
            erros.append(f'Linha {row_num}: dados inválidos.')
            continue
        if not codigo_farda or not nome:
            erros.append(f'Linha {row_num}: código e nome obrigatórios.')
            continue
        produto = StockFardamento.query.filter_by(codigo_farda=codigo_farda).first()
        if not produto:
            produto = StockFardamento(
                codigo_farda=codigo_farda,
                tipo=tipo,
                nome=nome,
                descricao=descricao
            )
            db.session.add(produto)
            db.session.flush()
        if StockFardamentoArmazem.query.filter_by(sub_codigo_farda=sub_codigo).first():
            erros.append(f'Linha {row_num}: sub-código {sub_codigo} já existe. Ignorado.')
            continue
        item = StockFardamentoArmazem(
            codigo_farda=codigo_farda,
            sub_codigo_farda=sub_codigo,
            tipo=tipo,
            nome=nome,
            descricao=descricao,
            tamanho=tamanho,
            stock=stock
        )
        db.session.add(item)
        linhas_importadas += 1
    db.session.commit()
    if erros:
        flash(f'{linhas_importadas} itens importados. {len(erros)} erro(s): ' + '; '.join(erros[:5]), 'warning')
    else:
        flash(f'{linhas_importadas} itens importados com sucesso.', 'success')
    return redirect(url_for('stock_fardamento'))


@app.route('/stock-fardamento/api/itens')
@login_required
def api_stock_fardamento_itens():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        return jsonify({'error': 'Acesso restrito'}), 403

    itens = StockFardamentoArmazem.query.filter(StockFardamentoArmazem.stock > 0).order_by(
        StockFardamentoArmazem.tipo.asc(),
        StockFardamentoArmazem.nome.asc(),
        StockFardamentoArmazem.tamanho.asc()
    ).all()

    resultado = []
    for item in itens:
        resultado.append({
            'id': item.id,
            'codigo_farda': item.codigo_farda,
            'sub_codigo_farda': item.sub_codigo_farda,
            'tipo': item.tipo,
            'nome': item.nome,
            'descricao': item.descricao,
            'tamanho': item.tamanho,
            'stock': item.stock
        })
    return jsonify(resultado)




@app.route('/tipos-farda-material/adicionar', methods=['POST'])
@login_required
def adicionar_tipo_farda_material():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    nome = request.form['nome'].strip()
    categoria = request.form.get('categoria', 'Farda')
    if nome:
        if not TipoFardaMaterial.query.filter_by(nome=nome, categoria=categoria).first():
            novo = TipoFardaMaterial(nome=nome, categoria=categoria)
            db.session.add(novo)
            db.session.commit()
            flash('Tipo adicionado.', 'success')
        else:
            flash('Tipo já existe.', 'warning')
    return redirect(url_for('stock_fardamento'))


@app.route('/tipos-farda-material/apagar/<int:id>')
@login_required
def apagar_tipo_farda_material(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_fardamento'))
    tipo = TipoFardaMaterial.query.get_or_404(id)
    db.session.delete(tipo)
    db.session.commit()
    flash('Tipo removido.', 'info')
    return redirect(url_for('stock_fardamento'))




#------------Farmaneto Atribuido-------------
@app.route('/fardamento-atribuido', methods=['GET', 'POST'])
@login_required
def fardamento_atribuido():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        bombeiro_id = request.form['bombeiro_id']
        tipo = request.form['tipo']
        nome = request.form['nome']
        tamanho = request.form['tamanho']
        data_entrega = datetime.strptime(request.form['data_entrega'], '%Y-%m-%d').date()

        novo = FardamentoAtribuido(
            bombeiro_id=bombeiro_id,
            tipo=tipo,
            nome=nome,
            tamanho=tamanho,
            data_entrega=data_entrega,
            estado='Entregue'
        )
        db.session.add(novo)
        db.session.commit()
        flash('Fardamento atribuído.', 'success')
        return redirect(url_for('fardamento_atribuido'))

    bombeiro_id = request.args.get('bombeiro_id', type=int)
    query = FardamentoAtribuido.query
    if bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)
    atribuicoes = query.order_by(FardamentoAtribuido.data_entrega.desc()).all()
    bombeiros = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()

    return render_template('fardamento_atribuido.html',
                           atribuicoes=atribuicoes,
                           bombeiros=bombeiros,
                           bombeiro_id=bombeiro_id)

#------------------Imprimir Fardamento


@app.route('/fardamento/imprimir-bombeiro')
@login_required
def imprimir_fardamento_bombeiro():
    bombeiro_id = request.args.get('bombeiro_id', type=int)
    if bombeiro_id:
        atribuicoes = FardamentoAtribuido.query.filter_by(
            bombeiro_id=bombeiro_id,
            estado='Entregue'
        ).order_by(FardamentoAtribuido.data_entrega.desc()).all()
        bombeiro = Bombeiro.query.get(bombeiro_id)
    else:
        atribuicoes = []
        bombeiro = None

    return render_template('imprimir_fardamento_bombeiro.html',
                           atribuicoes=atribuicoes,
                           bombeiro=bombeiro)

#------------------toggle de aprovação--------------

@app.route('/fardamento/toggle/<int:id>/<campo>')
@login_required
def toggle_aprovacao_fardamento(id, campo):
    if campo not in ('responsavel', 'comando'):
        flash('Campo inválido.', 'danger')
        return redirect(url_for('fardamento', tab='pedidos'))

    pedido = Fardamento.query.get_or_404(id)

    # Permissões
    if campo == 'comando':
        if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
            flash('Apenas Admin/Comando podem alterar este campo.', 'danger')
            return redirect(url_for('fardamento', tab='pedidos'))
        pedido.comando = not pedido.comando
    elif campo == 'responsavel':
        if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ('Comando', 'Fardamento'):
            flash('Sem permissão.', 'danger')
            return redirect(url_for('fardamento', tab='pedidos'))
        pedido.responsavel = not pedido.responsavel

    # Se ambos os vistos estão verdes, o pedido passa para "Análise"
    if pedido.responsavel and pedido.comando:
        pedido.estado = 'Análise'

    db.session.commit()
    return redirect(url_for('fardamento', tab='pedidos'))


# -----------------Resposta Fardamento--------

@app.route('/fardamento/resposta/<int:id>', methods=['POST'])
@login_required
def resposta_fardamento(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ('Comando', 'Fardamento'):
        flash('Sem permissão.', 'danger')
        return redirect(url_for('fardamento', tab='pedidos'))

    pedido = Fardamento.query.get_or_404(id)
    acao = request.form.get('acao')
    motivo = request.form.get('motivo', '')

    if acao == 'aguardar_stock':
        pedido.estado = 'Análise'
        flash('Pedido marcado como "Aguardar Stock".', 'info')
    elif acao == 'rejeitar':
        pedido.estado = 'Rejeitado'
        pedido.descricao_motivo = motivo
        db.session.flush()  # para ter a certeza que o pedido fica gravado antes de criar a mensagem

        # Enviar mensagem automática para o bombeiro
        msg = MensagemCorreio(
            remetente_id=current_user.id,
            destinatario_id=pedido.bombeiro_id,
            departamento=None,
            assunto='Pedido de fardamento rejeitado',
            corpo=f'O seu pedido de fardamento (ID {pedido.id}) foi rejeitado.\n\nMotivo: {motivo}',
            data_envio=datetime.utcnow(),
            lida=False,
            apagada_remetente=False,
            apagada_destinatario=False
        )
        db.session.add(msg)
        flash('Pedido rejeitado e mensagem enviada ao bombeiro.', 'warning')
    else:
        flash('Ação inválida.', 'danger')
        return redirect(url_for('fardamento', tab='pedidos'))

    db.session.commit()
    return redirect(url_for('fardamento', tab='pedidos'))


def gerar_codigo_stock_farmacia():
    ultimo = StockFarmacia.query.order_by(StockFarmacia.codigo.desc()).first()
    if ultimo and ultimo.codigo and ultimo.codigo.startswith('SF'):
        try:
            num = int(ultimo.codigo[2:]) + 1
        except:
            num = 1
    else:
        num = 1
    return f"SF{num:04d}"


#----------- Stock Farmácia--------------
@app.route('/stock-farmacia', methods=['GET', 'POST'])
@login_required
def stock_farmacia():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        categoria = request.form['categoria']
        nome = request.form['nome']
        tamanho = request.form.get('tamanho', '')
        stock = request.form.get('stock', 0, type=int)
        infstock = request.form.get('infstock', 0, type=int)
        data_validade_str = request.form.get('data_validade', '')

        data_validade = None
        if data_validade_str:
            try:
                data_validade = datetime.strptime(data_validade_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        codigo = gerar_codigo_stock_farmacia()

        novo = StockFarmacia(
            codigo=codigo,
            categoria=categoria,
            nome=nome,
            tamanho=tamanho if tamanho else None,
            stock=stock,
            infstock=infstock,
            data_validade=data_validade,
            data_atualizacao=datetime.utcnow()
        )
        db.session.add(novo)
        db.session.flush()   # para garantir que o ID é gerado (não é estritamente necessário)

        # Criar registo correspondente na Farmácia Central
        central = FarmaciaCentral(
            codigo=codigo,
            categoria=categoria,
            nome=nome,
            tamanho=tamanho if tamanho else None,
            stock=0,                      # stock inicial na central é 0
            stock_minimo=5,
            data_validade=data_validade
        )
        db.session.add(central)

        db.session.commit()
        flash(f'Produto {codigo} adicionado ao Stock Farmácia e à Farmácia Central.', 'success')
        return redirect(url_for('stock_farmacia'))

    # GET (mantém igual)
    itens = StockFarmacia.query.order_by(StockFarmacia.categoria, StockFarmacia.nome).all()
    categorias = CategoriaFarmacia.query.order_by(CategoriaFarmacia.nome).all()
    return render_template('stock_farmacia.html', itens=itens, categorias=categorias)

@app.route('/stock-farmacia/editar/<int:id>', methods=['POST'])
@login_required
def editar_stock_farmacia(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_farmacia'))

    item = StockFarmacia.query.get_or_404(id)
    item.categoria = request.form['categoria']
    item.nome = request.form['nome']
    item.tamanho = request.form.get('tamanho', '')
    novo_stock = request.form.get('stock', 0, type=int)
    novo_infstock = request.form.get('infstock', 0, type=int)
    data_validade_str = request.form.get('data_validade', '')

    if data_validade_str:
        try:
            item.data_validade = datetime.strptime(data_validade_str, '%Y-%m-%d').date()
        except ValueError:
            item.data_validade = None
    else:
        item.data_validade = None

    if novo_stock != item.stock or novo_infstock != item.infstock:
        item.data_atualizacao = datetime.utcnow()

    item.stock = novo_stock
    item.infstock = novo_infstock
    db.session.commit()

    # ----- Sincronizar com a Farmácia Central (mesmo código) -----
    central = FarmaciaCentral.query.filter_by(codigo=item.codigo).first()
    if central:
        central.categoria = item.categoria
        central.nome = item.nome
        central.tamanho = item.tamanho
        central.data_validade = item.data_validade
        central.ultima_atualizacao = datetime.utcnow()
        # NOTA: O stock da central NÃO é alterado aqui – apenas os dados descritivos
        db.session.commit()
    # ------------------------------------------------------------

    flash('Produto atualizado.', 'success')
    return redirect(url_for('stock_farmacia'))

@app.route('/stock-farmacia/apagar/<int:id>')
@login_required
def apagar_stock_farmacia(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_farmacia'))

    item = StockFarmacia.query.get_or_404(id)
    codigo = item.codigo

    # ----- Apagar também da Farmácia Central (se existir) -----
    central = FarmaciaCentral.query.filter_by(codigo=codigo).first()
    if central:
        db.session.delete(central)
    # ---------------------------------------------------------

    db.session.delete(item)
    db.session.commit()
    flash('Produto removido.', 'info')
    return redirect(url_for('stock_farmacia'))


@app.route('/farmacia-central')
@login_required
def listar_farmacia_central():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    # Obter parâmetros da query string
    categoria_filtro = request.args.get('categoria', '')
    pesquisa = request.args.get('pesquisa', '').strip()

    # Query base
    query = FarmaciaCentral.query

    # Filtro por categoria (correspondência exata)
    if categoria_filtro:
        query = query.filter(FarmaciaCentral.categoria == categoria_filtro)

    # Filtro por nome (case‑insensitive + ignorando acentos, usando unaccent)
    if pesquisa:
        # Remove acentos tanto da coluna como do termo de pesquisa
        query = query.filter(
            func.unaccent(FarmaciaCentral.nome).ilike(f'%{pesquisa}%')
        )

    # Ordenação por código ascendente
    itens = query.order_by(FarmaciaCentral.codigo.asc()).all()

    # Lista de categorias distintas para o combobox (já ordenada)
    categorias = db.session.query(FarmaciaCentral.categoria).distinct() \
                           .order_by(FarmaciaCentral.categoria).all()
    categorias = [c[0] for c in categorias]  # converte tuplas em lista simples

    return render_template('farmacia_central.html',
                           itens=itens,
                           categorias=categorias,
                           categoria_filtro=categoria_filtro,
                           pesquisa=pesquisa)



@app.route('/categorias-farmacia', methods=['GET', 'POST'])
@login_required
def categorias_farmacia():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nome = request.form['nome'].strip()
        checklist = request.form.get('checklist') == 'on'   # ← captura a checkbox

        if not nome:
            flash('Nome da categoria obrigatório.', 'warning')
        elif CategoriaFarmacia.query.filter_by(nome=nome).first():
            flash('Categoria já existe.', 'warning')
        else:
            nova = CategoriaFarmacia(nome=nome, checklist=checklist)
            db.session.add(nova)
            db.session.commit()
            flash('Categoria criada.', 'success')
        return redirect(url_for('categorias_farmacia'))

    # GET – listar categorias
    categorias = CategoriaFarmacia.query.order_by(CategoriaFarmacia.nome).all()
    return render_template('categorias_farmacia.html', categorias=categorias)


def verificar_stock_minimo(produto):
    if produto.infstock is None or produto.infstock == 0:
        return
    if produto.stock > produto.infstock:
        return

    # Destinatários: apenas Comando e responsáveis da Farmácia
    destinatarios = Bombeiro.query.filter(
        (Bombeiro.resp_departamento == 'Comando') |
        (func.lower(Bombeiro.resp_departamento) == 'farmacia'),
        Bombeiro.ativo == True
    ).all()

    # Remetente: utilizador "Sistema"
    remetente = Bombeiro.query.filter_by(mecanografico='SISTEMA').first()
    if not remetente:
        return

    corpo = (
        f"⚠️ Alerta de Stock Mínimo\n\n"
        f"O produto '{produto.nome}' (categoria: {produto.categoria}) "
        f"atingiu o stock mínimo definido ({produto.infstock} unidades).\n"
        f"Stock actual: {produto.stock} unidades.\n\n"
        f"Por favor, verifique e reponha o stock com urgência."
    )

    for dest in destinatarios:
        msg = MensagemCorreio(
            remetente_id=remetente.id,
            destinatario_id=dest.id,
            assunto=f'⚠️ Stock mínimo atingido – {produto.nome}',
            corpo=corpo,
            data_envio=datetime.utcnow(),
            lida=False,
            apagada_remetente=False,
            apagada_destinatario=False
        )
        db.session.add(msg)
    db.session.commit()


#----------- Exportar Stock Farmácia--------------
@app.route('/stock-farmacia/exportar')
@login_required
def exportar_stock_farmacia():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_farmacia'))

    itens = StockFarmacia.query.order_by(StockFarmacia.categoria, StockFarmacia.nome).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Farmacia"

    cabecalhos = ['ID', 'Categoria', 'Nome', 'Tamanho', 'Stock', 'Atualização']
    ws.append(cabecalhos)
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for i in itens:
        ws.append([i.id, i.categoria, i.nome, i.tamanho or '', i.stock,
                   i.data_atualizacao.strftime('%d/%m/%Y %H:%M') if i.data_atualizacao else ''])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='stock_farmacia.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#----------- Importar Stock Farmácia--------------
@app.route('/stock-farmacia/importar', methods=['POST'])
@login_required
def importar_stock_farmacia():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_farmacia'))

    if 'ficheiro' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('stock_farmacia'))
    ficheiro = request.files['ficheiro']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido.', 'danger')
        return redirect(url_for('stock_farmacia'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
        ws = wb.active
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('stock_farmacia'))

    linhas_importadas = 0
    erros = []

    # Assumindo que o cabeçalho tem as colunas na seguinte ordem:
    # 0: Código, 1: Categoria, 2: Nome, 3: Tamanho, 4: Stock, 5: Stock Mínimo, 6: Data Validade, 7: Atualização (ignorada)
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue

        # Extrair valores (com segurança)
        try:
            codigo = str(row[0]).strip() if len(row) > 0 and row[0] else ''
            categoria = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            nome = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            tamanho = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            stock = int(row[4]) if len(row) > 4 and row[4] is not None else 0
            infstock = int(row[5]) if len(row) > 5 and row[5] is not None else 0
            data_validade_str = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        except Exception as e:
            erros.append(f'Linha {row_num}: erro ao interpretar dados - {str(e)}')
            continue

        # Validação de campos obrigatórios
        if not categoria or not nome:
            erros.append(f'Linha {row_num}: categoria e nome são obrigatórios.')
            continue

        # Se o código não for fornecido, gerar automaticamente
        if not codigo:
            codigo = gerar_codigo_stock_farmacia()
        else:
            # Verificar se o código já existe na base de dados
            if StockFarmacia.query.filter_by(codigo=codigo).first():
                erros.append(f'Linha {row_num}: código {codigo} já existe. Ignorado.')
                continue

        # Converter data de validade
        data_validade = None
        if data_validade_str and data_validade_str.lower() != 'none':
            try:
                data_validade = datetime.strptime(data_validade_str, '%d/%m/%Y').date()
            except ValueError:
                try:
                    data_validade = datetime.strptime(data_validade_str, '%Y-%m-%d').date()
                except ValueError:
                    erros.append(f'Linha {row_num}: data de validade inválida "{data_validade_str}". Ignorada.')
                    data_validade = None

        # Criar novo produto
        novo = StockFarmacia(
            codigo=codigo,
            categoria=categoria,
            nome=nome,
            tamanho=tamanho if tamanho else None,
            stock=stock,
            infstock=infstock,
            data_validade=data_validade,
            data_atualizacao=datetime.utcnow()
        )
        db.session.add(novo)
        linhas_importadas += 1

    db.session.commit()

    if erros:
        flash(f'{linhas_importadas} produtos importados. {len(erros)} erro(s): ' + '; '.join(erros[:5]), 'warning')
    else:
        flash(f'{linhas_importadas} produtos importados com sucesso!', 'success')
    return redirect(url_for('stock_farmacia'))



#------------Imprimir Stock Farmacia----------
@app.route('/stock-farmacia/imprimir')
@login_required
def imprimir_stock_farmacia():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_farmacia'))

    itens = StockFarmacia.query.order_by(StockFarmacia.categoria, StockFarmacia.nome).all()
    return render_template('imprimir_stock_farmacia.html', itens=itens)



#----------- Categorias Farmácia--------------
@app.route('/categorias-farmacia/apagar/<int:id>')
@login_required
def apagar_categoria_farmacia(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('categorias_farmacia'))
    cat = CategoriaFarmacia.query.get_or_404(id)
    db.session.delete(cat)
    db.session.commit()
    flash('Categoria removida.', 'info')
    return redirect(url_for('categorias_farmacia'))


#----------- CheckList Ambulancia--------------

@app.route('/checklist-ambulancia', methods=['GET', 'POST'])
@login_required
def checklist_ambulancia():
    if request.method == 'POST':
        viatura_id = request.form['viatura_id']
        # Criar a checklist com o utilizador autenticado
        nova = ChecklistAmbulancia(viatura_id=viatura_id, bombeiro_id=current_user.id)
        db.session.add(nova)
        db.session.flush()

        # Produtos das categorias com checklist ativo
        categorias_check = CategoriaFarmacia.query.filter_by(checklist=True).all()
        produtos_disponiveis = []
        if categorias_check:
            nomes_categorias = [c.nome.lower() for c in categorias_check]
            produtos_disponiveis = StockFarmacia.query.filter(
                func.lower(StockFarmacia.categoria).in_(nomes_categorias)
            ).all()

        ids_selecionados = []
        for produto in produtos_disponiveis:
            if request.form.get(f'prod_{produto.id}') == 'on':
                ids_selecionados.append(produto.id)

        if not ids_selecionados:
            db.session.rollback()
            flash('Selecione pelo menos um produto.', 'warning')
            return redirect(url_for('checklist_ambulancia'))

        for pid in ids_selecionados:
            item = ChecklistAmbulanciaItem(checklist_id=nova.id, produto_id=pid, quantidade=0)
            db.session.add(item)

        db.session.commit()
        return redirect(url_for('preencher_quantidades', checklist_id=nova.id))

    # ---------- GET ----------
    viatura_id_filtro = request.args.get('viatura_id', type=int)
    query = ChecklistAmbulancia.query

    # Utilizador normal só vê as suas próprias checklists
    if current_user.tipo_user == 'User' and current_user.resp_departamento not in ['Comando', 'Farmacia', 'Socorrista']:
        query = query.filter(ChecklistAmbulancia.bombeiro_id == current_user.id)

    if viatura_id_filtro:
        query = query.filter_by(viatura_id=viatura_id_filtro)

    checklists = query.order_by(ChecklistAmbulancia.data_hora.desc()).all()
    viaturas = Viatura.query.order_by(Viatura.matricula).all()

    # Dados para o modal de criação (produtos das categorias com checklist ativo)
    categorias = CategoriaFarmacia.query.filter_by(checklist=True).order_by(CategoriaFarmacia.nome).all()
    produtos_por_categoria = []
    for cat in categorias:
        produtos = StockFarmacia.query.filter(
            func.lower(StockFarmacia.categoria) == cat.nome.lower()
        ).order_by(StockFarmacia.nome).all()
        if produtos:
            produtos_por_categoria.append((cat, produtos))

    return render_template('checklist_ambulancia.html',
                           checklists=checklists,
                           viaturas=viaturas,
                           produtos_por_categoria=produtos_por_categoria,
                           viatura_selecionada=viatura_id_filtro)


@app.route('/checklist-ambulancia/quantidades/<int:checklist_id>', methods=['GET', 'POST'])
@login_required
def preencher_quantidades(checklist_id):
    checklist = ChecklistAmbulancia.query.get_or_404(checklist_id)
    if checklist.finalizado:
        flash('Checklist já finalizada.', 'warning')
        return redirect(url_for('checklist_ambulancia'))

    if request.method == 'POST':
        # Atualizar quantidades nos itens da checklist e criar StockAmbulancia (pendente)
        for item in checklist.itens:
            qtd_str = request.form.get(f'qtd_{item.id}', '0')
            try:
                qtd = int(qtd_str)
            except ValueError:
                qtd = 0
            item.quantidade = qtd

            if qtd > 0:
                # Criar pedido no Stock Ambulância (pendente, sem abater)
                reposicao = StockAmbulancia(
                    ambulancia_id=checklist.viatura_id,
                    produto_id=item.produto_id,
                    quantidade=qtd,
                    solicitante_id=current_user.id,
                    responsavel_id=None,
                    checklist_id=checklist.id,
                    confirmado=False
                )
                db.session.add(reposicao)

        checklist.finalizado = True
        db.session.commit()
        flash('Quantidades registadas e pedidos enviados para Stock Ambulância.', 'success')
        return redirect(url_for('checklist_ambulancia'))


    # GET – mostrar formulário com os itens selecionados
    # Responsáveis: bombeiros cujo departamento é 'Farmácia'
    responsaveis = Bombeiro.query.filter_by(resp_departamento='Farmácia', ativo=True).all()
    return render_template('preencher_quantidades.html', checklist=checklist, responsaveis=responsaveis)

@app.route('/checklist-ambulancia/imprimir-lista')
@login_required
def imprimir_lista_checklists():
    viatura_id = request.args.get('viatura_id', type=int)
    query = ChecklistAmbulancia.query
    if viatura_id:
        query = query.filter_by(viatura_id=viatura_id)
    checklists = query.order_by(ChecklistAmbulancia.data_hora.desc()).all()
    return render_template('imprimir_lista_checklists.html', checklists=checklists)


@app.route('/checklist-ambulancia/imprimir/<int:checklist_id>')
@login_required
def imprimir_checklist(checklist_id):
    checklist = ChecklistAmbulancia.query.get_or_404(checklist_id)
    return render_template('imprimir_checklist_ambulancia.html', checklist=checklist)


#----------- Stock Ambulancia--------------

@app.route('/stock-ambulancia')
@login_required
def stock_ambulancia():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia', 'Socorrista', 'Central']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    reposicoes = StockAmbulancia.query.order_by(StockAmbulancia.data.desc()).all()
    return render_template('stock_ambulancia.html', reposicoes=reposicoes)



@app.route('/stock-ambulancia/confirmar/<int:id>')
@login_required
def confirmar_reposicao(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia', 'Socorrista', 'Central']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('stock_ambulancia'))

    reposicao = StockAmbulancia.query.get_or_404(id)
    if reposicao.confirmado:
        flash('Reposição já confirmada.', 'warning')
        return redirect(url_for('stock_ambulancia'))

    # Obter o produto principal (StockFarmacia) através da relação
    produto_principal = reposicao.produto   # relação definida em StockAmbulancia.produto
    if not produto_principal:
        flash('Produto não encontrado.', 'danger')
        return redirect(url_for('stock_ambulancia'))

    # Buscar o correspondente na Farmácia Central
    central = FarmaciaCentral.query.filter_by(codigo=produto_principal.codigo).first()
    if not central:
        flash('Produto não encontrado na Farmácia Central.', 'danger')
        return redirect(url_for('stock_ambulancia'))

    if central.stock < reposicao.quantidade:
        flash(f'Stock insuficiente na Farmácia Central para "{central.nome}". Disponível: {central.stock}.', 'danger')
        return redirect(url_for('stock_ambulancia'))

    # Abate da Farmácia Central
    central.stock -= reposicao.quantidade
    central.ultima_atualizacao = datetime.utcnow()

    reposicao.confirmado = True
    reposicao.responsavel_id = current_user.id
    db.session.commit()

    flash('Reposição confirmada e stock da Farmácia Central atualizado.', 'success')
    return redirect(url_for('stock_ambulancia'))


@app.route('/farmacia-central/transferir', methods=['POST'])
@login_required
def transferir_para_central():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Farmacia']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    codigo = request.form.get('codigo')
    quantidade = request.form.get('quantidade', type=int)

    produto_principal = StockFarmacia.query.filter_by(codigo=codigo).first()
    if not produto_principal or produto_principal.stock < quantidade:
        flash('Stock insuficiente na Farmácia Principal.', 'danger')
        return redirect(url_for('listar_farmacia_central'))

    central = FarmaciaCentral.query.filter_by(codigo=codigo).first()
    if not central:
        flash('Produto não encontrado na Farmácia Central.', 'danger')
        return redirect(url_for('listar_farmacia_central'))

    # Abate do StockFarmacia
    produto_principal.stock -= quantidade
    produto_principal.data_atualizacao = datetime.utcnow()

    # Adiciona à Farmácia Central
    central.stock += quantidade
    central.ultima_atualizacao = datetime.utcnow()

    db.session.commit()
    flash(f'Transferidos {quantidade} unidade(s) de {produto_principal.nome} para a Farmácia Central.', 'success')
    return redirect(url_for('listar_farmacia_central'))



#_____________________Central_____________

@app.route('/central')
@login_required
def central():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Central']:
        flash('Acesso restrito ao Departamento Central.', 'danger')
        return redirect(url_for('dashboard'))

    hoje = date.today()
    daqui_7_dias = hoje + timedelta(days=7)

    # Totais de escalados hoje por categoria
    categorias = ['Motorista', 'Socorrista', 'Centralista', 'EIP', 'ECIN', 'ELAC', 'Piquete', 'Bombeiro']
    totais = {}
    for cat in categorias:
        count = Escala.query.filter(
            func.date(Escala.data_inicio) <= hoje,
            func.date(Escala.data_fim) >= hoje,
            Escala.categoria == cat
        ).distinct(Escala.bombeiro_id).count()
        totais[cat] = count

    viaturas_inop = Viatura.query.filter(
        func.lower(Viatura.estado) == 'inoperacional'
    ).order_by(Viatura.matricula).all()
    notas_7dias = Nota.query.filter(
        Nota.data_evento >= hoje,
        Nota.data_evento <= daqui_7_dias
    ).order_by(Nota.data_evento.asc()).all()

    bombeiros = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()
    departamentos = list(set(b.resp_departamento for b in bombeiros if b.resp_departamento))

    return render_template('central.html',
                           hoje=hoje,
                           totais=totais,
                           viaturas_inop=viaturas_inop,
                           notas_7dias=notas_7dias,
                           bombeiros=bombeiros,
                           departamentos=departamentos)



@app.route('/central/dia/<string:data_str>')
@login_required
def central_dia(data_str):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Central':
        return jsonify({'erro': 'Acesso restrito'}), 403

    try:
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'erro': 'Formato de data inválido'}), 400

    categoria = request.args.get('categoria', '')  # novo parâmetro

    html = gerar_html_dia(data, categoria)
    return jsonify({'html': html})

@app.route('/central/totais/<string:data_str>')
@login_required
def central_totais_dia(data_str):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Central':
        return jsonify({'erro': 'Acesso restrito'}), 403
    try:
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'erro': 'Data inválida'}), 400

    categorias = ['Motorista', 'Socorrista', 'Centralista', 'EIP', 'ECIN', 'ELAC', 'Piquete', 'Bombeiro']
    totais = {}
    for cat in categorias:
        count = Escala.query.filter(
            func.date(Escala.data_inicio) <= data,
            func.date(Escala.data_fim) >= data,
            Escala.categoria == cat
        ).distinct(Escala.bombeiro_id).count()
        totais[cat] = count
    return jsonify(totais)


@app.route('/central/notas', methods=['POST'])
@login_required
def central_notas():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Central':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    descricao = request.form['descricao']
    data_evento_str = request.form.get('data_evento')
    data_evento = datetime.strptime(data_evento_str, '%Y-%m-%d').date() if data_evento_str else None

    nota = Nota(
        criador_id=current_user.id,
        descricao=descricao,
        data_evento=data_evento
    )
    db.session.add(nota)
    db.session.commit()
    flash('Nota registada com sucesso.', 'success')
    return redirect(url_for('central', tab='notas'))

@app.route('/central/notas/editar/<int:id>', methods=['POST'])
@login_required
def central_editar_nota(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Central':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('central', tab='notas'))
    nota = Nota.query.get_or_404(id)
    nota.descricao = request.form['descricao']
    data_evento_str = request.form.get('data_evento')
    nota.data_evento = datetime.strptime(data_evento_str, '%Y-%m-%d').date() if data_evento_str else None
    db.session.commit()
    flash('Nota atualizada.', 'success')
    return redirect(url_for('central', tab='notas'))

@app.route('/central/notas/apagar/<int:id>')
@login_required
def central_apagar_nota(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Central':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('central', tab='notas'))
    nota = Nota.query.get_or_404(id)
    db.session.delete(nota)
    db.session.commit()
    flash('Nota removida.', 'info')
    return redirect(url_for('central', tab='notas'))



@app.route('/central/atividade_mes/<int:ano>/<int:mes>')
@login_required
def central_atividade_mes(ano, mes):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Central':
        return jsonify({'erro': 'Acesso restrito'}), 403

    hoje = date.today()
    primeiro_dia = date(ano, mes, 1)
    ultimo_dia = date(ano, mes, calendar.monthrange(ano, mes)[1])

    atividade = {}
    for dia in range(1, ultimo_dia.day + 1):
        d = date(ano, mes, dia)
        tem_escalas = Escala.query.filter(
            func.date(Escala.data_inicio) <= d,
            func.date(Escala.data_fim) >= d
        ).first() is not None
        tem_trocas = TrocaServico.query.filter(
            (TrocaServico.data_origem == d) | (TrocaServico.data_destino == d),
            TrocaServico.estado == 'aprovada'
        ).first() is not None
        tem_dispensas = Dispensa.query.filter(
            Dispensa.data_inicio <= d,
            Dispensa.data_fim >= d,
            Dispensa.aprovada == True
        ).first() is not None

        atividade[str(d)] = {
            'escalas': tem_escalas,
            'trocas': tem_trocas,
            'dispensas': tem_dispensas
        }

    return jsonify(atividade)




#----------------Secção Correio---------------
@app.route('/correio')
@login_required
def correio():
    filtro = request.args.get('filtro', 'todas')  # 'todas', 'lidas', 'naolidas'

    # Caixa de entrada
    query_entrada = MensagemCorreio.query.filter(
        (
            (MensagemCorreio.destinatario_id == current_user.id) |
            (MensagemCorreio.departamento == current_user.resp_departamento)
        ),
        MensagemCorreio.remetente_id != current_user.id,
        MensagemCorreio.apagada_destinatario == False
    )

    if filtro == 'lidas':
        query_entrada = query_entrada.filter_by(lida=True)
    elif filtro == 'naolidas':
        query_entrada = query_entrada.filter_by(lida=False)

    caixa_entrada = query_entrada.order_by(MensagemCorreio.data_envio.desc()).all()

    # Enviadas
    enviadas = MensagemCorreio.query.filter(
        MensagemCorreio.remetente_id == current_user.id,
        MensagemCorreio.apagada_remetente == False
    ).order_by(MensagemCorreio.data_envio.desc()).all()

    bombeiros = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()
    departamentos = list(set(b.resp_departamento for b in bombeiros if b.resp_departamento))

    return render_template('correio.html',
                           caixa_entrada=caixa_entrada,
                           enviadas=enviadas,
                           bombeiros=bombeiros,
                           departamentos=departamentos,
                           filtro_atual=filtro)


@app.route('/correio/enviar', methods=['POST'])
@login_required
def correio_enviar():
    destinatario_tipo = request.form.get('destinatario_tipo')  # 'bombeiro' ou 'departamento'
    destinatario_id = request.form.get('destinatario_id', type=int)
    departamento = request.form.get('departamento') if destinatario_tipo == 'departamento' else None
    assunto = request.form.get('assunto', 'Sem assunto')
    corpo = request.form['corpo']

    nova = MensagemCorreio(
        remetente_id=current_user.id,
        destinatario_id=destinatario_id if destinatario_tipo == 'bombeiro' else None,
        departamento=departamento,
        assunto=assunto,
        corpo=corpo
    )
    db.session.add(nova)
    db.session.commit()
    flash('Mensagem enviada com sucesso.', 'success')
    return redirect(url_for('correio'))

@app.route('/correio/lida/<int:id>')
@login_required
def correio_marcar_lida(id):
    msg = MensagemCorreio.query.get_or_404(id)
    if msg.destinatario_id == current_user.id or msg.departamento == current_user.resp_departamento:
        msg.lida = True
        db.session.commit()
    return redirect(url_for('correio'))



@app.context_processor
def inject_novas_mensagens():
    if current_user.is_authenticated:
        naolidas = MensagemCorreio.query.filter(
            (
                (MensagemCorreio.destinatario_id == current_user.id) |
                (MensagemCorreio.departamento == current_user.resp_departamento)
            ),
            MensagemCorreio.remetente_id != current_user.id,
            MensagemCorreio.lida == False,
            MensagemCorreio.apagada_destinatario == False
        ).count()
    else:
        naolidas = 0
    return dict(msgs_naolidas=naolidas)

#------------------Importar 03-05-26----------------
# ========== FUNÇÕES AUXILIARES DE IMPORTAÇÃO ==========

def _importar_linha_bombeiros(row, row_num):
    numero = str(row[0]).strip() if row[0] else ''
    mecanografico = str(row[1]).strip() if len(row) > 1 and row[1] else ''
    nome = str(row[2]).strip() if len(row) > 2 and row[2] else ''
    nomecompleto = str(row[3]).strip() if len(row) > 3 and row[3] else ''
    email = str(row[4]).strip().lower() if len(row) > 4 and row[4] else ''
    telemovel = str(row[5]).strip() if len(row) > 5 and row[5] else None
    posto = str(row[6]).strip() if len(row) > 6 and row[6] else ''
    tipo_bombeiro = str(row[7]).strip() if len(row) > 7 and row[7] else 'Voluntário'
    departamento = str(row[8]).strip() if len(row) > 8 and row[8] else None
    tipo_user = str(row[9]).strip() if len(row) > 9 and row[9] else 'User'
    ativo = str(row[10]).strip().lower() == 'sim' if len(row) > 10 and row[10] else True
    password_hash = str(row[11]).strip() if len(row) > 11 and row[11] else generate_password_hash('123456')

    if not numero or not mecanografico or not nome or not email:
        return "Campos obrigatórios em falta (Nº Interno, Mecanográfico, Nome, Email)"
    if Bombeiro.query.filter(
        (Bombeiro.numero_interno == numero) |
        (Bombeiro.mecanografico == mecanografico) |
        (Bombeiro.email == email)
    ).first():
        return "Duplicado"
    b = Bombeiro(
        numero_interno=numero,
        mecanografico=mecanografico,
        nome=nome,
        nomecompleto=nomecompleto if nomecompleto else None,
        email=email,
        telemovel=telemovel if telemovel and telemovel != '' else None,
        posto=posto,
        tipo_bombeiro=tipo_bombeiro,
        resp_departamento=departamento if departamento and departamento != '' else None,
        tipo_user=tipo_user,
        ativo=ativo,
        password_hash=password_hash
    )
    db.session.add(b)
    db.session.flush()
    return None


def _importar_linha_viaturas(row, row_num):
    matricula = str(row[0]).strip() if row[0] else ''
    if not matricula:
        return "Matrícula obrigatória"
    if Viatura.query.filter_by(matricula=matricula).first():
        return "Matrícula já existe"
    tipo = str(row[1]).strip() if len(row) > 1 else ''
    nomenclatura = str(row[2]).strip() if len(row) > 2 else ''
    marca = str(row[3]).strip() if len(row) > 3 else ''
    modelo = str(row[4]).strip() if len(row) > 4 else ''
    ano = int(row[5]) if len(row) > 5 and row[5] else 0
    estado = str(row[6]).strip().lower() if len(row) > 6 else 'operacional'
    v = Viatura(matricula=matricula, tipo=tipo, nomenclatura=nomenclatura,
                marca=marca, modelo=modelo, ano=ano, estado=estado)
    db.session.add(v)
    db.session.flush()
    return None


def _importar_linha_categorias_farmacia(row, row_num):
    nome = str(row[0]).strip() if row[0] else ''
    if not nome:
        return "Nome da categoria obrigatório"
    if CategoriaFarmacia.query.filter_by(nome=nome).first():
        return "Categoria já existe"
    checklist = str(row[1]).strip().lower() == 'sim' if len(row) > 1 else False
    c = CategoriaFarmacia(nome=nome, checklist=checklist)
    db.session.add(c)
    db.session.flush()
    return None


def _importar_linha_stock_farmacia(row, row_num):
    try:
        cat = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        nome = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        tamanho_val = str(row[3]).strip()[:100] if len(row) > 3 and row[3] else ''
        # Tenta converter o stock, se falhar usa 0
        stock_val = 0
        if len(row) > 4 and row[4] is not None:
            try:
                stock_val = int(row[4])
            except (ValueError, TypeError):
                pass  # fica 0

        if not nome:
            return None

        s = StockFarmacia(
            categoria=cat,
            nome=nome,
            tamanho=tamanho_val if tamanho_val else None,
            stock=stock_val,
            data_atualizacao=datetime.utcnow()
        )
        return s
    except Exception:
        return None


def _importar_linha_stock_fardamento(row, row_num):
    try:
        id_original = None
        start_col = 0
        # Se a primeira coluna for um número inteiro, assumimos que é o ID
        if row[0] is not None:
            try:
                id_original = int(row[0])
                start_col = 1
            except (ValueError, TypeError):
                pass

        nome = str(row[start_col]).strip() if len(row) > start_col and row[start_col] else ''
        if not nome:
            return None

        descricao = str(row[start_col+1]).strip() if len(row) > start_col+1 and row[start_col+1] else ''
        tamanho = str(row[start_col+2]).strip() if len(row) > start_col+2 and row[start_col+2] else ''
        tipo = str(row[start_col+3]).strip() if len(row) > start_col+3 and row[start_col+3] else ''
        stock_val = int(row[start_col+4]) if len(row) > start_col+4 and row[start_col+4] else 0

        # Criar o objecto – se tiver ID, define-se explicitamente
        s = StockFardamento(
            id=id_original,
            nome=nome,
            descricao=descricao,
            tamanho=tamanho,
            tipo=tipo,
            stock=stock_val
        )
        return s
    except Exception:
        return None


def _importar_linha_ferias(row, row_num):
    """
    Espera uma linha com as colunas:
    0 - Mecanográfico do bombeiro
    1 - Data Início (dd/mm/aaaa)
    2 - Data Fim (dd/mm/aaaa)
    3 - Estado (Pendente/Aprovado/Rejeitado)
    4 - Nome do aprovador (ou vazio)
    5 - Data Pedido (dd/mm/aaaa HH:MM)
    """
    try:
        mecanografico = str(row[0]).strip() if row[0] else None
        inicio_str = str(row[1]).strip() if len(row) > 1 and row[1] else None
        fim_str = str(row[2]).strip() if len(row) > 2 and row[2] else None
        estado = str(row[3]).strip() if len(row) > 3 and row[3] else 'Pendente'
        nome_aprovador = str(row[4]).strip() if len(row) > 4 and row[4] else None
        data_pedido_str = str(row[5]).strip() if len(row) > 5 and row[5] else None

        if not mecanografico or not inicio_str or not fim_str:
            return None

        bombeiro = Bombeiro.query.filter_by(mecanografico=mecanografico).first()
        if not bombeiro:
            raise ValueError(f"Mecanográfico {mecanografico} não encontrado")

        data_inicio = _parse_data(inicio_str)
        data_fim = _parse_data(fim_str)
        if not data_inicio or not data_fim:
            raise ValueError("Data inválida")

        # Identificar o aprovador pelo nome (caso exista)
        aprovado_por = None
        if nome_aprovador:
            aprovador = Bombeiro.query.filter_by(nome=nome_aprovador).first()
            if aprovador:
                aprovado_por = aprovador.id

        data_pedido = _parse_datetime(data_pedido_str) or datetime.utcnow()

        ferias = Ferias(
            bombeiro_id=bombeiro.id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            estado=estado,
            aprovado_por=aprovado_por,
            data_pedido=data_pedido
        )
        return ferias
    except Exception as e:
        # A exceção será tratada no código chamador
        raise


def _importar_linha_disponibilidades(row, row_num):
    mec = str(row[0]).strip() if row[0] else None
    data_str = str(row[1]).strip() if len(row) > 1 else None
    turno_extra = str(row[2]).strip() if len(row) > 2 else ''
    categoria = str(row[3]).strip() if len(row) > 3 else ''
    confirmada = str(row[4]).strip().lower() == 'sim' if len(row) > 4 else False
    data = _parse_data(data_str) if data_str else None
    if not mec or not data:
        return "Mecanográfico ou data obrigatórios"
    bombeiro = Bombeiro.query.filter_by(mecanografico=mec).first()
    if not bombeiro:
        return f"Bombeiro {mec} não encontrado"
    d = Disponibilidade(bombeiro_id=bombeiro.id, data=data, turno_extra=turno_extra,
                        categoria=categoria, confirmada=confirmada)
    db.session.add(d)
    db.session.flush()
    return None


def _importar_linha_escalas(row, row_num):
    """Importa uma linha do Excel para a tabela Escala.
       Espera colunas: Mecanográfico | Nome | Início | Fim | Turno | Categoria | Função (opcional)
    """
    try:
        mecanografico = str(row[0]).strip() if row[0] else None
        # O nome não é usado directamente (é apenas para conferência)
        inicio_str = str(row[2]).strip() if len(row) > 2 and row[2] else None
        fim_str = str(row[3]).strip() if len(row) > 3 and row[3] else None
        turno = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        categoria = str(row[5]).strip() if len(row) > 5 and row[5] else 'Bombeiro'
        funcao = str(row[6]).strip() if len(row) > 6 and row[6] else None

        if not mecanografico or not inicio_str or not turno:
            return f"Linha {row_num}: Mecanográfico, data início ou turno em falta."

        bombeiro = Bombeiro.query.filter_by(mecanografico=mecanografico).first()
        if not bombeiro:
            return f"Linha {row_num}: Bombeiro com mecanográfico {mecanografico} não encontrado."

        # Converter a data/hora início
        try:
            # Tentar remover o ' 00:00:00' se existir, mas manter como datetime
            if ' ' in inicio_str:
                data_inicio = datetime.strptime(inicio_str, '%Y-%m-%d %H:%M:%S')
            else:
                data_inicio = datetime.strptime(inicio_str, '%Y-%m-%d')
        except ValueError:
            try:
                data_inicio = datetime.strptime(inicio_str, '%d/%m/%Y')
            except ValueError:
                return f"Linha {row_num}: Data início inválida '{inicio_str}'."

        # Determinar a hora de fim com base no turno (se a data fim não for fornecida ou for igual)
        # Usar a função turno_para_horas (já existente)
        inicio_str_hora, fim_str_hora = turno_para_horas(turno)
        # Converter as strings de hora em objetos time
        hora_inicio = datetime.strptime(inicio_str_hora, '%H:%M').time()
        hora_fim = datetime.strptime(fim_str_hora, '%H:%M').time()
        # Construir data_hora_inicio e data_hora_fim
        data_inicio = datetime.combine(data_inicio.date(), hora_inicio)
        data_fim = datetime.combine(data_inicio.date(), hora_fim)
        # Se a hora fim for menor ou igual à início, adiciona um dia
        if data_fim <= data_inicio:
            data_fim += timedelta(days=1)

        # Criar escala
        e = Escala(
            bombeiro_id=bombeiro.id,
            data_inicio=data_inicio,
            data_fim=data_fim,
            turno=turno,
            categoria=categoria,
            funcao=funcao if funcao and funcao != '' else None
        )
        db.session.add(e)
        return None  # sem erro

    except Exception as e:
        return f"Linha {row_num}: erro inesperado - {str(e)}"


def _importar_linha_avarias(row, row_num):
    codigo = str(row[0]).strip() if row[0] else None
    matricula = str(row[1]).strip() if len(row) > 1 else None
    descricao = str(row[2]).strip() if len(row) > 2 else ''
    reportador_mec = str(row[3]).strip() if len(row) > 3 else None
    kms = int(row[4]) if len(row) > 4 and row[4] else None
    resp_oficina = str(row[5]).strip().lower() == 'sim' if len(row) > 5 else False
    comando = str(row[6]).strip().lower() == 'sim' if len(row) > 6 else False
    estado = str(row[7]).strip() if len(row) > 7 else 'Pendente'
    data_str = str(row[8]).strip() if len(row) > 8 else None

    if not descricao or not matricula:
        return "Descrição e matrícula obrigatórias"
    viatura = Viatura.query.filter_by(matricula=matricula).first()
    if not viatura:
        return f"Viatura {matricula} não encontrada"
    reportador = Bombeiro.query.filter_by(mecanografico=reportador_mec).first() if reportador_mec else None
    data_reporte = _parse_datetime(data_str) or datetime.utcnow()

    a = Avaria(codigo=codigo, viatura_id=viatura.id,
               descricao=descricao, reportado_por=reportador.id if reportador else 1,
               kms=kms, responsavel_oficina=resp_oficina, comando_verificado=comando,
               estado=estado, data_reporte=data_reporte)
    db.session.add(a)
    db.session.flush()
    return None


def _importar_linha_trocas(row, row_num):
    mec_origem = str(row[0]).strip() if row[0] else None
    mec_destino = str(row[1]).strip() if len(row) > 1 else None
    data_origem_str = str(row[2]).strip() if len(row) > 2 else None
    data_destino_str = str(row[3]).strip() if len(row) > 3 else None
    motivo = str(row[4]).strip() if len(row) > 4 else ''
    estado = str(row[5]).strip() if len(row) > 5 else ''
    data_pedido_str = str(row[6]).strip() if len(row) > 6 else None

    b_orig = Bombeiro.query.filter_by(mecanografico=mec_origem).first() if mec_origem else None
    b_dest = Bombeiro.query.filter_by(mecanografico=mec_destino).first() if mec_destino else None
    data_orig = _parse_data(data_origem_str)
    data_dest = _parse_data(data_destino_str)
    data_pedido = _parse_datetime(data_pedido_str) or datetime.utcnow()
    if not b_orig or not b_dest or not data_orig or not data_dest:
        return "Dados de troca incompletos"
    t = TrocaServico(bombeiro_origem_id=b_orig.id, bombeiro_destino_id=b_dest.id,
                     data_origem=data_orig, data_destino=data_dest,
                     motivo=motivo, estado=estado, data_pedido=data_pedido)
    db.session.add(t)
    db.session.flush()
    return None


def _importar_linha_dispensas(row, row_num):
    mec = str(row[0]).strip() if row[0] else None
    inicio_str = str(row[1]).strip() if len(row) > 1 else None
    fim_str = str(row[2]).strip() if len(row) > 2 else None
    motivo = str(row[3]).strip() if len(row) > 3 else ''
    aprovada = str(row[4]).strip().lower() == 'sim' if len(row) > 4 else False
    inicio = _parse_data(inicio_str)
    fim = _parse_data(fim_str)
    if not mec or not inicio or not fim:
        return "Dados de dispensa incompletos"
    bombeiro = Bombeiro.query.filter_by(mecanografico=mec).first()
    if not bombeiro:
        return f"Bombeiro {mec} não encontrado"
    d = Dispensa(bombeiro_id=bombeiro.id, data_inicio=inicio, data_fim=fim,
                 motivo=motivo, aprovada=aprovada)
    db.session.add(d)
    db.session.flush()
    return None


def _importar_linha_creditos(row, row_num):
    mec = str(row[0]).strip() if row[0] else None
    data_str = str(row[1]).strip() if len(row) > 1 else None
    descricao = str(row[2]).strip() if len(row) > 2 else ''
    horas = int(row[3]) if len(row) > 3 and row[3] else 8
    estado = str(row[4]).strip() if len(row) > 4 else 'Não Gozado'
    data = _parse_data(data_str)
    if not mec or not data:
        return "Dados de crédito incompletos"
    bombeiro = Bombeiro.query.filter_by(mecanografico=mec).first()
    if not bombeiro:
        return f"Bombeiro {mec} não encontrado"
    c = CreditoDispensa(bombeiro_id=bombeiro.id, data=data,
                        descricao=descricao, horas=horas, observacao=estado)
    db.session.add(c)
    db.session.flush()
    return None


def _importar_linha_ecins(row, row_num):
    mec = str(row[0]).strip() if row[0] else None
    data_str = str(row[1]).strip() if len(row) > 1 else None
    turno = str(row[2]).strip() if len(row) > 2 else ''
    categoria = str(row[3]).strip() if len(row) > 3 else ''
    funcao = str(row[4]).strip() if len(row) > 4 else None
    estado = str(row[5]).strip() if len(row) > 5 else 'Pendente'
    data = _parse_data(data_str)
    if not mec or not data:
        return "Dados de ECIN incompletos"
    bombeiro = Bombeiro.query.filter_by(mecanografico=mec).first()
    if not bombeiro:
        return f"Bombeiro {mec} não encontrado"
    ec = Ecin(bombeiro_id=bombeiro.id, data=data, turno=turno,
              categoria=categoria, funcao=funcao, estado=estado)
    db.session.add(ec)
    db.session.flush()
    return None


def _importar_linha_fardamentos(row, row_num):
    mec = str(row[1]).strip() if len(row) > 1 and row[1] else None  # coluna 1 = mec
    tipo = str(row[2]).strip() if len(row) > 2 else ''
    nome = str(row[3]).strip() if len(row) > 3 else ''
    tamanho = str(row[4]).strip() if len(row) > 4 else ''
    motivo = str(row[5]).strip() if len(row) > 5 else ''
    estado = str(row[6]).strip() if len(row) > 6 else 'Pedido'
    data_reg_str = str(row[0]).strip() if len(row) > 0 and row[0] else None  # coluna 0 = data registo
    data_reg = _parse_datetime(data_reg_str) or datetime.utcnow()
    if not mec:
        return "Mecanográfico do bombeiro obrigatório"
    bombeiro = Bombeiro.query.filter_by(mecanografico=mec).first()
    if not bombeiro:
        return f"Bombeiro {mec} não encontrado"
    f = Fardamento(bombeiro_id=bombeiro.id, tipo=tipo, nome=nome,
                   tamanho=tamanho, motivo=motivo, estado=estado, data_registo=data_reg)
    db.session.add(f)
    db.session.flush()
    return None

def _importar_linha_fardamento_atribuido(row, row_num):
    try:
        id_original = int(row[0]) if row[0] else None
        bombeiro_id = int(row[1]) if len(row) > 1 and row[1] else None
        tipo = str(row[2]).strip() if len(row) > 2 else ''
        nome = str(row[3]).strip() if len(row) > 3 else ''
        tamanho = str(row[4]).strip() if len(row) > 4 else ''
        data_entrega_str = str(row[5]).strip() if len(row) > 5 else None
        data_entrega = _parse_data(data_entrega_str) if data_entrega_str else None
        estado = str(row[6]).strip() if len(row) > 6 else 'Entregue'
        idpedido = int(row[7]) if len(row) > 7 and row[7] else None

        if not bombeiro_id or not data_entrega:
            return None

        # Atualiza ou cria mantendo o ID
        existente = FardamentoAtribuido.query.get(id_original) if id_original else None
        if existente:
            existente.bombeiro_id = bombeiro_id
            existente.tipo = tipo
            existente.nome = nome
            existente.tamanho = tamanho
            existente.data_entrega = data_entrega
            existente.estado = estado
            existente.idpedido = idpedido
            return None  # já atualizado, não precisa adicionar novamente
        else:
            novo = FardamentoAtribuido(
                id=id_original,
                bombeiro_id=bombeiro_id,
                tipo=tipo,
                nome=nome,
                tamanho=tamanho,
                data_entrega=data_entrega,
                estado=estado,
                idpedido=idpedido
            )
            return novo
    except Exception:
        return None


def _importar_linha_oficina(row, row_num):
    codigo = str(row[0]).strip() if row[0] else ''
    nome_oficina = str(row[1]).strip() if len(row) > 1 else ''
    data_rec_str = str(row[2]).strip() if len(row) > 2 else None
    motivo = str(row[3]).strip() if len(row) > 3 else ''
    avaria_cod = str(row[4]).strip() if len(row) > 4 else None
    matricula = str(row[5]).strip() if len(row) > 5 else None
    kms = int(row[6]) if len(row) > 6 and row[6] else None
    estado = str(row[7]).strip() if len(row) > 7 else 'Oficina'

    if not nome_oficina:
        return "Nome da oficina obrigatório"
    viatura = Viatura.query.filter_by(matricula=matricula).first() if matricula else None
    avaria = Avaria.query.filter_by(codigo=avaria_cod).first() if avaria_cod else None
    data_recepcao = _parse_data(data_rec_str)
    o = Oficina(codigo=codigo, nome_oficina=nome_oficina,
                data_recepcao=data_recepcao,
                motivo=motivo, avaria_id=avaria.id if avaria else None,
                viatura_id=viatura.id if viatura else 1,
                kms=kms, estado=estado)
    db.session.add(o)
    db.session.flush()
    return None


def _importar_linha_gestao_frota(row, row_num):
    matricula = str(row[0]).strip() if row[0] else None
    if not matricula:
        return "Matrícula obrigatória"
    viatura = Viatura.query.filter_by(matricula=matricula).first()
    if not viatura:
        return f"Viatura {matricula} não encontrada"
    g = GestaoFrota(viatura_id=viatura.id)
    g.inspecao_periodica = _parse_data(str(row[1]).strip()) if len(row) > 1 and row[1] else None
    g.kms_ultima_revisao = int(row[2]) if len(row) > 2 and row[2] else None
    g.kms_proxima_revisao = int(row[3]) if len(row) > 3 and row[3] else None
    g.kms_pneus_dianteiros = int(row[4]) if len(row) > 4 and row[4] else None
    g.kms_pneus_trazeiros = int(row[5]) if len(row) > 5 and row[5] else None
    g.kms_correia = int(row[6]) if len(row) > 6 and row[6] else None
    g.outros_apontamentos = str(row[7]).strip() if len(row) > 7 and row[7] else ''
    db.session.add(g)
    db.session.flush()
    return None


def _importar_linha_stock_ambulancia(row, row_num):
    try:
        data_str = str(row[0]).strip() if row[0] else None
        matricula = str(row[1]).strip() if len(row) > 1 and row[1] else None
        nome_produto = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        quantidade = int(row[3]) if len(row) > 3 and row[3] else 0
        mec_solicitante = str(row[4]).strip() if len(row) > 4 and row[4] else None
        mec_responsavel = str(row[5]).strip() if len(row) > 5 and row[5] else None
        confirmado = str(row[6]).strip().lower() == 'sim' if len(row) > 6 else False

        if not matricula or not nome_produto:
            return None

        viatura = Viatura.query.filter_by(matricula=matricula).first()
        produto = StockFarmacia.query.filter_by(nome=nome_produto).first()
        if not viatura or not produto:
            return None   # saltar esta linha se a viatura ou o produto não existirem

        solicitante = Bombeiro.query.filter_by(mecanografico=mec_solicitante).first() if mec_solicitante else None
        responsavel = Bombeiro.query.filter_by(mecanografico=mec_responsavel).first() if mec_responsavel else None

        sa = StockAmbulancia(
            ambulancia_id=viatura.id,
            produto_id=produto.id,
            quantidade=quantidade,
            solicitante_id=solicitante.id if solicitante else 1,
            responsavel_id=responsavel.id if responsavel else None,
            checklist_id=None,
            confirmado=confirmado,
            data=_parse_datetime(data_str) or datetime.utcnow()
        )
        return sa
    except Exception:
        return None


def _importar_linha_notas_central(row, row_num):
    criador_mec = str(row[0]).strip() if row[0] else None
    data_criacao_str = str(row[1]).strip() if len(row) > 1 else None
    descricao = str(row[2]).strip() if len(row) > 2 else ''
    data_evento_str = str(row[3]).strip() if len(row) > 3 else None

    criador = Bombeiro.query.filter_by(mecanografico=criador_mec).first() if criador_mec else None
    data_criacao = _parse_datetime(data_criacao_str) or datetime.utcnow()
    data_evento = _parse_data(data_evento_str)

    n = Nota(criador_id=criador.id if criador else 1, data_criacao=data_criacao,
             descricao=descricao, data_evento=data_evento)
    db.session.add(n)
    db.session.flush()
    return None


def _importar_linha_mensagens_correio(row, row_num):
    remetente_mec = str(row[0]).strip() if row[0] else None
    destinatario_mec = str(row[1]).strip() if len(row) > 1 else None
    departamento = str(row[2]).strip() if len(row) > 2 else None
    assunto = str(row[3]).strip() if len(row) > 3 else ''
    corpo = str(row[4]).strip() if len(row) > 4 else ''
    data_envio_str = str(row[5]).strip() if len(row) > 5 else None
    lida = str(row[6]).strip().lower() == 'sim' if len(row) > 6 else False

    remetente = Bombeiro.query.filter_by(mecanografico=remetente_mec).first() if remetente_mec else None
    destinatario = Bombeiro.query.filter_by(mecanografico=destinatario_mec).first() if destinatario_mec else None
    data_envio = _parse_datetime(data_envio_str) or datetime.utcnow()

    m = MensagemCorreio(
        remetente_id=remetente.id if remetente else 1,
        destinatario_id=destinatario.id if destinatario else None,
        departamento=departamento if departamento and departamento != '' else None,
        assunto=assunto, corpo=corpo, data_envio=data_envio, lida=lida
    )
    db.session.add(m)
    db.session.flush()
    return None


@app.route('/backup/importar', methods=['POST'])
@login_required
def backup_importar():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    if 'ficheiro_backup' not in request.files:
        flash('Nenhum ficheiro enviado.', 'warning')
        return redirect(url_for('dashboard'))

    ficheiro = request.files['ficheiro_backup']
    if ficheiro.filename == '' or not ficheiro.filename.endswith(('.xlsx', '.xlsm')):
        flash('Formato inválido. Use .xlsx.', 'danger')
        return redirect(url_for('dashboard'))

    try:
        wb = openpyxl.load_workbook(ficheiro)
    except Exception as e:
        flash(f'Erro ao ler ficheiro: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))

    erros = []
    total_importado = 0

    # ---------- 1. APAGAR TODOS OS DADOS EXISTENTES ----------
    modelos_para_apagar = [
        NotaComando, Reuniao, TipoFardaMaterial,
        Nota, MensagemCorreio,
        StockAmbulancia,
        ChecklistAmbulanciaItem,
        ChecklistAmbulancia,
        StockFarmacia, CategoriaFarmacia,
        StockFardamento,
        Fardamento, FardamentoAtribuido, Ecin, GestaoFrota, Oficina,
        CreditoDispensa, Dispensa, TrocaServico, Escala,
        Avaria, Disponibilidade, Deslocacao, Ferias,  # ← apenas "Ferias"
        Viatura, Bombeiro
        ]

    for modelo in modelos_para_apagar:
        db.session.query(modelo).delete()
    db.session.flush()

    # ---------- 2. IMPORTAR BOMBEIROS ----------
    if 'Bombeiros' in wb.sheetnames:
        ws = wb['Bombeiros']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                b = _importar_linha_bombeiros(row, row_num)
                if b: db.session.add(b); total_importado += 1
            except Exception as e:
                erros.append(f"Bombeiros linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 3. IMPORTAR VIATURAS ----------
    if 'Viaturas' in wb.sheetnames:
        ws = wb['Viaturas']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                v = _importar_linha_viaturas(row, row_num)
                if v: db.session.add(v); total_importado += 1
            except Exception as e:
                erros.append(f"Viaturas linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 4. CATEGORIAS FARMÁCIA ----------
    if 'Categorias Farmacia' in wb.sheetnames:
        ws = wb['Categorias Farmacia']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                c = _importar_linha_categorias_farmacia(row, row_num)
                if c: db.session.add(c); total_importado += 1
            except Exception as e:
                erros.append(f"Categorias Farmácia linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 5. STOCK FARMÁCIA ----------
    if 'Stock Farmacia' in wb.sheetnames:
        ws = wb['Stock Farmacia']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                s = _importar_linha_stock_farmacia(row, row_num)
                if s: db.session.add(s); total_importado += 1
            except Exception as e:
                erros.append(f"Stock Farmácia linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 6. STOCK FARDAMENTO ----------

    # ---------- 6. STOCK FARDAMENTO ----------
    if 'Stock Fardamento' in wb.sheetnames:
        ws = wb['Stock Fardamento']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                s = _importar_linha_stock_fardamento(row, row_num)
                if s: db.session.add(s); total_importado += 1
            except Exception as e:
                erros.append(f"Stock Fardamento linha {row_num}: {str(e)[:100]}")
        db.session.flush()

        # Ajustar sequência no PostgreSQL (corrigido)
        if db.engine.dialect.name == 'postgresql':
            from sqlalchemy import text
            # Verifica se existe algum registo na tabela
            max_id = db.session.query(db.func.max(StockFardamento.id)).scalar()
            if max_id is not None:
                db.session.execute(text("SELECT setval('stock_fardamento_id_seq', :max_id)"), {'max_id': max_id})
            else:
                # Se a tabela está vazia, a próxima chave deve ser 1
                db.session.execute(text("SELECT setval('stock_fardamento_id_seq', 1, false)"))


    # ---------- 7. DISPONIBILIDADES ----------
    if 'Disponibilidades' in wb.sheetnames:
        ws = wb['Disponibilidades']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                d = _importar_linha_disponibilidades(row, row_num)
                if d: db.session.add(d); total_importado += 1
            except Exception as e:
                erros.append(f"Disponibilidades linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 8. ESCALAS ----------
    if 'Escalas' in wb.sheetnames:
        ws = wb['Escalas']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                e = _importar_linha_escalas(row, row_num)
                if e: db.session.add(e); total_importado += 1
            except Exception as e:
                erros.append(f"Escalas linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 9. AVARIAS ----------
    if 'Avarias' in wb.sheetnames:
        ws = wb['Avarias']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                a = _importar_linha_avarias(row, row_num)
                if a: db.session.add(a); total_importado += 1
            except Exception as e:
                erros.append(f"Avarias linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 10. TROCAS ----------
    if 'Trocas' in wb.sheetnames:
        ws = wb['Trocas']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                t = _importar_linha_trocas(row, row_num)
                if t: db.session.add(t); total_importado += 1
            except Exception as e:
                erros.append(f"Trocas linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 11. DISPENSAS ----------
    if 'Dispensas' in wb.sheetnames:
        ws = wb['Dispensas']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                d = _importar_linha_dispensas(row, row_num)
                if d: db.session.add(d); total_importado += 1
            except Exception as e:
                erros.append(f"Dispensas linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 12. CRÉDITOS ----------
    if 'Créditos' in wb.sheetnames:
        ws = wb['Créditos']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                c = _importar_linha_creditos(row, row_num)
                if c: db.session.add(c); total_importado += 1
            except Exception as e:
                erros.append(f"Créditos linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 13. ECINS ----------
    if 'ECINS' in wb.sheetnames:
        ws = wb['ECINS']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                ec = _importar_linha_ecins(row, row_num)
                if ec: db.session.add(ec); total_importado += 1
            except Exception as e:
                erros.append(f"ECINS linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 14. FARDAMENTOS ----------
    if 'Fardamentos' in wb.sheetnames:
        ws = wb['Fardamentos']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                f = _importar_linha_fardamentos(row, row_num)
                if f: db.session.add(f); total_importado += 1
            except Exception as e:
                erros.append(f"Fardamentos linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 15. OFICINA ----------
    if 'Oficina' in wb.sheetnames:
        ws = wb['Oficina']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                o = _importar_linha_oficina(row, row_num)
                if o: db.session.add(o); total_importado += 1
            except Exception as e:
                erros.append(f"Oficina linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 16. GESTÃO FROTA ----------
    if 'Gestao Frota' in wb.sheetnames:
        ws = wb['Gestao Frota']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                g = _importar_linha_gestao_frota(row, row_num)
                if g: db.session.add(g); total_importado += 1
            except Exception as e:
                erros.append(f"Gestão Frota linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 17. STOCK AMBULÂNCIA ----------
    if 'Stock Ambulância' in wb.sheetnames:
        ws = wb['Stock Ambulância']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                sa = _importar_linha_stock_ambulancia(row, row_num)
                if sa: db.session.add(sa); total_importado += 1
            except Exception as e:
                erros.append(f"Stock Ambulância linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 18. NOTAS ----------
    if 'Notas Central' in wb.sheetnames:
        ws = wb['Notas Central']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                n = _importar_linha_notas_central(row, row_num)
                if n: db.session.add(n); total_importado += 1
            except Exception as e:
                erros.append(f"Notas linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 19. MENSAGENS CORREIO ----------
    if 'Mensagens Correio' in wb.sheetnames:
        ws = wb['Mensagens Correio']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                m = _importar_linha_mensagens_correio(row, row_num)
                if m: db.session.add(m); total_importado += 1
            except Exception as e:
                erros.append(f"Mensagens linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 20. DESLOCAÇÕES ----------
    if 'Deslocacoes' in wb.sheetnames:
        ws = wb['Deslocacoes']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                d = _importar_linha_deslocacoes(row, row_num)
                if d: db.session.add(d); total_importado += 1
            except Exception as e:
                erros.append(f"Deslocações linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 21. FARDAMENTO ATRIBUÍDO ----------
    if 'Fardamento Atribuido' in wb.sheetnames:
        ws = wb['Fardamento Atribuido']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                fa = _importar_linha_fardamento_atribuido(row, row_num)
                if fa: db.session.add(fa); total_importado += 1
            except Exception as e:
                erros.append(f"Fardamento Atribuído linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 22. REUNIÕES ----------
    if 'Reunioes' in wb.sheetnames:
        ws = wb['Reunioes']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                r = _importar_linha_reunioes(row, row_num)
                if r: db.session.add(r); total_importado += 1
            except Exception as e:
                erros.append(f"Reuniões linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 23. NOTAS COMANDO ----------
    if 'Notas Comando' in wb.sheetnames:
        ws = wb['Notas Comando']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                nc = _importar_linha_notas_comando(row, row_num)
                if nc: db.session.add(nc); total_importado += 1
            except Exception as e:
                erros.append(f"Notas Comando linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 24. TIPOS FARDA/MATERIAL ----------
    if 'Tipos Farda Material' in wb.sheetnames:
        ws = wb['Tipos Farda Material']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                tf = _importar_linha_tipos_farda_material(row, row_num)
                if tf: db.session.add(tf); total_importado += 1
            except Exception as e:
                erros.append(f"Tipos Farda Material linha {row_num}: {str(e)[:100]}")
        db.session.flush()

    # ---------- 25. FÉRIAS ----------
    # ---------- 25. FÉRIAS ----------
    if 'Ferias' in wb.sheetnames:
        ws = wb['Ferias']
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(c is None for c in row):
                continue
            try:
                fer = _importar_linha_ferias(row, row_num)
                if fer:
                    db.session.add(fer)
                    total_importado += 1
            except Exception as e:
                erros.append(f"Férias linha {row_num}: {str(e)[:100]}")
        db.session.flush()

        # Ajustar sequência no PostgreSQL (se necessário)
        if db.engine.dialect.name == 'postgresql':
            from sqlalchemy import text
            max_id = db.session.query(db.func.max(Ferias.id)).scalar()
            if max_id is not None:
                db.session.execute(text("SELECT setval('ferias_id_seq', :max_id)"), {'max_id': max_id})
            else:
                db.session.execute(text("SELECT setval('ferias_id_seq', 1, false)"))

 # ---------- commit final e redireccionamento ----------
    db.session.commit()
    if erros:
        flash(f'{total_importado} registos importados. {len(erros)} erro(s): ' + '; '.join(erros[:5]), 'warning')
    else:
        flash(f'{total_importado} registos importados com sucesso!', 'success')
    return redirect(url_for('dashboard'))

#-----------------Backup-----------------------
@app.route('/backup/exportar')
@login_required
def backup_exportar():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)

    def escrever_cabecalho(ws, cabecalhos):
        ws.append(cabecalhos)
        for col in range(1, len(cabecalhos) + 1):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font

    # ---- 1. Bombeiros ----
    ws = wb.active
    ws.title = "Bombeiros"
    escrever_cabecalho(ws, ['Nº Interno', 'Mecanográfico', 'Nome', 'Nome Completo', 'Email', 'Telemóvel', 'Posto',
                            'Tipo Bombeiro', 'Resp. Departamento', 'Tipo Utilizador', 'Ativo', 'Password Hash'])
    for b in Bombeiro.query.order_by(Bombeiro.numero_interno).all():
        ws.append([b.numero_interno, b.mecanografico, b.nome, b.nomecompleto or '', b.email,
                   b.telemovel or '', b.posto, b.tipo_bombeiro, b.resp_departamento or '', b.tipo_user,
                   'Sim' if b.ativo else 'Não', b.password_hash])

    # ---- 2. Viaturas ----
    ws = wb.create_sheet("Viaturas")
    escrever_cabecalho(ws, ['Matrícula', 'Tipo', 'Nomenclatura', 'Marca', 'Modelo', 'Ano', 'Estado'])
    for v in Viatura.query.order_by(Viatura.matricula).all():
        ws.append([v.matricula, v.tipo, v.nomenclatura, v.marca, v.modelo, v.ano, v.estado])

    # ---- 3. Avarias ----
    ws = wb.create_sheet("Avarias")
    escrever_cabecalho(ws, ['Código', 'Viatura Matrícula', 'Descrição', 'Reportado por (mec.)', 'Kms',
                            'Resp. Oficina', 'Comando', 'Estado', 'Data Reporte'])
    for a in Avaria.query.order_by(Avaria.data_reporte.asc()).all():
        ws.append([a.codigo, a.viatura.matricula if a.viatura else '', a.descricao,
                   a.reportador.mecanografico if a.reportador else '', a.kms or '',
                   'Sim' if a.responsavel_oficina else 'Não', 'Sim' if a.comando_verificado else 'Não',
                   a.estado, a.data_reporte.strftime('%d/%m/%Y %H:%M') if a.data_reporte else ''])

    # ---- 4. Escalas ----
    ws = wb.create_sheet("Escalas")
    escrever_cabecalho(ws, ['Mecanográfico', 'Início', 'Fim', 'Turno', 'Categoria', 'Função'])
    for e in Escala.query.order_by(Escala.data_inicio.asc()).all():
        ws.append([e.bombeiro.mecanografico if e.bombeiro else '',
                   e.data_inicio.strftime('%d/%m/%Y %H:%M') if e.data_inicio else '',
                   e.data_fim.strftime('%d/%m/%Y %H:%M') if e.data_fim else '',
                   e.turno, e.categoria, e.funcao or ''])

    # ---- 5. Trocas ----
    ws = wb.create_sheet("Trocas")
    escrever_cabecalho(ws, ['Origem (mec.)', 'Destino (mec.)', 'Data Origem', 'Data Destino', 'Motivo', 'Estado', 'Data Pedido'])
    for t in TrocaServico.query.order_by(TrocaServico.data_pedido.asc()).all():
        ws.append([t.bombeiro_origem.mecanografico if t.bombeiro_origem else '',
                   t.bombeiro_destino.mecanografico if t.bombeiro_destino else '',
                   t.data_origem.strftime('%d/%m/%Y') if t.data_origem else '',
                   t.data_destino.strftime('%d/%m/%Y') if t.data_destino else '',
                   t.motivo or '', t.estado or '',
                   t.data_pedido.strftime('%d/%m/%Y %H:%M') if t.data_pedido else ''])

    # ---- 6. Dispensas ----
    ws = wb.create_sheet("Dispensas")
    escrever_cabecalho(ws, ['Bombeiro (mec.)', 'Início', 'Fim', 'Motivo', 'Aprovada'])
    for d in Dispensa.query.order_by(Dispensa.data_inicio.asc()).all():
        ws.append([d.bombeiro.mecanografico if d.bombeiro else '',
                   d.data_inicio.strftime('%d/%m/%Y') if d.data_inicio else '',
                   d.data_fim.strftime('%d/%m/%Y') if d.data_fim else '',
                   d.motivo or '', 'Sim' if d.aprovada else 'Não'])

    # ---- 7. Disponibilidades ----
    ws = wb.create_sheet("Disponibilidades")
    escrever_cabecalho(ws, ['Bombeiro (mec.)', 'Data', 'Turno Extra', 'Categoria', 'Confirmada'])
    for d in Disponibilidade.query.order_by(Disponibilidade.data.asc()).all():
        ws.append([d.bombeiro.mecanografico if d.bombeiro else '',
                   d.data.strftime('%d/%m/%Y') if d.data else '',
                   d.turno_extra or '', d.categoria or '',
                   'Sim' if d.confirmada else 'Não'])

    # ---- 8. Créditos ----
    ws = wb.create_sheet("Créditos")
    escrever_cabecalho(ws, ['Bombeiro (mec.)', 'Data', 'Descrição', 'Horas', 'Estado'])
    for c in CreditoDispensa.query.order_by(CreditoDispensa.data.asc()).all():
        ws.append([c.bombeiro.mecanografico if c.bombeiro else '',
                   c.data.strftime('%d/%m/%Y') if c.data else '',
                   c.descricao or '', c.horas, c.observacao or ''])


    # ---- 9. Stock Fardamento ----
    ws = wb.create_sheet("Stock Fardamento")
    cabecalhos = ['ID', 'Nome', 'Descrição', 'Tamanho', 'Tipo', 'Stock']
    ws.append(cabecalhos)
    for col in range(1, len(cabecalhos) + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
    for s in StockFardamento.query.order_by(StockFardamento.nome).all():
        ws.append([s.id, s.nome, s.descricao or '', s.tamanho or '', s.tipo, s.stock])



    # ---- 10. Stock Farmácia ----
    ws = wb.create_sheet("Stock Farmacia")
    cabecalhos = ['ID', 'Categoria', 'Nome', 'Tamanho', 'Stock', 'Stock Mínimo', 'Última Atualização']
    ws.append(cabecalhos)
    for col in range(1, len(cabecalhos) + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font
    for s in StockFarmacia.query.order_by(StockFarmacia.nome).all():
        ws.append([s.id, s.categoria, s.nome, s.tamanho or '', s.stock, s.infstock or 0,
                   s.data_atualizacao.strftime('%d/%m/%Y %H:%M') if s.data_atualizacao else ''])

    # ---- 11. Categorias Farmácia ----
    ws = wb.create_sheet("Categorias Farmacia")
    escrever_cabecalho(ws, ['Nome', 'Checklist'])
    for cat in CategoriaFarmacia.query.order_by(CategoriaFarmacia.nome).all():
        ws.append([cat.nome, 'Sim' if cat.checklist else 'Não'])

    # ---- 12. Stock Ambulância ----
    ws = wb.create_sheet("Stock Ambulância")
    escrever_cabecalho(ws, ['Data', 'Ambulância', 'Produto', 'Quantidade', 'Solicitante (mec.)', 'Responsável (mec.)', 'Confirmado'])
    for sa in StockAmbulancia.query.order_by(StockAmbulancia.data.asc()).all():
        ws.append([sa.data.strftime('%d/%m/%Y %H:%M') if sa.data else '',
                   sa.ambulancia.matricula if sa.ambulancia else '',
                   sa.produto_stock.nome if sa.produto_stock else '',
                   sa.quantidade,
                   sa.solicitante.mecanografico if sa.solicitante else '',
                   sa.responsavel.mecanografico if sa.responsavel else '',
                   'Sim' if sa.confirmado else 'Não'])

    # ---- 13. Checklist Ambulância ----
    ws = wb.create_sheet("Checklist Ambulância")
    escrever_cabecalho(ws, ['ID Checklist', 'Data/Hora', 'Viatura', 'Bombeiro (mec.)', 'Finalizado'])
    for ch in ChecklistAmbulancia.query.order_by(ChecklistAmbulancia.data_hora.asc()).all():
        ws.append([ch.id, ch.data_hora.strftime('%d/%m/%Y %H:%M') if ch.data_hora else '',
                   ch.viatura.matricula if ch.viatura else '',
                   ch.bombeiro.mecanografico if ch.bombeiro else '',
                   'Sim' if ch.finalizado else 'Não'])

    # ---- 14. Checklist Ambulância Itens ----
    ws = wb.create_sheet("Checklist Itens")
    escrever_cabecalho(ws, ['Checklist ID', 'Produto', 'Quantidade'])
    for item in ChecklistAmbulanciaItem.query.all():
        ws.append([item.checklist_id, item.produto.nome if item.produto else '', item.quantidade])

    # ---- 15. Mensagens Correio ----
    ws = wb.create_sheet("Mensagens Correio")
    escrever_cabecalho(ws, ['Remetente (mec.)', 'Destinatário (mec.)', 'Departamento', 'Assunto', 'Corpo', 'Data', 'Lida'])
    for m in MensagemCorreio.query.order_by(MensagemCorreio.data_envio.asc()).all():
        ws.append([m.remetente.mecanografico if m.remetente else '',
                   m.destinatario.mecanografico if m.destinatario else '',
                   m.departamento or '', m.assunto, m.corpo,
                   m.data_envio.strftime('%d/%m/%Y %H:%M') if m.data_envio else '',
                   'Sim' if m.lida else 'Não'])

    # ---- 16. Notas Central ----
    ws = wb.create_sheet("Notas Central")
    escrever_cabecalho(ws, ['Criador (mec.)', 'Data Criação', 'Descrição', 'Data Evento'])
    for n in Nota.query.order_by(Nota.data_criacao.asc()).all():
        ws.append([n.criador.mecanografico if n.criador else '',
                   n.data_criacao.strftime('%d/%m/%Y %H:%M') if n.data_criacao else '',
                   n.descricao, n.data_evento.strftime('%d/%m/%Y') if n.data_evento else ''])

    # ---- 17. ECINS ----
    ws = wb.create_sheet("ECINS")
    escrever_cabecalho(ws, ['Bombeiro (mec.)', 'Data', 'Turno', 'Categoria', 'Função', 'Estado', 'Valor'])
    for ec in Ecin.query.order_by(Ecin.data.asc()).all():
        ws.append([ec.bombeiro.mecanografico if ec.bombeiro else '',
                   ec.data.strftime('%d/%m/%Y') if ec.data else '',
                   ec.turno, ec.categoria or '', ec.funcao or '', ec.estado, ec.valor or 0.0])

    # ---- 18. Oficina ----
    ws = wb.create_sheet("Oficina")
    escrever_cabecalho(ws, ['Código', 'Nome Oficina', 'Data Recepção', 'Motivo', 'Nº Avaria', 'Viatura', 'Kms', 'Estado'])
    for o in Oficina.query.order_by(Oficina.data_registo.asc()).all():
        ws.append([o.codigo, o.nome_oficina,
                   o.data_recepcao.strftime('%d/%m/%Y') if o.data_recepcao else '',
                   o.motivo or '', o.avaria.codigo if o.avaria else '',
                   o.viatura.matricula if o.viatura else '', o.kms or '', o.estado])

    # ---- 19. Gestão Frota ----
    ws = wb.create_sheet("Gestao Frota")
    escrever_cabecalho(ws, ['Matrícula', 'Inspeção', 'Kms Últ. Revisão', 'Kms Próx. Revisão', 'Kms Pneus Diant.',
                            'Kms Pneus Tras.', 'Kms Correia', 'Apontamentos'])
    for g in GestaoFrota.query.all():
        v = g.viatura
        ws.append([v.matricula if v else '',
                   g.inspecao_periodica.strftime('%d/%m/%Y') if g.inspecao_periodica else '',
                   g.kms_ultima_revisao or '', g.kms_proxima_revisao or '',
                   g.kms_pneus_dianteiros or '', g.kms_pneus_trazeiros or '',
                   g.kms_correia or '', g.outros_apontamentos or ''])

    # ---- 20. Fardamentos (Pedidos) ----
    ws = wb.create_sheet("Fardamentos")
    escrever_cabecalho(ws, ['Data Registo', 'Bombeiro (mec.)', 'Tipo', 'Nome', 'Tamanho', 'Motivo', 'Estado'])
    for f in Fardamento.query.order_by(Fardamento.data_registo.asc()).all():
        ws.append([f.data_registo.strftime('%d/%m/%Y %H:%M') if f.data_registo else '',
                   f.bombeiro.mecanografico if f.bombeiro else '',
                   f.tipo, f.nome, f.tamanho, f.motivo, f.estado])

    # ---- 21. Fardamento Atribuído ----
    ws = wb.create_sheet("Fardamento Atribuido")
    escrever_cabecalho(ws, ['Bombeiro (mec.)', 'Tipo', 'Nome', 'Tamanho', 'Data Entrega', 'Estado', 'ID Pedido'])
    for fa in FardamentoAtribuido.query.order_by(FardamentoAtribuido.data_entrega.asc()).all():
        ws.append([fa.bombeiro.mecanografico if fa.bombeiro else '',
                   fa.tipo, fa.nome, fa.tamanho,
                   fa.data_entrega.strftime('%d/%m/%Y') if fa.data_entrega else '',
                   fa.estado, fa.idpedido or ''])

    # ---- 22. Deslocações ----
    ws = wb.create_sheet("Deslocacoes")
    escrever_cabecalho(ws, ['Data', 'Hora', 'Serviço', 'Origem', 'Destino', 'Valor', 'Viatura', 'Nº Serviço', 'Bombeiro (mec.)'])
    for d in Deslocacao.query.order_by(Deslocacao.data.asc()).all():
        ws.append([d.data.strftime('%d/%m/%Y') if d.data else '',
                   d.hora_inicio, d.servico, d.local_origem or '', d.local_destino or '',
                   d.valor or 0.0,
                   d.viatura.matricula if d.viatura else '',
                   d.n_servico or '',
                   d.bombeiro.mecanografico if d.bombeiro else ''])

    # ---- 23. Reuniões ----
    ws = wb.create_sheet("Reunioes")
    escrever_cabecalho(ws, ['Data', 'Hora', 'Assunto', 'Descrição', 'Criador (mec.)'])
    for r in Reuniao.query.order_by(Reuniao.data.asc()).all():
        ws.append([r.data.strftime('%d/%m/%Y') if r.data else '',
                   r.hora or '', r.assunto, r.descricao or '',
                   r.criador.mecanografico if r.criador else ''])

    # ---- 24. Notas Comando ----
    ws = wb.create_sheet("Notas Comando")
    escrever_cabecalho(ws, ['Criador (mec.)', 'Data Criação', 'Descrição', 'Data Evento'])
    for n in NotaComando.query.order_by(NotaComando.data_criacao.asc()).all():
        ws.append([n.criador.mecanografico if n.criador else '',
                   n.data_criacao.strftime('%d/%m/%Y %H:%M') if n.data_criacao else '',
                   n.descricao, n.data_evento.strftime('%d/%m/%Y') if n.data_evento else ''])

    # ---- 25. Tipos Farda/Material ----
    ws = wb.create_sheet("Tipos Farda Material")
    escrever_cabecalho(ws, ['Nome', 'Categoria'])
    for t in TipoFardaMaterial.query.order_by(TipoFardaMaterial.nome).all():
        ws.append([t.nome, t.categoria])

    # ---- 26. Férias ----
    ws = wb.create_sheet("Ferias")
    cabecalhos = ['Mecanográfico', 'Início', 'Fim', 'Estado', 'Aprovado por (nome)', 'Data Pedido']
    escrever_cabecalho(ws, cabecalhos)

    for f in Ferias.query.order_by(Ferias.data_inicio).all():
        mecanografico = f.bombeiro.mecanografico if f.bombeiro else ''
        nome_aprovador = ''
        if f.aprovado_por:
            aprovador = Bombeiro.query.get(f.aprovado_por)
            if aprovador:
                nome_aprovador = aprovador.nome

        ws.append([
            mecanografico,
            f.data_inicio.strftime('%d/%m/%Y') if f.data_inicio else '',
            f.data_fim.strftime('%d/%m/%Y') if f.data_fim else '',
            f.estado,
            nome_aprovador,
            f.data_pedido.strftime('%d/%m/%Y %H:%M') if f.data_pedido else ''
        ])


    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='backup_quartel.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


#----------------------Painel Comando__________________

@app.route('/painel-comando')
@login_required
def painel_comando():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito ao Comando.', 'danger')
        return redirect(url_for('dashboard'))

    hoje = date.today()
    daqui_7_dias = hoje + timedelta(days=7)

    # ----- Total escalados hoje -----
    escalas_hoje = Escala.query.filter(
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje
    ).order_by(Escala.categoria, Escala.turno, Escala.data_inicio).all()

    total_hoje = len(set(e.bombeiro_id for e in escalas_hoje))

    categorias_hoje = {}
    for e in escalas_hoje:
        cat = e.categoria or 'Outros'
        if cat not in categorias_hoje:
            categorias_hoje[cat] = []
        categorias_hoje[cat].append({
            'nome': e.bombeiro.nome,
            'mecanografico': e.bombeiro.mecanografico,
            'turno': e.turno,
            'posto': e.bombeiro.posto
        })

    # ----- Avarias ativas -----
    avarias_ativas = Avaria.query.filter(Avaria.estado.in_(['Pendente', 'Analisar'])).count()

    # ----- Trocas pendentes -----
    trocas_pendentes = TrocaServico.query.filter_by(estado='aceite_colega').count()

    # ----- Dispensas por aprovar -----
    dispensas_pendentes = Dispensa.query.filter_by(aprovada=False).count()

    # ----- Disponibilidades por confirmar (substitui Créditos) -----
    disponibilidades_pendentes = Disponibilidade.query.filter_by(confirmada=False).count()

    # ----- Viaturas Inoperacionais -----
    viaturas_inop = Viatura.query.filter(
        func.lower(Viatura.estado) == 'inoperacional'
    ).order_by(Viatura.matricula).all()

    # ----- Reuniões (próximos 7 dias) -----
    reunioes = Reuniao.query.filter(
        Reuniao.data >= hoje,
        Reuniao.data <= daqui_7_dias
    ).order_by(Reuniao.data.asc(), Reuniao.hora.asc()).all()

    # ----- Notas (data_evento nos próximos 7 dias) -----
    notas = NotaComando.query.filter(
        NotaComando.data_evento >= hoje,
        NotaComando.data_evento <= daqui_7_dias
    ).order_by(NotaComando.data_evento.asc()).all()

    return render_template('painel_comando.html',
                           hoje=hoje,
                           total_hoje=total_hoje,
                           categorias_hoje=categorias_hoje,
                           avarias_ativas=avarias_ativas,
                           trocas_pendentes=trocas_pendentes,
                           dispensas_pendentes=dispensas_pendentes,
                           disponibilidades_pendentes=disponibilidades_pendentes,
                           viaturas_inop=viaturas_inop,
                           reunioes=reunioes,
                           notas=notas)

def obter_turno_atual():
    """Retorna o nome do turno (ex: '1 - 00h/08h') correspondente à hora atual."""
    hora = datetime.now().hour
    turnos = [
        ('1 - 00h/08h', (0, 8)),
        ('2 - 08h/16h', (8, 16)),
        ('3 - 16h/24h', (16, 24)),
        ('4 - 11h/19h', (11, 19)),
        ('5 - 10h/18h', (10, 18)),
        ('6 - 07h/19h', (7, 19)),
        ('8 - 08h/20h', (8, 20)),
        ('9 - 20h/08h', (20, 24)),  # 20h-24h parte
        ('7 - 19h/07h', None),       # tratado separadamente
    ]
    # Procurar turnos com intervalo simples
    for nome, (ini, fim) in turnos:
        if nome == '7 - 19h/07h':
            continue
        if ini <= hora < fim:
            return nome
    # Turnos que cruzam a meia-noite
    if hora >= 19 or hora < 7:
        return '7 - 19h/07h'
    if hora >= 20 or hora < 8:
        return '9 - 20h/08h'
    return 'Indefinido'

@app.route('/painel/reuniao/adicionar', methods=['POST'])
@login_required
def adicionar_reuniao():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('painel_comando'))
    data = datetime.strptime(request.form['data'], '%Y-%m-%d').date()
    hora = request.form.get('hora', '')
    assunto = request.form['assunto']
    descricao = request.form.get('descricao', '')
    nova = Reuniao(data=data, hora=hora if hora else None, assunto=assunto, descricao=descricao, criador_id=current_user.id)
    db.session.add(nova)
    db.session.commit()
    flash('Reunião agendada.', 'success')
    return redirect(url_for('painel_comando'))

@app.route('/painel/reuniao/editar/<int:id>', methods=['POST'])
@login_required
def editar_reuniao(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('painel_comando'))
    r = Reuniao.query.get_or_404(id)
    r.data = datetime.strptime(request.form['data'], '%Y-%m-%d').date()
    r.hora = request.form.get('hora', '')
    r.assunto = request.form['assunto']
    r.descricao = request.form.get('descricao', '')
    db.session.commit()
    flash('Reunião atualizada.', 'success')
    return redirect(url_for('painel_comando'))

@app.route('/painel/reuniao/apagar/<int:id>')
@login_required
def apagar_reuniao(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('painel_comando'))
    r = Reuniao.query.get_or_404(id)
    db.session.delete(r)
    db.session.commit()
    flash('Reunião removida.', 'info')
    return redirect(url_for('painel_comando'))

@app.route('/painel/notas/adicionar', methods=['POST'])
@login_required
def adicionar_nota_comando():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('painel_comando'))
    descricao = request.form['descricao']
    data_evento_str = request.form.get('data_evento')
    data_evento = datetime.strptime(data_evento_str, '%Y-%m-%d').date() if data_evento_str else None
    nota = NotaComando(
        criador_id=current_user.id,
        descricao=descricao,
        data_evento=data_evento
    )
    db.session.add(nota)
    db.session.commit()
    flash('Nota do comando registada.', 'success')
    return redirect(url_for('painel_comando'))


@app.route('/painel/notas/editar/<int:id>', methods=['POST'])
@login_required
def editar_nota_comando(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('painel_comando'))
    nota = NotaComando.query.get_or_404(id)
    nota.descricao = request.form['descricao']
    data_evento_str = request.form.get('data_evento')
    nota.data_evento = datetime.strptime(data_evento_str, '%Y-%m-%d').date() if data_evento_str else None
    db.session.commit()
    flash('Nota do comando atualizada.', 'success')
    return redirect(url_for('painel_comando'))


@app.route('/painel/notas/apagar/<int:id>')
@login_required
def apagar_nota_comando(id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('painel_comando'))
    nota = NotaComando.query.get_or_404(id)
    db.session.delete(nota)
    db.session.commit()
    flash('Nota do comando removida.', 'info')
    return redirect(url_for('painel_comando'))


@app.route('/admin/apagar-tudo')
@login_required
def apagar_tudo():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    # Ordem inversa para respeitar as chaves estrangeiras
    modelos_para_apagar = [
        Nota, MensagemCorreio, ChecklistAmbulanciaItem, ChecklistAmbulancia,
        StockAmbulancia, StockFarmacia, CategoriaFarmacia,
        Fardamento, Ecin, GestaoFrota, Oficina,
        CreditoDispensa, Dispensa, TrocaServico, Escala,
        Avaria, Disponibilidade, Viatura, Bombeiro
    ]
    for modelo in modelos_para_apagar:
        db.session.query(modelo).delete()
    db.session.commit()
    flash('Todos os dados foram eliminados com sucesso!', 'success')
    return redirect(url_for('dashboard'))

from datetime import date, timedelta


@app.context_processor
def inject_pendencias():
    pendencias = {}
    if current_user.is_authenticated:
        user = current_user
        total = 0

        # ========== ADMIN / COMANDO ==========
        if user.tipo_user == 'Admin' or user.resp_departamento == 'Comando':
            # Avarias pendentes
            pendencias['avarias'] = Avaria.query.filter(Avaria.estado.in_(['Pendente', 'Analisar'])).count()
            total += pendencias['avarias']

            # Trocas pendentes de aprovação
            pendencias['trocas'] = TrocaServico.query.filter_by(estado='aceite_colega').count()
            total += pendencias['trocas']

            # Dispensas pendentes
            pendencias['dispensas'] = Dispensa.query.filter_by(aprovada=False).count()
            total += pendencias['dispensas']

            # Créditos em análise
            pendencias['creditos'] = CreditoDispensa.query.filter_by(observacao='Em Análise').count()
            total += pendencias['creditos']

            # Fardamento pendente
            pendencias['fardamento'] = Fardamento.query.filter_by(estado='Pedido').count()
            total += pendencias['fardamento']

            # ECIN pendentes
            pendencias['ecins'] = Ecin.query.filter_by(estado='Pendente').count()
            total += pendencias['ecins']

            # Stock Farmácia abaixo do mínimo
            pendencias['stock_farmacia_minimo'] = StockFarmacia.query.filter(
                StockFarmacia.infstock > 0,
                StockFarmacia.stock <= StockFarmacia.infstock
            ).count()
            total += pendencias['stock_farmacia_minimo']

            # Stock Farmácia Central abaixo do mínimo
            pendencias['central_stock_minimo'] = FarmaciaCentral.query.filter(
                FarmaciaCentral.stock_minimo > 0,
                FarmaciaCentral.stock <= FarmaciaCentral.stock_minimo
            ).count()
            total += pendencias['central_stock_minimo']

            # Reposições de ambulância pendentes
            pendencias['stock_ambulancia'] = StockAmbulancia.query.filter_by(confirmado=False).count()
            total += pendencias['stock_ambulancia']

            # Inspeções periódicas próximas
            hoje = date.today()
            limite = hoje + timedelta(days=30)
            pendencias['inspecoes_proximas'] = GestaoFrota.query.join(Viatura).filter(
                GestaoFrota.inspecao_periodica != None,
                GestaoFrota.inspecao_periodica >= hoje,
                GestaoFrota.inspecao_periodica <= limite
            ).count()
            total += pendencias['inspecoes_proximas']

        # ========== CENTRAL ==========
        elif user.resp_departamento == 'Central' and user.tipo_user != 'Admin':
            # Trocas pendentes de aprovação (para avisar o Comando)
            pendencias['trocas_pendentes'] = TrocaServico.query.filter_by(estado='aceite_colega').count()
            total += pendencias['trocas_pendentes']

            # Dispensas pendentes de aprovação
            pendencias['dispensas_pendentes'] = Dispensa.query.filter_by(aprovada=False).count()
            total += pendencias['dispensas_pendentes']

            # Contagem total no badge do menu
            pendencias['total_pendentes'] = pendencias['trocas_pendentes'] + pendencias['dispensas_pendentes']

        # ========== SECRETARIA ==========
        elif user.resp_departamento == 'Secretaria' and user.tipo_user != 'Admin':
            # ECIN pendentes
            pendencias['ecins'] = Ecin.query.filter_by(estado='Pendente').count()
            total += pendencias['ecins']

            # Créditos em análise
            pendencias['creditos'] = CreditoDispensa.query.filter_by(observacao='Em Análise').count()
            total += pendencias['creditos']

        # ========== OFICINA ==========
        elif user.resp_departamento == 'Oficina' and user.tipo_user != 'Admin':
            # Avarias pendentes
            pendencias['avarias'] = Avaria.query.filter(Avaria.estado.in_(['Pendente', 'Analisar'])).count()
            total += pendencias['avarias']

        # ========== FARMÁCIA ==========
        elif user.resp_departamento == 'Farmacia' and user.tipo_user != 'Admin':
            # Stock abaixo do mínimo
            pendencias['stock_farmacia_minimo'] = StockFarmacia.query.filter(
                StockFarmacia.infstock > 0,
                StockFarmacia.stock <= StockFarmacia.infstock
            ).count()
            total += pendencias['stock_farmacia_minimo']

            pendencias['central_stock_minimo'] = FarmaciaCentral.query.filter(
                FarmaciaCentral.stock_minimo > 0,
                FarmaciaCentral.stock <= FarmaciaCentral.stock_minimo
            ).count()
            total += pendencias['central_stock_minimo']

            # Reposições de ambulância pendentes
            pendencias['stock_ambulancia'] = StockAmbulancia.query.filter_by(confirmado=False).count()
            total += pendencias['stock_ambulancia']

        # ========== FARDAMENTO ==========
        elif user.resp_departamento == 'Fardamento' and user.tipo_user != 'Admin':
            # Pedidos pendentes
            pendencias['fardamento'] = Fardamento.query.filter_by(estado='Pedido').count()
            total += pendencias['fardamento']

        pendencias['total'] = total
    return dict(pendencias=pendencias)


#---------------------------Deslocações----------------
from datetime import date

from datetime import date, datetime

@app.route('/deslocacoes', methods=['GET', 'POST'])
@login_required
def deslocacoes():
    if request.method == 'POST':
        # Dados principais
        data_str = request.form['data']
        hora_inicio = request.form['hora_inicio']
        servico = request.form['servico']
        local_origem = request.form.get('local_origem', '')
        local_destino = request.form.get('local_destino', '')
        valor_str = request.form.get('valor', '')
        viatura_id = request.form.get('viatura_id', type=int)
        n_servico = request.form.get('n_servico', '')

        # Novos campos
        data_fim_str = request.form.get('data_fim', '')
        hora_fim = request.form.get('hora_fim', '')

        # Validar data início
        try:
            data = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            flash('Data inválida.', 'danger')
            return redirect(url_for('deslocacoes'))

        # Validar hora início
        if len(hora_inicio) > 10:
            try:
                dt = datetime.strptime(hora_inicio, '%d/%m/%Y %H:%M')
                hora_inicio = dt.strftime('%H:%M')
            except ValueError:
                try:
                    dt = datetime.strptime(hora_inicio, '%Y-%m-%d %H:%M')
                    hora_inicio = dt.strftime('%H:%M')
                except ValueError:
                    flash('Formato de hora início inválido. Use HH:MM.', 'danger')
                    return redirect(url_for('deslocacoes'))
        elif len(hora_inicio) != 5 or hora_inicio[2] != ':':
            flash('Hora início deve estar no formato HH:MM.', 'danger')
            return redirect(url_for('deslocacoes'))

        # Validar data fim (opcional)
        data_fim = None
        if data_fim_str:
            try:
                data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Data fim inválida.', 'danger')
                return redirect(url_for('deslocacoes'))

        # Validar hora fim (opcional)
        if hora_fim and (len(hora_fim) != 5 or hora_fim[2] != ':'):
            flash('Hora fim deve estar no formato HH:MM.', 'danger')
            return redirect(url_for('deslocacoes'))

        # Valor (apenas para perfis autorizados)
        valor = None
        if current_user.tipo_user == 'Admin' or current_user.resp_departamento in ['Comando', 'Secretaria']:
            if valor_str:
                try:
                    valor = float(valor_str)
                except ValueError:
                    pass

        nova = Deslocacao(
            bombeiro_id=current_user.id,
            data=data,
            hora_inicio=hora_inicio,
            data_fim=data_fim,
            hora_fim=hora_fim if hora_fim else None,
            servico=servico,
            local_origem=local_origem,
            local_destino=local_destino,
            valor=valor,
            viatura_id=viatura_id if viatura_id else None,
            n_servico=n_servico if n_servico else None
        )
        db.session.add(nova)
        db.session.commit()
        flash('Deslocação registada.', 'success')
        return redirect(url_for('deslocacoes'))

    # GET – listagem
    query = Deslocacao.query.order_by(Deslocacao.data.desc())
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria']:
        query = query.filter_by(bombeiro_id=current_user.id)

    deslocacoes_lista = query.all()
    viaturas = Viatura.query.order_by(Viatura.matricula).all()

    return render_template('deslocacoes.html',
                           deslocacoes=deslocacoes_lista,
                           viaturas=viaturas,
                           now=date.today())

@app.route('/administrativo')
@login_required
def administrativo():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Secretaria']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    # ---------- Filtros Deslocações ----------
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    bombeiro_id_desl = request.args.get('bombeiro_id', type=int)

    query_desl = Deslocacao.query
    if data_inicio:
        try:
            d = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query_desl = query_desl.filter(Deslocacao.data >= d)
        except ValueError:
            pass
    if data_fim:
        try:
            d = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query_desl = query_desl.filter(Deslocacao.data <= d)
        except ValueError:
            pass
    if bombeiro_id_desl:
        query_desl = query_desl.filter_by(bombeiro_id=bombeiro_id_desl)

    deslocacoes = query_desl.order_by(Deslocacao.data.desc()).all()
    total_valor_desl = sum(d.valor for d in deslocacoes if d.valor)
    bombeiros_ativos = Bombeiro.query.filter_by(ativo=True).order_by(Bombeiro.nome).all()

    # ---------- Filtros ECIN/ELAC (apenas registos com estado diferente de Pendente/Não Escalado) ----------
    mes_ecin = request.args.get('mes_ecin', type=int, default=date.today().month)
    ano_ecin = request.args.get('ano_ecin', type=int, default=date.today().year)
    cat_ecin = request.args.get('cat_ecin', 'todas')
    bombeiro_id_ecin = request.args.get('bombeiro_id_ecin', type=int)

    query_ecin = Ecin.query.filter(
        db.extract('month', Ecin.data) == mes_ecin,
        db.extract('year', Ecin.data) == ano_ecin
    )
    # --- EXCLUIR estados Pendente e Não Escalado ---
    query_ecin = query_ecin.filter(~Ecin.estado.in_(['Pendente', 'Não Escalado']))

    if cat_ecin == 'ECIN':
        query_ecin = query_ecin.filter(Ecin.categoria == 'ECIN')
    elif cat_ecin == 'ELAC':
        query_ecin = query_ecin.filter(Ecin.categoria == 'ELAC')
    if bombeiro_id_ecin:
        query_ecin = query_ecin.filter_by(bombeiro_id=bombeiro_id_ecin)

    ecins = query_ecin.order_by(Ecin.data.desc()).all()
    total_valor_ecin = sum(ec.valor for ec in ecins if ec.valor)
    turnos_ecin = query_ecin.filter(Ecin.categoria == 'ECIN').count()
    turnos_elac = query_ecin.filter(Ecin.categoria == 'ELAC').count()

    # Lista de bombeiros que aparecem nos ECINs filtrados (para o dropdown)
    bombeiros_ativos_ecin = Bombeiro.query.join(Ecin).filter(
        Ecin.bombeiro_id == Bombeiro.id,
        db.extract('month', Ecin.data) == mes_ecin,
        db.extract('year', Ecin.data) == ano_ecin,
        ~Ecin.estado.in_(['Pendente', 'Não Escalado'])
    ).distinct().order_by(Bombeiro.nome).all()

    # Atribuir 42.00 € automaticamente a todos os ECINs do filtro que ainda não têm valor
    ecins_sem_valor = [ec for ec in ecins if ec.valor is None]
    if ecins_sem_valor:
        for ec in ecins_sem_valor:
            ec.valor = 42.0
        db.session.commit()
        # Recarregar para refletir os novos valores
        ecins = query_ecin.order_by(Ecin.data.desc()).all()
        total_valor_ecin = sum(ec.valor for ec in ecins if ec.valor)

    return render_template('administrativo.html',
                           deslocacoes=deslocacoes,
                           total_valor_desl=total_valor_desl,
                           ecins=ecins,
                           total_valor_ecin=total_valor_ecin,
                           turnos_ecin=turnos_ecin,
                           turnos_elac=turnos_elac,
                           bombeiros_ativos=bombeiros_ativos,
                           bombeiros_ativos_ecin=bombeiros_ativos_ecin,
                           data_inicio=data_inicio,
                           data_fim=data_fim,
                           bombeiro_id_desl=bombeiro_id_desl,
                           mes_ecin=mes_ecin,
                           ano_ecin=ano_ecin,
                           cat_ecin=cat_ecin,
                           bombeiro_id_ecin=bombeiro_id_ecin,
                           now=date.today())



@app.route('/administrativo/atualizar-valor-ecin', methods=['POST'])
@login_required
def atualizar_valor_ecin():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('administrativo'))

    ec_id = request.form.get('ecin_id', type=int)
    novo_valor = request.form.get('valor', type=float)
    ec = Ecin.query.get_or_404(ec_id)
    ec.valor = novo_valor if novo_valor else None
    db.session.commit()

    # Preservar os filtros que o utilizador tinha
    mes = request.form.get('mes_atual', type=int)
    ano = request.form.get('ano_atual', type=int)
    cat = request.form.get('cat_atual', 'todas')
    bombeiro = request.form.get('bombeiro_atual', type=int)
    flash('Valor atualizado.', 'success')
    return redirect(url_for('administrativo', tab='ecins',
                            mes_ecin=mes, ano_ecin=ano, cat_ecin=cat,
                            bombeiro_id_ecin=bombeiro))



@app.route('/administrativo/mobilidade/criar', methods=['POST'])
@login_required
def criar_mobilidade():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        return jsonify({'erro': 'Acesso restrito'}), 403

    ecin_id = request.form.get('ecin_id', type=int)
    bombeiro_substituto_id = request.form.get('bombeiro_id', type=int)
    horas = request.form.get('horas', type=float)

    if not ecin_id or not bombeiro_substituto_id or not horas:
        return jsonify({'erro': 'Parâmetros inválidos'}), 400

    original = Ecin.query.get_or_404(ecin_id)
    if original.categoria not in ['ECIN', 'ELAC']:
        return jsonify({'erro': 'Registo não é ECIN/ELAC'}), 400

    if Mobilidade.query.filter_by(ecin_original_id=original.id).first():
        return jsonify({'erro': 'Este registo já possui uma mobilidade.'}), 400

    valor_substituto = 3.50 * horas
    valor_base = original.valor if original.valor is not None else 42.0
    novo_valor_original = max(0, valor_base - valor_substituto)

    novo_ecin = Ecin(
        bombeiro_id=bombeiro_substituto_id,
        data=original.data,
        turno=original.turno,
        categoria=original.categoria,
        funcao=original.funcao,
        estado='Mobilizado',
        valor=valor_substituto
    )
    db.session.add(novo_ecin)
    db.session.flush()

    mobilidade = Mobilidade(
        ecin_original_id=original.id,
        bombeiro_substituto_id=bombeiro_substituto_id,
        horas=horas,
        valor_pago=valor_substituto
    )
    db.session.add(mobilidade)
    original.valor = novo_valor_original
    db.session.commit()

    substituto = Bombeiro.query.get(bombeiro_substituto_id)
    return jsonify({
        'sucesso': True,
        'substituto_nome': substituto.nome,
        'original_nome': original.bombeiro.nome,
        'novo_valor_original': novo_valor_original,
        'novo_registo': {
            'id': novo_ecin.id,
            'data': novo_ecin.data.strftime('%d/%m/%Y'),
            'turno': novo_ecin.turno,
            'categoria': novo_ecin.categoria,
            'bombeiro': substituto.nome,
            'funcao': novo_ecin.funcao or '-',
            'estado': novo_ecin.estado,
            'valor': novo_ecin.valor
        }
    })


@app.route('/administrativo/mobilidade/apagar/<int:ecin_id>', methods=['POST'])
@login_required
def apagar_mobilidade(ecin_id):
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        return jsonify({'erro': 'Acesso restrito'}), 403

    original = Ecin.query.get_or_404(ecin_id)
    mobilidade = Mobilidade.query.filter_by(ecin_original_id=original.id).first()
    if not mobilidade:
        return jsonify({'erro': 'Mobilidade não encontrada'}), 404

    substituto_ecin = Ecin.query.filter_by(
        bombeiro_id=mobilidade.bombeiro_substituto_id,
        data=original.data,
        turno=original.turno,
        categoria=original.categoria
    ).first()
    if substituto_ecin:
        db.session.delete(substituto_ecin)

    original.valor = (original.valor or 0) + float(mobilidade.valor_pago)
    db.session.delete(mobilidade)
    db.session.commit()

    return jsonify({'sucesso': True, 'novo_valor_original': float(original.valor)})


@app.route('/ecins/imprimir-turno-diario')
@login_required
def imprimir_turno_diario_ecin():
    # Apenas perfis autorizados
    if current_user.tipo_user not in ['Admin'] and current_user.resp_departamento not in ['Comando', 'ECIN', 'Central']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('dashboard'))

    # Data actual (ou pode receber por parâmetro)
    data = request.args.get('data', type=str)
    if data:
        try:
            data = datetime.strptime(data, '%Y-%m-%d').date()
        except ValueError:
            data = date.today()
    else:
        data = date.today()

    # Buscar ECINs escalados para esta data (categoria ECIN e estado do tipo Motorista/Chefe/Guarnição)
    ecins = Ecin.query.filter(
        Ecin.data == data,
        Ecin.categoria == 'ECIN',
        Ecin.estado.in_(['Motorista ECIN', 'Chefe ECIN', 'Guarnição ECIN'])
    ).order_by(Ecin.turno, Ecin.funcao).all()

    # Separar por turno e ordem: Chefe, Motorista, depois Guarnição (até 3)
    turno_manha = []
    turno_noite = []
    for ec in ecins:
        bombeiro = ec.bombeiro
        info = {
            'mecanografico': bombeiro.mecanografico,
            'nome': bombeiro.nome,
            'posto': bombeiro.posto,
            'funcao': ec.funcao,   # 'Motorista', 'Chefe', 'Guarnição'
            'turno': ec.turno
        }
        if ec.turno == '07h/19h':
            turno_manha.append(info)
        else:
            turno_noite.append(info)

    # Ordenar cada turno: Chefe, Motorista, Guarnição (por nome)
    def ordenar_turno(lista):
        ordem = {'Chefe': 0, 'Motorista': 1, 'Guarnição': 2}
        return sorted(lista, key=lambda x: (ordem.get(x['funcao'], 3), x['nome']))

    turno_manha = ordenar_turno(turno_manha)
    turno_noite = ordenar_turno(turno_noite)

    # Garantir 5 linhas por turno: 1 Chefe, 1 Motorista, 3 Guarnição
    def preparar_linhas(turno):
        linhas = []
        # Chefe
        chefe = next((b for b in turno if b['funcao'] == 'Chefe'), None)
        linhas.append(chefe if chefe else {'mecanografico': '', 'nome': '', 'posto': '', 'funcao': 'CHEFE -'})
        # Motorista
        motorista = next((b for b in turno if b['funcao'] == 'Motorista'), None)
        linhas.append(motorista if motorista else {'mecanografico': '', 'nome': '', 'posto': '', 'funcao': 'MOT -'})
        # Guarnições (até 3)
        guarnicoes = [b for b in turno if b['funcao'] == 'Guarnição']
        for i in range(3):
            if i < len(guarnicoes):
                linhas.append(guarnicoes[i])
            else:
                linhas.append({'mecanografico': '', 'nome': '', 'posto': '', 'funcao': ''})
        return linhas

    linhas_manha = preparar_linhas(turno_manha)
    linhas_noite = preparar_linhas(turno_noite)

    meses = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    return render_template('imprimir_turno_ecin_diario.html',
                           data=data,
                           linhas_manha=linhas_manha,
                           linhas_noite=linhas_noite,
                           meses=meses)


@app.route('/administrativo/exportar-deslocacoes')
@login_required
def exportar_deslocacoes_administrativo():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('administrativo'))

    # Aplicar os mesmos filtros que a página usa
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    bombeiro_id = request.args.get('bombeiro_id', type=int)

    query = Deslocacao.query
    if data_inicio:
        try:
            d = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query = query.filter(Deslocacao.data >= d)
        except ValueError:
            pass
    if data_fim:
        try:
            d = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query = query.filter(Deslocacao.data <= d)
        except ValueError:
            pass
    if bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)

    desl = query.order_by(Deslocacao.data.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deslocacoes"

    cabecalhos = ['Data', 'Hora', 'Serviço', 'Origem', 'Destino', 'Valor', 'Viatura', 'Nº Serviço', 'Bombeiro']
    ws.append(cabecalhos)

    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True)
    for col in range(1, len(cabecalhos)+1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for d in desl:
        ws.append([
            d.data.strftime('%d/%m/%Y') if d.data else '',
            d.hora_inicio,
            d.servico,
            d.local_origem or '',
            d.local_destino or '',
            d.valor or 0.0,
            d.viatura.matricula if d.viatura else '',
            d.n_servico or '',
            d.bombeiro.nome if d.bombeiro else ''
        ])

    # Ajustar largura das colunas
    col_widths = [12, 8, 18, 20, 20, 10, 15, 12, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='deslocacoes.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/administrativo/exportar-resumo-deslocacoes')
@login_required
def exportar_resumo_deslocacoes():
    # totais por serviço e por viatura
    desl = Deslocacao.query.all()
    from collections import defaultdict
    por_servico = defaultdict(float)
    por_viatura = defaultdict(float)
    for d in desl:
        if d.valor:
            por_servico[d.servico] += d.valor
            mat = d.viatura.matricula if d.viatura else 'N/A'
            por_viatura[mat] += d.valor

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Contabilistico"
    ws.append(['Serviço', 'Total (€)'])
    for serv, val in por_servico.items():
        ws.append([serv, val])
    ws.append([])
    ws.append(['Viatura', 'Total (€)'])
    for mat, val in por_viatura.items():
        ws.append([mat, val])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='resumo_contabilistico.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/administrativo/imprimir-deslocacoes')
@login_required
def imprimir_deslocacoes_administrativo():
    # aplicar filtros se passados
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    bombeiro_id = request.args.get('bombeiro_id', type=int)
    query = Deslocacao.query
    if data_inicio:
        try:
            d = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query = query.filter(Deslocacao.data >= d)
        except ValueError: pass
    if data_fim:
        try:
            d = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query = query.filter(Deslocacao.data <= d)
        except ValueError: pass
    if bombeiro_id:
        query = query.filter_by(bombeiro_id=bombeiro_id)
    desl = query.order_by(Deslocacao.data.desc()).all()
    return render_template('imprimir_deslocacoes.html', deslocacoes=desl)


#-----------------exportar ecins administrativo----------------
@app.route('/administrativo/exportar-ecins')
@login_required
def exportar_ecins_administrativo():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('administrativo'))

    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)
    cat = request.args.get('cat_ecin', 'todas')

    query = Ecin.query.filter(
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    )
    if cat == 'ECIN':
        query = query.filter(Ecin.categoria == 'ECIN')
    elif cat == 'ELAC':
        query = query.filter(Ecin.categoria == 'ELAC')

    ecins = query.order_by(Ecin.data).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ECINs"
    cabecalhos = ['Data', 'Turno', 'Categoria', 'Bombeiro', 'Função', 'Estado', 'Valor']
    ws.append(cabecalhos)
    for ec in ecins:
        ws.append([ec.data.strftime('%d/%m/%Y') if ec.data else '',
                   ec.turno, ec.categoria or '',
                   ec.bombeiro.nome if ec.bombeiro else '',
                   ec.funcao or '', ec.estado,
                   ec.valor or 0.0])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='ecins.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/administrativo/enviar-contagem', methods=['POST'])
@login_required
def enviar_contagem_ecins():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('administrativo', tab='ecins'))

    mes = request.form.get('mes_ecin', type=int, default=date.today().month)
    ano = request.form.get('ano_ecin', type=int, default=date.today().year)
    cat = request.form.get('cat_ecin', 'todas')

    query = Ecin.query.filter(
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    )
    if cat == 'ECIN':
        query = query.filter(Ecin.categoria == 'ECIN')
    elif cat == 'ELAC':
        query = query.filter(Ecin.categoria == 'ELAC')

    ecins = query.order_by(Ecin.bombeiro_id).all()

    from collections import defaultdict
    contagem = defaultdict(lambda: {'turnos': 0, 'valor': 0.0})
    for ec in ecins:
        contagem[ec.bombeiro_id]['turnos'] += 1
        contagem[ec.bombeiro_id]['valor'] += ec.valor or 0.0

    enviadas = 0
    # Guardar dados para o resumo do remetente
    resumo_remetente = "Contagem de turnos enviada:\n\n"

    for bombeiro_id, dados in contagem.items():
        bombeiro = Bombeiro.query.get(bombeiro_id)
        if not bombeiro:
            continue

        categoria_nome = 'ECIN/ELAC'
        if cat == 'ECIN':
            categoria_nome = 'ECIN'
        elif cat == 'ELAC':
            categoria_nome = 'ELAC'

        total_turnos = dados['turnos']
        total_valor = dados['valor']
        corpo = (
            f"O(A) bombeiro(a) {bombeiro.nome} teve {total_turnos} turno(s) de {categoria_nome} "
            f"no mês de {mes}/{ano}, no valor de {total_valor:.2f} €."
        )

        msg = MensagemCorreio(
            remetente_id=current_user.id,
            destinatario_id=bombeiro_id,
            departamento=None,
            assunto='Envio de Contagem de Turno',
            corpo=corpo,
            data_envio=datetime.utcnow(),
            lida=False,
            apagada_remetente=False,
            apagada_destinatario=False
        )
        db.session.add(msg)
        enviadas += 1

        # Adicionar ao resumo do remetente
        resumo_remetente += f"{bombeiro.nome}: {total_turnos} turnos, {total_valor:.2f} €\n"

    # Enviar cópia da contagem ao remetente
    msg_remetente = MensagemCorreio(
        remetente_id=current_user.id,
        destinatario_id=current_user.id,
        departamento=None,
        assunto='Cópia da Contagem de Turno Enviada',
        corpo=resumo_remetente,
        data_envio=datetime.utcnow(),
        lida=False,
        apagada_remetente=False,
        apagada_destinatario=False
    )
    db.session.add(msg_remetente)

    db.session.commit()
    flash(f'Contagens enviadas para {enviadas} bombeiros. Uma cópia foi guardada na sua caixa de entrada.', 'success')
    return redirect(url_for('administrativo', tab='ecins',
                            mes_ecin=mes, ano_ecin=ano, cat_ecin=cat))


@app.route('/administrativo/imprimir-contabilidade-ecin')
@login_required
def imprimir_contabilidade_ecin():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('administrativo', tab='ecins'))

    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)

    registos = Ecin.query.filter(
        Ecin.categoria == 'ECIN',
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    ).order_by(Ecin.bombeiro_id, Ecin.data).all()

    from collections import defaultdict
    dados = defaultdict(lambda: {'valor': 0.0})

    for ec in registos:
        if ec.bombeiro and ec.valor:
            dados[ec.bombeiro]['valor'] += ec.valor

    # Converter valor em horas (3.50 €/h) e depois em turnos completos (12h) + horas restantes
    for bombeiro, info in dados.items():
        horas_totais = info['valor'] / 3.5
        turnos_int = int(horas_totais // 12)
        horas_rest = round(horas_totais % 12, 1)
        if horas_rest > 11.9:
            turnos_int += 1
            horas_rest = 0
        info['turnos_int'] = turnos_int
        info['horas_rest'] = horas_rest

    # Paginação
    bombeiros_ordenados = sorted(dados.items(), key=lambda item: item[0].nome)
    POR_PAGINA = 25
    paginas = []
    pagina_atual = []
    subtotal_turnos_int = 0
    subtotal_horas_rest = 0
    subtotal_valor = 0.0
    total_geral_turnos_int = 0
    total_geral_horas_rest = 0
    total_geral_valor = 0.0

    for i, (bombeiro, info) in enumerate(bombeiros_ordenados):
        pagina_atual.append((bombeiro, info))
        subtotal_turnos_int += info['turnos_int']
        subtotal_horas_rest += info['horas_rest']
        if subtotal_horas_rest >= 12:
            extra = int(subtotal_horas_rest // 12)
            subtotal_turnos_int += extra
            subtotal_horas_rest = round(subtotal_horas_rest % 12, 1)
        subtotal_valor += info['valor']

        total_geral_turnos_int += info['turnos_int']
        total_geral_horas_rest += info['horas_rest']
        if total_geral_horas_rest >= 12:
            extra_geral = int(total_geral_horas_rest // 12)
            total_geral_turnos_int += extra_geral
            total_geral_horas_rest = round(total_geral_horas_rest % 12, 1)
        total_geral_valor += info['valor']

        if len(pagina_atual) == POR_PAGINA or i == len(bombeiros_ordenados) - 1:
            paginas.append({
                'bombeiros': pagina_atual,
                'subtotal_turnos_int': subtotal_turnos_int,
                'subtotal_horas_rest': subtotal_horas_rest,
                'subtotal_valor': subtotal_valor
            })
            pagina_atual = []
            subtotal_turnos_int = 0
            subtotal_horas_rest = 0
            subtotal_valor = 0.0

    return render_template('imprimir_contabilidade_ecin.html',
                           mes=mes, ano=ano,
                           paginas=paginas,
                           total_geral_turnos_int=total_geral_turnos_int,
                           total_geral_horas_rest=total_geral_horas_rest,
                           total_geral_valor=total_geral_valor)



@app.route('/administrativo/imprimir-contabilidade-elac')
@login_required
def imprimir_contabilidade_elac():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Secretaria':
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('administrativo', tab='ecins'))

    mes = request.args.get('mes', type=int, default=date.today().month)
    ano = request.args.get('ano', type=int, default=date.today().year)

    registos = Ecin.query.filter(
        Ecin.categoria == 'ELAC',   # ← única diferença
        db.extract('month', Ecin.data) == mes,
        db.extract('year', Ecin.data) == ano
    ).order_by(Ecin.bombeiro_id, Ecin.data).all()

    # ... (mesmo código da função anterior, igual)
    from collections import defaultdict
    dados = defaultdict(lambda: {'valor': 0.0})

    for ec in registos:
        if ec.bombeiro and ec.valor:
            dados[ec.bombeiro]['valor'] += ec.valor

    for bombeiro, info in dados.items():
        horas_totais = info['valor'] / 3.5
        turnos_int = int(horas_totais // 12)
        horas_rest = round(horas_totais % 12, 1)
        if horas_rest > 11.9:
            turnos_int += 1
            horas_rest = 0
        info['turnos_int'] = turnos_int
        info['horas_rest'] = horas_rest

    bombeiros_ordenados = sorted(dados.items(), key=lambda item: item[0].nome)
    POR_PAGINA = 25
    paginas = []
    pagina_atual = []
    subtotal_turnos_int = 0
    subtotal_horas_rest = 0
    subtotal_valor = 0.0
    total_geral_turnos_int = 0
    total_geral_horas_rest = 0
    total_geral_valor = 0.0

    for i, (bombeiro, info) in enumerate(bombeiros_ordenados):
        pagina_atual.append((bombeiro, info))
        subtotal_turnos_int += info['turnos_int']
        subtotal_horas_rest += info['horas_rest']
        if subtotal_horas_rest >= 12:
            extra = int(subtotal_horas_rest // 12)
            subtotal_turnos_int += extra
            subtotal_horas_rest = round(subtotal_horas_rest % 12, 1)
        subtotal_valor += info['valor']

        total_geral_turnos_int += info['turnos_int']
        total_geral_horas_rest += info['horas_rest']
        if total_geral_horas_rest >= 12:
            extra_geral = int(total_geral_horas_rest // 12)
            total_geral_turnos_int += extra_geral
            total_geral_horas_rest = round(total_geral_horas_rest % 12, 1)
        total_geral_valor += info['valor']

        if len(pagina_atual) == POR_PAGINA or i == len(bombeiros_ordenados) - 1:
            paginas.append({
                'bombeiros': pagina_atual,
                'subtotal_turnos_int': subtotal_turnos_int,
                'subtotal_horas_rest': subtotal_horas_rest,
                'subtotal_valor': subtotal_valor
            })
            pagina_atual = []
            subtotal_turnos_int = 0
            subtotal_horas_rest = 0
            subtotal_valor = 0.0

    return render_template('imprimir_contabilidade_elac.html',
                           mes=mes, ano=ano,
                           paginas=paginas,
                           total_geral_turnos_int=total_geral_turnos_int,
                           total_geral_horas_rest=total_geral_horas_rest,
                           total_geral_valor=total_geral_valor)



@app.route('/correio/apagar/<int:id>')
@login_required
def correio_apagar(id):
    msg = MensagemCorreio.query.get_or_404(id)
    if msg.remetente_id == current_user.id:
        msg.apagada_remetente = True
    elif msg.destinatario_id == current_user.id or msg.departamento == current_user.resp_departamento:
        msg.apagada_destinatario = True
    else:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('correio'))
    db.session.commit()
    flash('Mensagem removida.', 'info')
    return redirect(url_for('correio'))

@app.route('/monitor')
@login_required
def monitor():
    config = Monitor.query.filter_by(bombeiro_id=current_user.id).first()
    if not config:
        config = Monitor(bombeiro_id=current_user.id)
        db.session.add(config)
        db.session.commit()
    return render_template('monitor.html', config=config)


@app.route('/monitor/config', methods=['GET', 'POST'])
@login_required
def monitor_config():
    if request.method == 'GET':
        config = Monitor.query.filter_by(bombeiro_id=current_user.id).first()
        if not config:
            config = Monitor(bombeiro_id=current_user.id)
            db.session.add(config)
            db.session.commit()
        return jsonify({
            'fogos': config.fogos,
            'google_maps': config.google_maps,
            'bombeiros_pt': config.bombeiros_pt,
            'ipma': config.ipma,
            'pontoagua': config.pontoagua if hasattr(config, 'pontoagua') else True
        })
    else:  # POST
        data = request.get_json()
        config = Monitor.query.filter_by(bombeiro_id=current_user.id).first()
        if not config:
            config = Monitor(bombeiro_id=current_user.id)
            db.session.add(config)

        config.fogos = data.get('fogos', True)
        config.google_maps = data.get('google_maps', True)
        config.bombeiros_pt = data.get('bombeiros_pt', True)
        config.ipma = data.get('ipma', True)
        config.pontoagua = data.get('pontoagua', True)  # ← corrigido: usar pontoagua (não pontos_agua)

        db.session.commit()
        return jsonify({'success': True})


from geopy.geocoders import Nominatim  # pip install geopy
from sqlalchemy import func

geolocator = Nominatim(user_agent="bombeiros_nisa")


@app.route('/api/pontos-agua', methods=['GET'])
@login_required
def get_pontos_agua():
    """Retorna todos os pontos de água em GeoJSON"""
    freguesia = request.args.get('freguesia', '')
    tipo = request.args.get('tipo', '')

    query = PontoAgua.query
    if freguesia:
        query = query.filter(PontoAgua.freguesia == freguesia)
    if tipo:
        query = query.filter(PontoAgua.tipo == tipo)

    pontos = query.all()

    features = []
    for p in pontos:
        features.append({
            'type': 'Feature',
            'geometry': {
                'type': 'Point',
                'coordinates': [p.longitude, p.latitude]
            },
            'properties': {
                'id': p.id,
                'nome': p.nome,
                'tipo': p.tipo,
                'freguesia': p.freguesia,
                'descricao': p.descricao,
                'capacidade': p.capacidade
            }
        })

    return jsonify({
        'type': 'FeatureCollection',
        'features': features
    })


@app.route('/api/pontos-agua', methods=['POST'])
@login_required
def add_ponto_agua():
    """Adiciona um novo ponto de água"""
    # Verificar permissões
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        return jsonify({'error': 'Acesso restrito'}), 403

    data = request.get_json()

    nome = data.get('nome', '').strip()
    tipo = data.get('tipo', 'Hidrante')
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    freguesia = data.get('freguesia', '').strip()  # ← campo freguesia
    descricao = data.get('descricao', '')
    capacidade = data.get('capacidade', '')

    # Validações
    if not nome:
        return jsonify({'error': 'Nome é obrigatório'}), 400
    if latitude is None or longitude is None:
        return jsonify({'error': 'Coordenadas inválidas'}), 400

    try:
        latitude = float(latitude)
        longitude = float(longitude)
    except (ValueError, TypeError):
        return jsonify({'error': 'Coordenadas inválidas'}), 400

    ponto = PontoAgua(
        nome=nome,
        tipo=tipo,
        latitude=latitude,
        longitude=longitude,
        freguesia=freguesia if freguesia else None,
        descricao=descricao,
        capacidade=capacidade,
        criado_por=current_user.id
    )

    db.session.add(ponto)
    db.session.commit()

    return jsonify({'success': True, 'id': ponto.id})


@app.route('/api/pontos-agua/<int:id>', methods=['DELETE'])
@login_required
def delete_ponto_agua(id):
    """Remove um ponto de água"""
    # Verificar permissões
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento not in ['Comando', 'Fardamento']:
        return jsonify({'error': 'Acesso restrito'}), 403

    ponto = PontoAgua.query.get_or_404(id)
    db.session.delete(ponto)
    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/concelhos', methods=['GET'])
@login_required
def get_concelhos():
    """Retorna lista de concelhos com pontos de água"""
    concelhos = db.session.query(PontoAgua.concelho).distinct().all()
    return jsonify([c[0] for c in concelhos if c[0]])



@app.route('/correio/apagar-em-massa', methods=['POST'])
@login_required
def apagar_correio_massa():
    ids = request.form.getlist('ids[]')
    print(f"DEBUG: IDs recebidos: {ids}")  # para ver nos logs do Render
    if not ids:
        flash('Nenhuma mensagem selecionada.', 'warning')
        return redirect(url_for('correio'))

    for msg_id in ids:
        try:
            msg = MensagemCorreio.query.get(int(msg_id))
            if msg:
                if msg.destinatario_id == current_user.id or msg.departamento == current_user.resp_departamento:
                    msg.apagada_destinatario = True
                elif msg.remetente_id == current_user.id:
                    msg.apagada_remetente = True
        except Exception as e:
            print(f"DEBUG: Erro ao apagar msg {msg_id}: {e}")

    db.session.commit()
    flash(f'{len(ids)} mensagens removidas com sucesso.', 'success')
    return redirect(url_for('correio'))


# ==================== QUADRO OPERACIONAL ====================

@app.route('/quadro-operacional')
@login_required
def quadro_operacional():
    # Verificar permissões: Admin, Comando ou bombeiros escalados para hoje
    hoje = date.today()
    hora_atual = datetime.now().hour

    # Verificar se o utilizador está escalado para hoje
    is_escalado_hoje = Escala.query.filter(
        Escala.bombeiro_id == current_user.id,
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje
    ).first() is not None

    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando' and not is_escalado_hoje:
        flash('Acesso restrito. Apenas Admin, Comando ou bombeiros escalados para hoje podem aceder.', 'danger')
        return redirect(url_for('dashboard'))

    # ========== 1. COMANDO ==========
    comando = Bombeiro.query.filter(
        Bombeiro.resp_departamento == 'Comando',
        Bombeiro.ativo == True
    ).first()

    # ========== 2. CENTRAL (depende do turno) ==========
    if 8 <= hora_atual < 20:
        turno_central = '8 - 08h/20h'
        turno_central_desc = "Turno Diurno (08h-20h)"
    else:
        turno_central = '9 - 20h/08h'
        turno_central_desc = "Turno Noturno (20h-08h)"

    central = Escala.query.filter(
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje,
        Escala.categoria == 'Centralista',
        Escala.turno == turno_central
    ).first()

    # ========== 3. ECIN (depende do turno atual) ==========
    if 7 <= hora_atual < 19:
        turno_ecin = '6 - 07h/19h'
        turno_ecin_desc = "ECIN - Turno Diurno (07h-19h)"
    else:
        turno_ecin = '7 - 19h/07h'
        turno_ecin_desc = "ECIN - Turno Noturno (19h-07h)"

    # Buscar ECINs para o turno atual
    ecins = Ecin.query.filter(
        Ecin.data == hoje,
        Ecin.categoria == 'ECIN',
        Ecin.turno == turno_ecin,
        Ecin.estado.in_(['Motorista ECIN', 'Chefe ECIN', 'Guarnição ECIN'])
    ).order_by(Ecin.funcao).all()

    # Se não encontrar ECINs no turno específico, buscar qualquer ECIN do dia
    if not ecins:
        ecins = Ecin.query.filter(
            Ecin.data == hoje,
            Ecin.categoria == 'ECIN',
            Ecin.estado.in_(['Motorista ECIN', 'Chefe ECIN', 'Guarnição ECIN'])
        ).order_by(Ecin.funcao).all()
        if ecins:
            turno_ecin_desc = "ECIN (Turno não especificado)"

    # Organizar ECIN por função
    ecin_chefe = None
    ecin_motorista = None
    ecin_guarnicao = []

    for ec in ecins:
        if ec.funcao == 'Chefe':
            ecin_chefe = ec
        elif ec.funcao == 'Motorista':
            ecin_motorista = ec
        elif ec.funcao == 'Guarnição' and len(ecin_guarnicao) < 3:
            ecin_guarnicao.append(ec)

    # ========== 4. EIP (5 bombeiros do turno) ==========
    eips = Escala.query.filter(
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje,
        Escala.categoria == 'EIP',
        Escala.turno == turno_ecin
    ).order_by(Escala.bombeiro_id).limit(5).all()

    if not eips:
        eips = Escala.query.filter(
            func.date(Escala.data_inicio) <= hoje,
            func.date(Escala.data_fim) >= hoje,
            Escala.categoria == 'EIP'
        ).order_by(Escala.bombeiro_id).limit(5).all()

    # ========== 5. INEM ==========
    if 7 <= hora_atual < 19:
        turno_inem = '6 - 07h/19h'
        turno_inem_desc = "Turno Diurno (07h-19h)"
    else:
        turno_inem = '7 - 19h/07h'
        turno_inem_desc = "Turno Noturno (19h-07h)"

    inem_um = Escala.query.filter(
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje,
        Escala.categoria == 'Socorrista',
        Escala.turno == turno_inem
    ).first()

    # ========== 6. MOTORISTAS DO TURNO PARA INEM ==========
    motoristas_turno = Escala.query.join(Bombeiro).filter(
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje,
        Escala.categoria == 'Motorista',
        Escala.turno == turno_inem
    ).order_by(Bombeiro.nome).all()

    if not motoristas_turno:
        motoristas_turno = Escala.query.join(Bombeiro).filter(
            func.date(Escala.data_inicio) <= hoje,
            func.date(Escala.data_fim) >= hoje,
            Escala.categoria == 'Motorista'
        ).order_by(Escala.turno, Bombeiro.nome).all()

    # ========== 7. TODOS OS EIP PARA AS COMBOBOX DA RESERVA ==========
    todos_eip = Escala.query.join(Bombeiro).filter(
        func.date(Escala.data_inicio) <= hoje,
        func.date(Escala.data_fim) >= hoje,
        Escala.categoria == 'EIP',
        Escala.turno == turno_ecin
    ).order_by(Bombeiro.nome).all()

    if not todos_eip:
        todos_eip = Escala.query.join(Bombeiro).filter(
            func.date(Escala.data_inicio) <= hoje,
            func.date(Escala.data_fim) >= hoje,
            Escala.categoria == 'EIP'
        ).order_by(Bombeiro.nome).all()

    # ========== 8. BUSCAR CONFIGURAÇÃO SALVA ==========
    config = QuadroOperacional.query.filter_by(data=hoje).first()

    # ========== 9. BUSCAR VIATURAS ==========
    viaturas_vfci = Viatura.query.filter(
        Viatura.tipo.ilike('%VFCI%'),
        Viatura.estado == 'operacional'
    ).order_by(Viatura.matricula).all()

    viaturas_absc = Viatura.query.filter(
        Viatura.tipo.ilike('%ABSC%'),
        Viatura.estado == 'operacional'
    ).order_by(Viatura.matricula).all()

    viaturas_vcot = Viatura.query.filter(
        Viatura.tipo.ilike('%VCOT%'),
        Viatura.estado == 'operacional'
    ).order_by(Viatura.matricula).all()

    # Determinar turno atual para exibição
    if 7 <= hora_atual < 19:
        turno_atual = "Turno Diurno (07h00 - 19h00)"
    else:
        turno_atual = "Turno Noturno (19h00 - 07h00)"

    return render_template('quadro_operacional.html',
                           hoje=hoje,
                           hora_atual=hora_atual,
                           turno_atual=turno_atual,
                           turno_central=turno_central_desc,
                           turno_ecin=turno_ecin_desc,
                           turno_inem=turno_inem_desc,
                           comando=comando,
                           central=central,
                           ecin_chefe=ecin_chefe,
                           ecin_motorista=ecin_motorista,
                           ecin_guarnicao=ecin_guarnicao,
                           eips=eips,
                           inem_um=inem_um,
                           motoristas_turno=motoristas_turno,
                           todos_eip=todos_eip,
                           viaturas_vfci=viaturas_vfci,
                           viaturas_absc=viaturas_absc,
                           viaturas_vcot=viaturas_vcot,
                           config=config)


@app.route('/api/salvar-quadro-operacional', methods=['POST'])
@login_required
def salvar_quadro_operacional():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        return jsonify({'error': 'Acesso restrito'}), 403

    data = request.get_json()
    hoje = date.today()

    # Verificar se já existe configuração para hoje
    quadro = QuadroOperacional.query.filter_by(data=hoje).first()
    if not quadro:
        quadro = QuadroOperacional(data=hoje, criado_por=current_user.id)
        db.session.add(quadro)

    # Atualizar viaturas
    quadro.viatura_ecin_id = data.get('viatura_ecin') or None
    quadro.viatura_eip_id = data.get('viatura_eip') or None
    quadro.viatura_inem_id = data.get('viatura_inem') or None
    quadro.viatura_reserva_id = data.get('viatura_reserva') or None
    quadro.viatura_comando_id = data.get('viatura_comando') or None

    # Atualizar motorista INEM
    quadro.motorista_inem_id = data.get('motorista_inem_id') or None
    quadro.motorista_inem_numero = data.get('motorista_inem_numero') or None
    quadro.motorista_inem_mec = data.get('motorista_inem_mec') or None

    # Atualizar reservas
    quadro.reserva_1_id = data.get('reserva_1_id') or None
    quadro.reserva_1_numero = data.get('reserva_1_numero') or None
    quadro.reserva_1_mec = data.get('reserva_1_mec') or None

    quadro.reserva_2_id = data.get('reserva_2_id') or None
    quadro.reserva_2_numero = data.get('reserva_2_numero') or None
    quadro.reserva_2_mec = data.get('reserva_2_mec') or None

    db.session.commit()

    return jsonify({'success': True})


@app.route('/api/exportar-quadro-operacional', methods=['POST'])
@login_required
def exportar_quadro_operacional():
    data = request.get_json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quadro Operacional"

    # Estilos
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True, size=12)
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_font = Font(bold=True, size=14, color='FFFFFF')

    # Título
    ws.merge_cells('A1:E1')
    cell = ws['A1']
    cell.value = f"QUADRO OPERACIONAL - {data['data']}"
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # Subtítulo
    ws.merge_cells('A2:E2')
    ws['A2'].value = f"Turno Atual: {data.get('turno_atual', '')}"
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')

    row = 4

    # ECIN
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1).value = "ECIN - Equipa de Combate a Incêndios"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "Função"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Nº Interno"
    ws.cell(row=row, column=4).value = "Mecanográfico"
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    ws.cell(row=row, column=1).value = "Chefe"
    ws.cell(row=row, column=2).value = data['ecin']['chefe']
    ws.cell(row=row, column=3).value = data['ecin']['chefe_numero']
    ws.cell(row=row, column=4).value = data['ecin']['chefe_mec']
    row += 1

    ws.cell(row=row, column=1).value = "Motorista"
    ws.cell(row=row, column=2).value = data['ecin']['motorista']
    ws.cell(row=row, column=3).value = data['ecin']['motorista_numero']
    ws.cell(row=row, column=4).value = data['ecin']['motorista_mec']
    row += 1

    for i, g in enumerate(data['ecin']['guarnicao'], 1):
        if g and g.get('nome'):
            ws.cell(row=row, column=1).value = f"Guarnição {i}"
            ws.cell(row=row, column=2).value = g.get('nome', '')
            ws.cell(row=row, column=3).value = g.get('numero', '')
            ws.cell(row=row, column=4).value = g.get('mec', '')
            row += 1

    ws.cell(row=row, column=1).value = "Viatura ECIN"
    ws.cell(row=row, column=2).value = data['ecin']['viatura']
    row += 2

    # EIP
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1).value = "EIP - Equipa de Intervenção Permanente"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='000000')
    row += 1

    ws.cell(row=row, column=1).value = "#"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Nº Interno"
    ws.cell(row=row, column=4).value = "Mecanográfico"
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    for i, e in enumerate(data['eip']['elementos'], 1):
        if e and e.get('nome'):
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = e.get('nome', '')
            ws.cell(row=row, column=3).value = e.get('numero', '')
            ws.cell(row=row, column=4).value = e.get('mec', '')
            row += 1

    ws.cell(row=row, column=1).value = "Viatura EIP"
    ws.cell(row=row, column=2).value = data['eip']['viatura']
    row += 2

    # INEM
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1).value = f"INEM - Equipa de Suporte (Turno: {data['inem']['turno']})"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "#"
    ws.cell(row=row, column=2).value = "Função/Nome"
    ws.cell(row=row, column=3).value = "Nº Interno"
    ws.cell(row=row, column=4).value = "Mecanográfico"
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    ws.cell(row=row, column=1).value = "1"
    ws.cell(row=row, column=2).value = f"Socorrista\n{data['inem']['socorrista']}"
    ws.cell(row=row, column=3).value = data['inem']['socorrista_numero']
    ws.cell(row=row, column=4).value = data['inem']['socorrista_mec']
    row += 1

    ws.cell(row=row, column=1).value = "2"
    ws.cell(row=row, column=2).value = f"Motorista\n{data['inem']['motorista']}"
    ws.cell(row=row, column=3).value = data['inem']['motorista_numero']
    ws.cell(row=row, column=4).value = data['inem']['motorista_mec']
    row += 1

    ws.cell(row=row, column=1).value = "Viatura INEM"
    ws.cell(row=row, column=2).value = data['inem']['viatura']
    row += 2

    # RESERVA
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1).value = "RESERVA (EIP)"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='0DCAF0', end_color='0DCAF0', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='000000')
    row += 1

    ws.cell(row=row, column=1).value = "#"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Nº Interno"
    ws.cell(row=row, column=4).value = "Mecanográfico"
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    for i, r in enumerate(data['reserva']['elementos'], 1):
        if r and r.get('nome'):
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = r.get('nome', '')
            ws.cell(row=row, column=3).value = r.get('numero', '')
            ws.cell(row=row, column=4).value = r.get('mec', '')
            row += 1

    ws.cell(row=row, column=1).value = "Viatura Reserva"
    ws.cell(row=row, column=2).value = data['reserva']['viatura']
    row += 2

    # COMANDO
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1).value = "COMANDO"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "Função"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Nº Interno"
    ws.cell(row=row, column=4).value = "Mecanográfico"
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    ws.cell(row=row, column=1).value = "Comandante"
    ws.cell(row=row, column=2).value = data['comando']['nome']
    ws.cell(row=row, column=3).value = data['comando']['numero']
    ws.cell(row=row, column=4).value = data['comando']['mec']
    row += 1

    ws.cell(row=row, column=1).value = "Viatura Comando"
    ws.cell(row=row, column=2).value = data['comando']['viatura']
    row += 2

    # CENTRAL
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1).value = f"CENTRAL (Turno: {data['central']['turno']})"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "Função"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Nº Interno"
    ws.cell(row=row, column=4).value = "Mecanográfico"
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    ws.cell(row=row, column=1).value = "Central"
    ws.cell(row=row, column=2).value = data['central']['nome']
    ws.cell(row=row, column=3).value = data['central']['numero']
    ws.cell(row=row, column=4).value = data['central']['mec']

    # Ajustar larguras
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 30

    # Ajustar altura das linhas
    for r in range(1, row + 5):
        ws.row_dimensions[r].height = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name=f'quadro_operacional_{data["data"]}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/quadro-operacional-config', methods=['GET'])
@login_required
def get_quadro_operacional_config():
    hoje = date.today()
    quadro = QuadroOperacional.query.filter_by(data=hoje).first()

    if quadro:
        return jsonify({
            'viatura_ecin': quadro.viatura_ecin_id,
            'viatura_eip': quadro.viatura_eip_id,
            'viatura_inem': quadro.viatura_inem_id,
            'viatura_reserva': quadro.viatura_reserva_id,
            'viatura_comando': quadro.viatura_comando_id,
            'motorista_inem_id': quadro.motorista_inem_id,
            'motorista_inem_numero': quadro.motorista_inem_numero,
            'motorista_inem_mec': quadro.motorista_inem_mec,
            'reserva_1_id': quadro.reserva_1_id,
            'reserva_1_numero': quadro.reserva_1_numero,
            'reserva_1_mec': quadro.reserva_1_mec,
            'reserva_2_id': quadro.reserva_2_id,
            'reserva_2_numero': quadro.reserva_2_numero,
            'reserva_2_mec': quadro.reserva_2_mec
        })
    return jsonify({})


@app.route('/api/salvar-quadro-operacional', methods=['POST'])
@login_required
def salvar_quadro_operacional():
    if current_user.tipo_user != 'Admin' and current_user.resp_departamento != 'Comando':
        return jsonify({'error': 'Acesso restrito'}), 403

    data = request.get_json()
    hoje = date.today()

    quadro = QuadroOperacional.query.filter_by(data=hoje).first()
    if not quadro:
        quadro = QuadroOperacional(data=hoje, criado_por=current_user.id)
        db.session.add(quadro)

    # Atualizar viaturas
    quadro.viatura_ecin_id = data.get('viatura_ecin') or None
    quadro.viatura_eip_id = data.get('viatura_eip') or None
    quadro.viatura_inem_id = data.get('viatura_inem') or None
    quadro.viatura_reserva_id = data.get('viatura_reserva') or None
    quadro.viatura_comando_id = data.get('viatura_comando') or None

    # Atualizar motorista INEM
    quadro.motorista_inem_id = data.get('motorista_inem_id') or None
    quadro.motorista_inem_numero = data.get('motorista_inem_numero') or None
    quadro.motorista_inem_mec = data.get('motorista_inem_mec') or None

    # Atualizar reservas
    quadro.reserva_1_id = data.get('reserva_1_id') or None


@app.route('/api/exportar-quadro-operacional', methods=['POST'])
@login_required
def exportar_quadro_operacional():
    data = request.get_json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Quadro Operacional"

    # Estilos
    header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    header_font = Font(bold=True, size=12)
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_font = Font(bold=True, size=14, color='FFFFFF')

    # Título
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = f"QUADRO OPERACIONAL - {data['data']}"
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # Subtítulo com turno
    ws.merge_cells('A2:D2')
    ws['A2'].value = f"Turno Atual: {data.get('turno_atual', '')}"
    ws['A2'].alignment = openpyxl.styles.Alignment(horizontal='center')

    row = 4

    # ECIN
    ws.cell(row=row, column=1).value = "ECIN - Equipa de Combate a Incêndios"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "Função"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Mecanográfico"
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    ws.cell(row=row, column=1).value = "Chefe"
    ws.cell(row=row, column=2).value = data['ecin']['chefe']
    row += 1
    ws.cell(row=row, column=1).value = "Motorista"
    ws.cell(row=row, column=2).value = data['ecin']['motorista']
    row += 1
    for i, g in enumerate(data['ecin']['guarnicao'], 1):
        if g:
            ws.cell(row=row, column=1).value = f"Guarnição {i}"
            ws.cell(row=row, column=2).value = g
            row += 1

    ws.cell(row=row, column=1).value = "Viatura ECIN"
    ws.cell(row=row, column=2).value = data['ecin']['viatura']
    row += 2

    # EIP
    ws.cell(row=row, column=1).value = "EIP - Equipa de Intervenção Permanente"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='000000')
    row += 1

    ws.cell(row=row, column=1).value = "#"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Mecanográfico"
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    for i, e in enumerate(data['eip']['elementos'], 1):
        if e:
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = e
            row += 1

    ws.cell(row=row, column=1).value = "Viatura EIP"
    ws.cell(row=row, column=2).value = data['eip']['viatura']
    row += 2

    # INEM
    ws.cell(row=row, column=1).value = f"INEM - Equipa de Suporte (Turno: {data['inem']['turno']})"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "#"
    ws.cell(row=row, column=2).value = "Função/Nome"
    ws.cell(row=row, column=3).value = "Mecanográfico"
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    ws.cell(row=row, column=1).value = "1"
    ws.cell(row=row, column=2).value = f"Socorrista ({data['inem']['turno']})\n{data['inem']['socorrista']}"
    row += 1
    ws.cell(row=row, column=1).value = "2"
    ws.cell(row=row, column=2).value = f"Motorista\n{data['inem']['motorista']}"
    row += 1

    ws.cell(row=row, column=1).value = "Viatura INEM"
    ws.cell(row=row, column=2).value = data['inem']['viatura']
    row += 2

    # RESERVA
    ws.cell(row=row, column=1).value = "RESERVA (EIP)"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='0DCAF0', end_color='0DCAF0', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='000000')
    row += 1

    ws.cell(row=row, column=1).value = "#"
    ws.cell(row=row, column=2).value = "Nome"
    ws.cell(row=row, column=3).value = "Mecanográfico"
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).font = header_font
    row += 1

    for i, r in enumerate(data['reserva']['elementos'], 1):
        if r:
            ws.cell(row=row, column=1).value = i
            ws.cell(row=row, column=2).value = r
            row += 1

    ws.cell(row=row, column=1).value = "Viatura Reserva"
    ws.cell(row=row, column=2).value = data['reserva']['viatura']
    row += 2

    # COMANDO
    ws.cell(row=row, column=1).value = "COMANDO"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "Comandante"
    ws.cell(row=row, column=2).value = data['comando']['nome']
    row += 1
    ws.cell(row=row, column=1).value = "Viatura Comando"
    ws.cell(row=row, column=2).value = data['comando']['viatura']
    row += 2

    # CENTRAL
    ws.cell(row=row, column=1).value = f"CENTRAL (Turno: {data['central']['turno']})"
    ws.cell(row=row, column=1).fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1

    ws.cell(row=row, column=1).value = "Central"
    ws.cell(row=row, column=2).value = data['central']['nome']

    # Ajustar larguras
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 30

    # Ajustar altura das linhas
    for r in range(1, row + 5):
        ws.row_dimensions[r].height = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True,
                     download_name=f'quadro_operacional_{data["data"]}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


#if __name__ == '__main__':
    #app.run(debug=True)